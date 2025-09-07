import os
import re
import io
from xlsxwriter import Workbook
from PyQt5 import QtCore
from PyQt5.QtCore import *
from PyQt5 import QtGui
from PyQt5.QtGui import *
from PyQt5.QtGui import QColor
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import*
from PyQt5.QtWidgets import QMessageBox
import sys
from PyQt5 import uic
import pandas as pd
import numpy as np
import scipy as sp
from scipy.interpolate import interp1d
import ctypes
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from datetime import date
from ui_windows.ui_newVersion import*
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib import style
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.ticker import FuncFormatter
from mpl_point_clicker import clicker
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Mm
from openpyxl.styles import PatternFill, Border, Side
import time
import win32com.client as win32

myappid = 'mycompany.myproduct.subproduct.version' # arbitrary string
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)

stylesheet2=""" #centralwidget 
{ background-image:url(Images for Program/png_worldwide.png)
} 
"""
desabledButtonStyleSheet=""" QPushButton{
font: 12pt "Siemens Sans Black";
color: rgb(255, 255, 255);
background-color: rgb(138, 138, 138);
border-radius:5px;
}
QPushButton:hover{
	
	background-color: rgb(95, 107, 112);
}
"""
enabledButtonStyleSheet=""" QPushButton{
font: 12pt "Siemens Sans Black";
color: rgb(255, 255, 255);
background-color: rgb(56, 63, 66);
border-radius:5px;
}
QPushButton:hover{
	
	background-color: rgb(255, 205, 0);
}
QPushButton:pressed{
background-color:rgb(250, 171, 72)
}
"""

class CustomMessageBox(QMessageBox):
    def closeEvent(self, event):
        if self.clickedButton() not in (self.button(QMessageBox.Cancel),):
            event.ignore()


class WorkerSignals(QObject):
    finished = pyqtSignal()
    result = pyqtSignal(object)
    progress = pyqtSignal(int)

class Worker(QRunnable):
    """Worker thread

    Inherits from QRunnable to handler worker thread setup, signals and wrap-"""
    def __init__(self, fn,*args, **kwargs):
        super(Worker, self).__init__()
        self.fn=fn
        self.args=args
        self.kwargs=kwargs
        self.signals = WorkerSignals()

        self.kwargs['progress_bar'] = self.signals.progress
    
    @pyqtSlot()
    def run(self):
        """Initialise the runner function with passed args, kwargs."""
        result=self.fn(*self.args, **self.kwargs)
        self.signals.result.emit(result)
        self.signals.finished.emit()

class Calculation_window(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi("ui_windows/all_calculation_window.ui",self)
        #self.ExportButton1.clicked.connect(self.Export_to_Excel)
        self.setStyleSheet(stylesheet2)
        self.setWindowTitle('all calculation table')
        self.setWindowIcon(QtGui.QIcon("Images for Program/logo.ico"))
        self.actionExcel_xlsx.triggered.connect(self.Export_to_Excel)

        self.results_table_1.installEventFilter(self)
        self.results_table_2.installEventFilter(self)
        self.results_table_3.installEventFilter(self)

    def eventFilter(self,source,event):
        ## _____ if the cursor is on the 1st table (results_table_1)
        if source==self.results_table_1:
            selected_table=self.results_table_1
            if event.type()== QtCore.QEvent.KeyPress:
                if event==QtGui.QKeySequence.Copy:
                    self.copySelection1()
                    return True
                elif event==QtGui.QKeySequence.Paste:
                    self.pasteSelection1()
                    return True
            elif event.type() == QtCore.QEvent.ContextMenu:
                # a context menu for the copy/paste operations
                menu = QtWidgets.QMenu()
                removeRow=menu.addAction('Delete selected row')
                removeRow.triggered.connect(self.removeRow1)

                removeColumn=menu.addAction('Delete selected column')
                removeColumn.triggered.connect(self.removeColumn1)
                
                copyAction = menu.addAction('Copy')
                copyAction.triggered.connect(self.copySelection1)
                
                pasteAction = menu.addAction('Paste')
                pasteAction.triggered.connect(self.pasteSelection1)

                addRowSubMenu=menu.addMenu('Add Row')
                addBefore=addRowSubMenu.addAction('add before')
                addBefore.triggered.connect(self.addRowBefore1)
                addAfter=addRowSubMenu.addAction('add after')
                addAfter.triggered.connect(self.addRowAfter1)

                addColumnSubMenu=menu.addMenu('Add Column')
                addColumnBefore=addColumnSubMenu.addAction('Add before')
                addColumnBefore.triggered.connect(self.addColumnBefore1)
                addColumnAfter=addColumnSubMenu.addAction('Add after')
                addColumnAfter.triggered.connect(self.addColumnAfter1)
                
                self.LineEdit1=QLineEdit('Rows number before',self)
                widget1=QWidgetAction(self)
                widget1.setDefaultWidget(self.LineEdit1)

                self.LineEdit2=QLineEdit('Rows number After',self)
                widget2=QWidgetAction(self)
                widget2.setDefaultWidget(self.LineEdit2)

                self.LineEdit3=QLineEdit('Column number before',self)
                widget3=QWidgetAction(self)
                widget3.setDefaultWidget(self.LineEdit3)

                self.LineEdit4=QLineEdit('Column number after')
                widget4=QWidgetAction(self)
                widget4.setDefaultWidget(self.LineEdit4)

                addMultipleRowsMenu=menu.addMenu('Add Rows')
                addMultipleRowsBefore=addMultipleRowsMenu.addAction(widget1)
                self.LineEdit1.returnPressed.connect(self.addMultipleRowBefore1)
                addMultipleRowsAfter=addMultipleRowsMenu.addAction(widget2)
                self.LineEdit2.returnPressed.connect(self.addMultipleRowAfter1)

                addMultipleColumnsMenu=menu.addMenu('Add Columns')
                addMultipleColumnsBefore=addMultipleColumnsMenu.addAction(widget3)
                self.LineEdit3.returnPressed.connect(self.addMultipleColumnsBefore1)
                addMultipleColumnsAfter=addMultipleColumnsMenu.addAction(widget4)
                self.LineEdit4.returnPressed.connect(self.addMultipleColumnsAfter1)
                
                if not selected_table.selectedIndexes() :
                    # no selection available, both copy and paste are disabled
                    copyAction.setEnabled(False)
                    pasteAction.setEnabled(False)
                    removeRow.setEnabled(False)
                    removeColumn.setEnabled(False)
                    addRowSubMenu.setEnabled(False)
                    addMultipleRowsMenu.setEnabled(False)
                menu.exec(event.globalPos())
                return True

## _____ if the cursor is on the 2nd table (results_table_2) 
        elif source==self.results_table_2:
            selected_table=self.results_table_2
            if event.type()== QtCore.QEvent.KeyPress:
                if event==QtGui.QKeySequence.Copy:
                    self.copySelection2()
                    return True
                elif event==QtGui.QKeySequence.Paste:
                    self.pasteSelection2()
                    return True
            elif event.type() == QtCore.QEvent.ContextMenu:
                # a context menu for the copy/paste operations
                menu = QtWidgets.QMenu()
                removeRow=menu.addAction('Delete selected row')
                removeRow.triggered.connect(self.removeRow2)

                removeColumn=menu.addAction('Delete selected column')
                removeColumn.triggered.connect(self.removeColumn2)
                
                copyAction = menu.addAction('Copy')
                copyAction.triggered.connect(self.copySelection2)
                
                pasteAction = menu.addAction('Paste')
                pasteAction.triggered.connect(self.pasteSelection2)

                addRowSubMenu=menu.addMenu('Add Row')
                addBefore=addRowSubMenu.addAction('add before')
                addBefore.triggered.connect(self.addRowBefore2)
                
                addAfter=addRowSubMenu.addAction('add after')
                addAfter.triggered.connect(self.addRowAfter2)

                addColumnSubMenu=menu.addMenu('Add column')
                addColumnBefore=addColumnSubMenu.addAction('Add column before')
                addColumnBefore.triggered.connect(self.addColumnBefore2)
                addColumnAfter=addColumnSubMenu.addAction('Add column after')
                addColumnAfter.triggered.connect(self.addColumnAfter2)
    
                self.LineEdit1=QLineEdit('Rows number before',self)
                widget1=QWidgetAction(self)
                widget1.setDefaultWidget(self.LineEdit1)

                self.LineEdit2=QLineEdit('Rows number After',self)
                widget2=QWidgetAction(self)
                widget2.setDefaultWidget(self.LineEdit2)

                self.LineEdit3=QLineEdit('Columns number before')
                widget3=QWidgetAction(self)
                widget3.setDefaultWidget(self.LineEdit3)

                self.LineEdit4=QLineEdit('Columns number after')
                widget4=QWidgetAction(self)
                widget4.setDefaultWidget(self.LineEdit4)

                addMultipleRowsMenu=menu.addMenu('Add Rows')
                addMultipleRowsBefore=addMultipleRowsMenu.addAction(widget1)
                self.LineEdit1.returnPressed.connect(self.addMultipleRowBefore2)
                addMultipleRowsAfter=addMultipleRowsMenu.addAction(widget2)
                self.LineEdit2.returnPressed.connect(self.addMultipleRowAfter2)

                addMultipleColumnsMenu=menu.addMenu('Add columns')
                addMultipleColumnsBefore=addMultipleColumnsMenu.addAction(widget3)
                self.LineEdit3.returnPressed.connect(self.addMultipleColumnsBefore2)
                addMultipleColumnsAfter=addMultipleColumnsMenu.addAction(widget4)
                self.LineEdit4.returnPressed.connect(self.addMultipleColumnsAfter2)

                if not selected_table.selectedIndexes() :
                    # no selection available, both copy and paste are disabled
                    copyAction.setEnabled(False)
                    pasteAction.setEnabled(False)
                    removeRow.setEnabled(False)
                    removeColumn.setEnabled(False)
                    addRowSubMenu.setEnabled(False)
                    addMultipleRowsMenu.setEnabled(False)
                menu.exec(event.globalPos())
                return True

## _____ if the cursor is on the 3rd table (results_table_3)
        elif source==self.results_table_3:
            selected_table=self.results_table_3
            if event.type()== QtCore.QEvent.KeyPress:
                if event==QtGui.QKeySequence.Copy:
                    self.copySelection3()
                    return True
                elif event==QtGui.QKeySequence.Paste:
                    self.pasteSelection3()
                    return True
            elif event.type() == QtCore.QEvent.ContextMenu:
                # a context menu for the copy/paste operations
                menu = QtWidgets.QMenu()
                removeRow=menu.addAction('Delete selected row')
                removeRow.triggered.connect(self.removeRow3)

                removeColumn=menu.addAction('Delete selected column')
                removeColumn.triggered.connect(self.removeColumn3)
                
                copyAction = menu.addAction('Copy')
                copyAction.triggered.connect(self.copySelection3)
                
                pasteAction = menu.addAction('Paste')
                pasteAction.triggered.connect(self.pasteSelection3)

                addRowSubMenu=menu.addMenu('Add Row')
                addBefore=addRowSubMenu.addAction('add before')
                addBefore.triggered.connect(self.addRowBefore3)
                addAfter=addRowSubMenu.addAction('add after')
                addAfter.triggered.connect(self.addRowAfter3)

                addColumnSubMenu=menu.addMenu('Add column')
                addColumnBefore=addColumnSubMenu.addAction('Add before')
                addColumnBefore.triggered.connect(self.addColumnBefore3)
                addColumnAfter=addColumnSubMenu.addAction('Add after')
                addColumnAfter.triggered.connect(self.addColumnAfter3)

                self.LineEdit1=QLineEdit('Rows number before',self)
                widget1=QWidgetAction(self)
                widget1.setDefaultWidget(self.LineEdit1)

                self.LineEdit2=QLineEdit('Rows number after',self)
                widget2=QWidgetAction(self)
                widget2.setDefaultWidget(self.LineEdit2)

                self.LineEdit3=QLineEdit('Columns number before')
                widget3=QWidgetAction(self)
                widget3.setDefaultWidget(self.LineEdit3)

                self.LineEdit4=QLineEdit('Columns number after')
                widget4=QWidgetAction(self)
                widget4.setDefaultWidget(self.LineEdit4)

                addMultipleRowsMenu=menu.addMenu('Add Rows')
                addMultipleRowsBefore=addMultipleRowsMenu.addAction(widget1)
                self.LineEdit1.returnPressed.connect(self.addMultipleRowBefore3)
                addMultipleRowsAfter=addMultipleRowsMenu.addAction(widget2)
                self.LineEdit2.returnPressed.connect(self.addMultipleRowAfter3)

                addMultipleColumnsMenu=menu.addMenu('Add columns')
                addMultipleColumnsBefore=addMultipleColumnsMenu.addAction(widget3)
                self.LineEdit3.returnPressed.connect(self.addMultipleColumnsBefore3)
                addMultipleColumnsAfter=addMultipleColumnsMenu.addAction(widget4)
                self.LineEdit4.returnPressed.connect(self.addMultipleColumnsAfter3)

                if not selected_table.selectedIndexes() :
                    # no selection available, both copy and paste are disabled
                    copyAction.setEnabled(False)
                    pasteAction.setEnabled(False)
                    removeRow.setEnabled(False)
                    removeColumn.setEnabled(False)
                    addRowSubMenu.setEnabled(False)
                    addMultipleRowsMenu.setEnabled(False)
                menu.exec(event.globalPos())
                return True
        
        return super(Calculation_window, self).eventFilter(source, event)
     
### ______________________ event_Functions for results_table_1  __________________________ ###

    def copySelection1(self):
        # clear the current contents of the clipboard
        self.clipboard=''
        copied_cells = sorted(self.results_table_1.selectedIndexes())
        max_column = copied_cells[-1].column()
        for c in copied_cells:
            self.clipboard += self.results_table_1.item(c.row(), c.column()).text()
            if c.column() == max_column:
                self.clipboard += '\n'
            else:
                self.clipboard += '\t'
        QApplication.clipboard().setText(self.clipboard)

    def pasteSelection1(self):
        selection = self.results_table_1.selectedIndexes()

        if selection:
            row_anchor = selection[0].row()
            column_anchor = selection[0].column()

            clipboard = QApplication.clipboard()

            rows = clipboard.text().split('\n')
            for indx_row, row in enumerate(rows):
                values = row.split('\t')
                for indx_col, value in enumerate(values):
                    item = QTableWidgetItem(value)
                    self.results_table_1.setItem(row_anchor + indx_row, column_anchor + indx_col, item)

    def removeColumn1(self):
        index_list=[]
        for modelidx in self.results_table_1.selectionModel().selectedColumns():
            idx=QtCore.QPersistentModelIndex(modelidx)
            index_list.append(idx)
        for index in index_list:
            self.results_table_1.removeColumn(index.column())

    def removeRow1(self):
        index_list=[]
        for modelidx in self.results_table_1.selectionModel().selectedRows():
            idx=QtCore.QPersistentModelIndex(modelidx)
            index_list.append(idx)
        for index in index_list:
            self.results_table_1.removeRow(index.row())

    def addRowBefore1(self):
        selection=self.results_table_1.selectedIndexes()
        rowPosition=selection[0].row()
        #print(rowPosition)
        self.results_table_1.insertRow(rowPosition)

    def addRowAfter1(self):
        selection=self.results_table_1.selectedIndexes()
        rowPosition=selection[0].row()
        #print(rowPosition)
        self.results_table_1.QAbstractItemModel.insertRow(rowPosition+1)

    def addColumnBefore1(self):
        selection=self.results_table_1.selectedIndexes()
        columnPosition=selection[0].column()
        self.results_table_1.insertColumn(columnPosition)

    def addColumnAfter1(self):
        selection=self.results_table_1.selectedIndexes()
        columnPosition=selection[0].column()
        self.results_table_1.insertColumn(columnPosition+1)

    def addMultipleRowBefore1(self):
        selection=self.results_table_1.selectedIndexes()
        rowPosition=selection[0].row()
        #print(rowPosition)
        for r in range(int(self.LineEdit1.text())):
            self.results_table_1.insertRow(rowPosition)

    def addMultipleRowAfter1(self):
        selection=self.results_table_1.selectedIndexes()
        rowPosition=selection[0].row()
        #print(rowPosition)
        for r in range(int(self.LineEdit2.text())):
            self.results_table_1.insertRow(rowPosition+1)

    def addMultipleColumnsBefore1(self):
        selection=self.results_table_1.selectedIndexes()
        columnPosition=selection[0].column()
        for c in range(int(self.LineEdit3.text())):
            self.results_table_1.insertColumn(columnPosition)

    def addMultipleColumnsAfter1(self):
        selection=self.results_table_1.selectedIndexes()
        columnPosition=selection[0].column()
        for c in range(int(self.LineEdit4.text())):
            self.results_table_1.insertColumn(columnPosition+1)

### ____________________________ event_Functions for results_table_2  ___________________________ ###

    def copySelection2(self):
        # clear the current contents of the clipboard
        self.clipboard=''
        copied_cells = sorted(self.results_table_2.selectedIndexes())
        #print(copied_cells)
        max_column = copied_cells[-1].column()
        for c in copied_cells:
            self.clipboard += self.results_table_2.item(c.row(), c.column()).text()
            if c.column() == max_column:
                self.clipboard += '\n'
            else:
                self.clipboard += '\t'
        QApplication.clipboard().setText(self.clipboard)

    def pasteSelection2(self):
        selection = self.results_table_2.selectedIndexes()

        if selection:
            row_anchor = selection[0].row()
            column_anchor = selection[0].column()

            clipboard = QApplication.clipboard()

            rows = clipboard.text().split('\n')
            for indx_row, row in enumerate(rows):
                values = row.split('\t')
                for indx_col, value in enumerate(values):
                    item = QTableWidgetItem(value)
                    self.results_table_2.setItem(row_anchor + indx_row, column_anchor + indx_col, item)

    def removeColumn2(self):
       index_list=[]
       for modelidx in self.results_table_2.selectionModel().selectedColumns():
           idx=QtCore.QPersistentModelIndex(modelidx)
           index_list.append(idx)
       for index in index_list:
           self.results_table_2.removeColumn(index.column())

    def removeRow2(self):
        index_list=[]
        for modelidx in self.results_table_2.selectionModel().selectedRows():
            idx=QtCore.QPersistentModelIndex(modelidx)
            index_list.append(idx)
        for index in index_list:
            self.results_table_2.removeRow(index.row())

    def addRowBefore2(self):
        selection=self.results_table_2.selectedIndexes()
        rowPosition=selection[0].row()
        #print(rowPosition)
        self.results_table_2.insertRow(rowPosition)

    def addRowAfter2(self):
        selection=self.results_table_2.selectedIndexes()
        rowPosition=selection[0].row()
        #print(rowPosition)
        self.results_table_2.insertRow(rowPosition+1)

    def addColumnBefore2(self):
        selection=self.results_table_2.selectedIndexes()
        columnPosition=selection[0].column()
        self.results_table_2.insertColumn(columnPosition)

    def addColumnAfter2(self):
        selection=self.results_table_2.selectedIndexes()
        columnPosition=selection[0].column()
        self.results_table_2.insertColumn(columnPosition+1)

    def addMultipleRowBefore2(self):
        selection=self.results_table_2.selectedIndexes()
        rowPosition=selection[0].row()
        #print(rowPosition)
        for r in range(int(self.LineEdit1.text())):
            self.results_table_2.insertRow(rowPosition)

    def addMultipleRowAfter2(self):
        selection=self.results_table_2.selectedIndexes()
        rowPosition=selection[0].row()
        #print(rowPosition)
        for r in range(int(self.LineEdit2.text())):
            self.results_table_2.insertRow(rowPosition+1)

    def addMultipleColumnsBefore2(self):
        selection=self.results_table_2.selectedindexes()
        columnPosition=selection[0].column()
        for c in range(int(self.LineEdit3.text())):
            self.results_table_2.insertColumn(columnPosition)

    def addMultipleColumnsAfter2(self):
        selection=self.results_table_2.selectedIndexes()
        columnPosition=selection[0].column()
        for c in range(int(self.LineEdit4.text())):
            self.results_table_2.insertColumn(columnPosition+1)

### __________________________________ event_Functions for results_table_3  __________________________ ###
    def copySelection3(self):
        # clear the current contents of the clipboard
        self.clipboard=''
        copied_cells = sorted(self.results_table_3.selectedIndexes())
        #print(copied_cells)
        max_column = copied_cells[-1].column()
        for c in copied_cells:
            self.clipboard += self.results_table_3.item(c.row(), c.column()).text()
            if c.column() == max_column:
                self.clipboard += '\n'
            else:
                self.clipboard += '\t'
        QApplication.clipboard().setText(self.clipboard)

    def pasteSelection3(self):
        selection = self.results_table_3.selectedIndexes()

        if selection:
            row_anchor = selection[0].row()
            column_anchor = selection[0].column()

            clipboard = QApplication.clipboard()

            rows = clipboard.text().split('\n')
            for indx_row, row in enumerate(rows):
                values = row.split('\t')
                for indx_col, value in enumerate(values):
                    item = QTableWidgetItem(value)
                    self.results_table_3.setItem(row_anchor + indx_row, column_anchor + indx_col, item)

    def removeColumn3(self):
        index_list=[]
        for modelidx in self.results_table_3.selectionModel().selectedColumns():
            idx=QtCore.QPersistentModelIndex(modelidx)
            index_list.append(idx)
        for index in index_list:
            self.results_table_3.removeColumn(index.column())

    def removeRow3(self):
        index_list=[]
        for modelidx in self.results_table_3.selectionModel().selectedRows():
            idx=QtCore.QPersistentModelIndex(modelidx)
            index_list.append(idx)
        for index in index_list:
            self.results_table_3.removeRow(index.row())

    def addRowBefore3(self):
        selection=self.results_table_3.selectedIndexes()
        rowPosition=selection[0].row()
        #print(rowPosition)
        self.results_table_3.insertRow(rowPosition)

    def addRowAfter3(self):
        selection=self.results_table_3.selectedIndexes()
        rowPosition=selection[0].row()
        #print(rowPosition)
        self.results_table_3.insertRow(rowPosition+1)

    def addColumnBefore3(self):
        selection=self.results_table_3.selectedIndexes()
        columnPosition=selection[0].column()
        self.results_table_3.insertColumn(columnPosition)

    def addColumnAfter3(self):
        selection=self.results_table_3.selectedIndexes()
        columnPosition=selection[0].column()
        self.results_table_3.insertColumn(columnPosition+1)

    def addMultipleRowBefore3(self):
        selection=self.results_table_3.selectedIndexes()
        rowPosition=selection[0].row()
        #print(rowPosition)
        for r in range(int(self.LineEdit1.text())):
            self.results_table_3.insertRow(rowPosition)

    def addMultipleRowAfter3(self):
        selection=self.results_table_3.selectedIndexes()
        rowPosition=selection[0].row()
        #print(rowPosition)
        for r in range(int(self.LineEdit2.text())):
            self.results_table_3.insertRow(rowPosition+1)

    def addMultipleColumnsBefore3(self):
        selection=self.results_table_3.selectedIndexes()
        columnPosition=selection[0].column()
        for c in range(int(self.LineEdit3.text())):
            self.results_table_3.insertColumn(columnPosition)

    def addMultipleColumnsAfter3(self):
        selection=self.results_table_3.selectedIndexes()
        columnPosition=selection[0].column()
        for c in range(int(self.LineEdit4.text())):
            self.results_table_3.insertColumn(columnPosition+1)

    def Export_to_Excel(self):

        #self.results_Specific=Calculation_window_Specific()
        # create column header and data frame:
        columnHeaders=[]
        for j in range(self.results_table_1.model().columnCount()):
            columnHeaders.append(self.results_table_1.horizontalHeaderItem(j).text())
            
        self.df=pd.DataFrame(columns=columnHeaders)

        # inserting data in the Data Frame:

        for row3 in range(self.results_table_3.rowCount()):
            for col3 in range(self.results_table_3.columnCount()):
                self.df.at[row3,columnHeaders[col3]]= self.results_table_3.item(row3,col3).text()

        for row1 in range(self.results_table_1.rowCount()):
            for col in range(self.results_table_1.columnCount()):
                row1s=row1+self.results_table_3.rowCount()
                self.df.at[row1s,columnHeaders[col]]= self.results_table_1.item(row1,col).text()

        for row2 in range(self.results_table_2.rowCount()):
            for col2 in range(self.results_table_2.columnCount()):
                row2s=row2+self.results_table_1.rowCount()+self.results_table_3.rowCount()
                self.df.at[row2s,columnHeaders[col2]]= self.results_table_2.item(row2,col2).text()
        
        response,filter=QFileDialog.getSaveFileName(self,"Bitte die Vorlage speichern",f'Motor-Daten-{self.results_table_3.item(0,3).text()}')
        if response!='':
            self.df.to_excel(rf'{response}.xlsx',index=False)
        else:
            pass

class Calculation_window_Specific(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi("ui_windows/specific_calculation_window.ui",self)
        self.setStyleSheet(stylesheet2)
        self.setWindowTitle('specific calculation table')
        self.setWindowIcon(QtGui.QIcon("Images for Program/logo.ico"))
        self.actionExcel_xlsx.triggered.connect(self.Export_to_Excel2) 

        self.results_table_Specific.installEventFilter(self)     
    
    def eventFilter(self,source,event):
        if event.type()== QtCore.QEvent.KeyPress:
            if event==QtGui.QKeySequence.Copy:
                self.copySelection()
                return True
            elif event==QtGui.QKeySequence.Paste:
                self.pasteSelection()
                return True
        elif event.type() == QtCore.QEvent.ContextMenu:
            # a context menu for the copy/paste operations
            menu = QtWidgets.QMenu()
            removeRow=menu.addAction('Delete selected row')
            removeRow.triggered.connect(self.removeRow)

            removeColumn=menu.addAction('Delete selected column')
            removeColumn.triggered.connect(self.removeColumn)
            
            copyAction = menu.addAction('Copy')
            copyAction.triggered.connect(self.copySelection)
            
            pasteAction = menu.addAction('Paste')
            pasteAction.triggered.connect(self.pasteSelection)

            addRowSubMenu=menu.addMenu('Add Row')
            addBefore=addRowSubMenu.addAction('add before')
            addBefore.triggered.connect(self.addRowBefore)
            addAfter=addRowSubMenu.addAction('add after')
            addAfter.triggered.connect(self.addRowAfter)

            addColumnSubMenu=menu.addMenu('Add Column')
            addColumnBefore=addColumnSubMenu.addAction('Add before')
            addColumnBefore.triggered.connect(self.addColumnBefore)
            addColumnAfter=addColumnSubMenu.addAction('Add after')
            addColumnAfter.triggered.connect(self.addColumnAfter)

            self.LineEdit1=QLineEdit('Rows number before',self)
            widget1=QWidgetAction(self)
            widget1.setDefaultWidget(self.LineEdit1)

            self.LineEdit2=QLineEdit('Rows number After',self)
            widget2=QWidgetAction(self)
            widget2.setDefaultWidget(self.LineEdit2)

            self.LineEdit3=QLineEdit('Columns number before')
            widget3=QWidgetAction(self)
            widget3.setDefaultWidget(self.LineEdit3)

            self.LineEdit4=QLineEdit('Columns number after')
            widget4=QWidgetAction(self)
            widget4.setDefaultWidget(self.LineEdit4)

            addMultipleRowsMenu=menu.addMenu('Add Rows')
            addMultipleRowsBefore=addMultipleRowsMenu.addAction(widget1)
            self.LineEdit1.returnPressed.connect(self.addMultipleRowBefore)
            addMultipleRowsAfter=addMultipleRowsMenu.addAction(widget2)
            self.LineEdit2.returnPressed.connect(self.addMultipleRowAfter)

            addmultiplecolumnsMenu=menu.addMenu('Add Columns')
            addMultipleColumnsBefore=addmultiplecolumnsMenu.addAction(widget3)
            self.LineEdit3.returnPressed.connect(self.addMultipleColumnsBefore)
            addMultiplecolumnsAfter=addmultiplecolumnsMenu.addAction(widget4)
            self.LineEdit4.returnPressed.connect(self.addMultipleColumnsAfter)

            if not self.results_table_Specific.selectedIndexes():
                # no selection available, both copy and paste are disabled
                copyAction.setEnabled(False)
                pasteAction.setEnabled(False)
                removeRow.setEnabled(False)
                removeColumn.setEnabled(False)
                addRowSubMenu.setEnabled(False)
                addMultipleRowsMenu.setEnabled(False)
            menu.exec(event.globalPos())
            return True
        return super(Calculation_window_Specific, self).eventFilter(source, event)
    
    def copySelection(self):
        # clear the current contents of the clipboard
        self.clipboard=''
        copied_cells = sorted(self.results_table_Specific.selectedIndexes())
        #print(copied_cells)
        max_column = copied_cells[-1].column()
        for c in copied_cells:
            self.clipboard += self.results_table_Specific.item(c.row(), c.column()).text()
            if c.column() == max_column:
                self.clipboard += '\n'
            else:
                self.clipboard += '\t'
        QApplication.clipboard().setText(self.clipboard)

    def pasteSelection(self):
        selection = self.results_table_Specific.selectedIndexes()

        if selection:
            row_anchor = selection[0].row()
            column_anchor = selection[0].column()

            clipboard = QApplication.clipboard()

            rows = clipboard.text().split('\n')
            for indx_row, row in enumerate(rows):
                values = row.split('\t')
                for indx_col, value in enumerate(values):
                    item = QTableWidgetItem(value)
                    self.results_table_Specific.setItem(row_anchor + indx_row, column_anchor + indx_col, item)

    def removeColumn(self):
        selection=sorted(self.results_table_Specific.selectionModel().selectedColumns())
        for idx in selection:
            self.results_table_Specific.removeColumn(idx.column())

    def removeRow(self):
        selection=sorted(self.results_table_Specific.selectionModel().selectedRows())
        for idx in selection:
            self.results_table_Specific.removeRow(idx.row())

    def addRowBefore(self):
        selection=self.results_table_Specific.selectedIndexes()
        rowPosition=selection[0].row()
        #print(rowPosition)
        self.results_table_Specific.insertRow(rowPosition)

    def addRowAfter(self):
        selection=self.results_table_Specific.selectedIndexes()
        rowPosition=selection[0].row()
        #print(rowPosition)
        self.results_table_Specific.insertRow(rowPosition+1)
    
    def addColumnBefore(self):
        selection=self.results_table_Specific.selectedIndexes()
        columnPosition=selection[0].column()
        self.results_table_Specific.insertColumn(columnPosition)

    def addColumnAfter(self):
        selection=self.results_table_Specific.selectedIndexes()
        columnPosition=selection[0].column()
        self.results_table_Specific.insertColumn(columnPosition+1)

    def addMultipleRowBefore(self):
        selection=self.results_table_Specific.selectedIndexes()
        rowPosition=selection[0].row()
        #print(rowPosition)
        for r in range(int(self.LineEdit1.text())):
            self.results_table_Specific.insertRow(rowPosition)

    def addMultipleRowAfter(self):
        selection=self.results_table_Specific.selectedIndexes()
        rowPosition=selection[0].row()
        #print(rowPosition)
        for r in range(int(self.LineEdit2.text())):
            self.results_table_Specific.insertRow(rowPosition+1)

    def addMultipleColumnsBefore(self):
        selection=self.results_table_Specific.selectedIndexes()
        ColumnPosition=selection[0].column()
        for c in range(int(self.LineEdit3.text())):
            self.results_table_Specific.insertColumn(ColumnPosition)

    def addMultipleColumnsAfter(self):
        selection=self.results_table_Specific.selectedIndexes()
        columnPosition=selection[0].column()
        for c in range(int(self.LineEdit4.text())):
            self.results_table_Specific.insertColumn(columnPosition+1)

    def Export_to_Excel2(self):
        # create column header and data frame:
        columnHeaders=[]
        for j in range(self.results_table_Specific.model().columnCount()):
            columnHeaders.append(self.results_table_Specific.horizontalHeaderItem(j).text())

        self.df=pd.DataFrame(columns=columnHeaders)
        # inserting data in the Data Frame:
        for row in range(self.results_table_Specific.rowCount()):
            for col in range(self.results_table_Specific.columnCount()):
                self.df.at[row,columnHeaders[col]]= self.results_table_Specific.item(row,col).text()
        
        response,filter=QFileDialog.getSaveFileName(self,'Save the file')
        if response!='':
            self.df.to_excel(rf'{response}.xlsx',index=False)
        else:
            pass
## ________________________________________________________________ The Main Window ____________________________________________________ #
class MainWindow(QMainWindow): 

    #progress=pyqtSignal(int)   
    def __init__(self): 
        super().__init__()
        self.ui=Ui_MainWindow()   ## Created by: Qt User Interface Compiler version 5.14.1.   imported from ui_newVersion.py
        self.ui.setupUi(self)

        self.threadpool = QThreadPool()

        self.normalstyle="""
QComboBox {
	border: 2px solid rgb(255, 255, 255);
	border-radius: 5px;
	background-color:rgb(255, 255, 255);
}

QComboBox:on{
	border:2px solid #383f42;
	border-radius:5px;
}

/*style for dropdown area*/
QComboBox::drop-down{
	border:0px;
}

/*style drop down arrow*/
QComboBox::down-arrow{
	image: url(:/icons/Images for Program/icons/chevron-down.svg);
	width: 20px;
	height: 20px;
	margin-right:5px;
}

/*style for menu list*/
QComboBox QListView{
	border:1px solid #383f42;
	padding:5px;
	background-color:#fff;
	outline:0px;
}
        QComboBox QScrollBar:vertical {
        background-color:transparent  ;
        width: 15px;
        margin: 0px 0px 0px 0px;
    }

    QComboBox QScrollBar::handle:vertical {
        background-color:rgb(56, 63, 66) ;
        min-height: 20px;
        margin: 16px 2px 16px 2px;
        border-radius:5px;
    }
    QComboBox QScrollBar::handle:vertical:hover {
        background-color: rgb(255, 205, 0);
    }
    QComboBox QScrollBar::handle:vertical:pressed {
        background-color: rgb(250, 171, 72);
    }
    QComboBox QScrollBar::add-line:vertical, QComboBox QScrollBar::sub-line:vertical {
        background: transparent;
        height: 15px;
        width: 15px;
    }

    QComboBox QScrollBar::add-page:vertical, QComboBox QScrollBar::sub-page:vertical {
        background: none;
        height: 5px;
        width: 5px;
    }

	QComboBox QScrollBar::sub-line:vertical:hover, QComboBox QScrollBar::add-line:vertical:hover {
			background-color:rgb(255, 205, 0);
			border-radius:7px;
            }

	QComboBox QScrollBar::sub-line:vertical:pressed, QComboBox QScrollBar::add-line:vertical:pressed {
			background-color:rgb(250, 171, 72);
			border-radius:7px;
            }

    QComboBox QScrollBar::down-arrow:vertical {
        background-color: transparent;
        border: none;
        width: 15px;
        height: 15px;
        margin: 0px 0px 0px 0px;
            /* replace with the path to your custom image */
        image: url(:/icons/Images for Program/icons/down-arrow.png);
    }

    QComboBox QScrollBar::up-arrow:vertical{
        background-color:transparent;
        border: none;
        width: 15px;
        height: 15px;
        margin: 0px 0px 0px 0px;
            /* replace with the path to your custom image */
        image: url(:/icons/Images for Program/icons/up-arrow.png);
    }

    QComboBox QScrollBar::up-arrow:vertical {
        subcontrol-position: top;
    }

    QComboBox QScrollBar::down-arrow:vertical {
        subcontrol-position: bottom;
    }
"""
        
        self.errorstyle_boxes="""
    QComboBox {
	border: 2px solid rgb(251, 215, 69);
	border-radius: 5px;
    background-color:rgb(255, 252, 155);

}
QComboBox::drop-down{
	border:0px;
}
QComboBox::down-arrow{
	image: url(:/icons/Images for Program/icons/chevron-down.svg);
	width: 20px;
	height: 20px;
	margin-right:5px;
}
QComboBox:on{
	border:3px solid #383f42;
}
QComboBox QListView{
	border:1px solid #383f42;
	padding:5px;
	background-color:#fff;
	outline:0px;
    }
        QComboBox QScrollBar:vertical {
        background-color:transparent  ;
        width: 15px;
        margin: 0px 0px 0px 0px;
    }

    QComboBox QScrollBar::handle:vertical {
        background-color:rgb(56, 63, 66) ;
        min-height: 20px;
        margin: 16px 2px 16px 2px;
        border-radius:5px;
    }
    QComboBox QScrollBar::handle:vertical:hover {
        background-color: rgb(255, 205, 0);
    }
    QComboBox QScrollBar::handle:vertical:pressed {
        background-color: rgb(250, 171, 72);
    }
    QComboBox QScrollBar::add-line:vertical, QComboBox QScrollBar::sub-line:vertical {
        background: transparent;
        height: 15px;
        width: 15px;
    }

    QComboBox QScrollBar::add-page:vertical, QComboBox QScrollBar::sub-page:vertical {
        background: none;
        height: 5px;
        width: 5px;
    }

    QComboBox QScrollBar::sub-line:vertical:hover, QComboBox QScrollBar::add-line:vertical:hover {
			background-color:rgb(255, 205, 0);
			border-radius:7px;
            }

	QComboBox QScrollBar::sub-line:vertical:pressed, QComboBox QScrollBar::add-line:vertical:pressed {
			background-color:rgb(250, 171, 72);
			border-radius:7px;
            }

    QComboBox QScrollBar::down-arrow:vertical {
        background-color: transparent;

        border: none;
        width: 15px;
        height: 15px;
        margin: 0px 0px 0px 0px;
            /* replace with the path to your custom image */
        image: url(:/icons/Images for Program/icons/down-arrow.png);
    }

    QComboBox QScrollBar::up-arrow:vertical{
        background-color:transparent;
        border: none;
        width: 15px;
        height: 15px;
        margin: 0px 0px 0px 0px;
            /* replace with the path to your custom image */
        image: url(:/icons/Images for Program/icons/up-arrow.png);
    }

    QComboBox QScrollBar::up-arrow:vertical {
        subcontrol-position: top;
    }

    QComboBox QScrollBar::down-arrow:vertical {
        subcontrol-position: bottom;
    }
"""

        self.successstyle=""" QLineEdit{border-radius:5px;} 
                            QLineEdit:focus{border: 2px solid #383f42; border-radius:5px;
}"""
        
        self.errorstyle= """ QLineEdit{border-radius:5px;} 
                            QLineEdit:focus{border: 2px solid rgb(251, 215, 69);
                            border-radius:5px;
                            background-color:rgb(255, 252, 155)}"""
        
        self.what_to_calculat=['Anzahl der Pole','DeltaTetagesamt','therm.Widerstand (Wicklung-Blech)','therm.Widerstand (Blech Kühlung)',\
        'therm.Kapazität (Wicklung)','therm.Kapazität (Blech)','Eisenverluste','kupferverluste','Gesamtverluste','Bemessungsleistung',\
        'S6_Leistung','kaltwiderstand [20C°]','Warmwiderstand [100C°]','Thermischer Gesamtwiderstand','Strangwiderstand','Motorinduktivität',\
        'Frequenz','Bemessungsdrehzahl','Puls-Drehzahl','Maximaldrehzahl mechanisch','Maximaldrehzahl [1.8*n_N]',\
        'Bemessungsspannung','Polradspannung','Spannungskonstante','Leiterspannung/1000rpm',\
        'Bemessungsstrom','Puls-Strom','stillstandsstrom','Kurzschlussstrom','Grenzstrom','Strom S1 (100% Air chilled)',\
        'Strom S1 (100% Water chilled)','Strom S6 (40% water chilled)','Drehmoment_Bemessungs','Puls-Drehmoment','Stillstandsdrehmoment','Drehmoment S1 (100% Air chilled)',\
        'Drehmoment S1 (100% Water chilled)','Drehmoment S6 (40% Water chilled)','Drehmomntkonstante','Einsatzdrehzahl_FS']
       
###___________ HIDE ALL "errorlabels"_____________ ###
        self.errorlabels=[self.ui.errorlabel1_1,self.ui.errorlabel1_2,self.ui.errorlabel1_3,self.ui.errorlabel1_4,self.ui.errorlabel1_5,\
        self.ui.errorlabel1_6,self.ui.errorlabel1_7,self.ui.errorlabel1_8,self.ui.errorlabel1_9,self.ui.errorlabel1_10,\
        self.ui.errorlabel1_11,self.ui.errorlabel1_12,self.ui.errorlabel1_13,self.ui.errorlabel1_14,self.ui.errorlabel1_15,\
        self.ui.errorlabel1_16,self.ui.errorlabel1_17,self.ui.errorlabel1_18,self.ui.errorlabel1_19,self.ui.errorlabel1_20,\
        self.ui.errorlabel1_21,self.ui.errorlabel1_22,self.ui.errorlabel1_23,self.ui.errorlabel1_24,self.ui.errorlabel1_25,\
        self.ui.errorlabel1_26,self.ui.errorlabel1_27,self.ui.errorlabel1_28,self.ui.errorlabel1_29,self.ui.errorlabel1_30,\
        self.ui.errorlabel1_31,self.ui.errorlabel1_32,self.ui.errorlabel1_33,self.ui.errorlabel1_34,self.ui.errorlabel1_35,\
        self.ui.errorlabel2_1,self.ui.errorlabel1_36,self.ui.errorlabel1_37,self.ui.errorlabel1_38,self.ui.errorlabel1_39,\
        self.ui.errorlabel1_40, self.ui.progressBarLabel,self.ui.errorlabel2_2,self.ui.errorlabel1_41,self.ui.errorlabel1_42,\
        self.ui.errorlabel_20, self.ui.errorlabel_21, self.ui.errorlabel_22]


### _________________ Desable export Buttons: ____________________ ###

        self.ui.creatword.setEnabled(False)
        self.ui.exportBasisdaten_btn.setEnabled(False)

        for errorlabel in self.errorlabels:
            self.keep_space=errorlabel.sizePolicy()
            self.keep_space.setRetainSizeWhenHidden(True)
            errorlabel.setSizePolicy(self.keep_space)
            errorlabel.hide()

###_________ blue line on the bottom to indicate the selected page from the left side menu ______###         
        self.selectedPageIndex=self.ui.stackedWidget.currentIndex()
        if self.selectedPageIndex==0:
            self.ui.Data_input_btn.setStyleSheet("border-bottom: 4px solid #0261a5")
        elif self.selectedPageIndex==1:
            self.ui.calculat_btn.setStyleSheet("border-bottom: 4px solid #0261a5")
        elif self.selectedPageIndex==2:
            self.ui.Data_sheet_btn.setStyleSheet("border-bottom: 4px solid #0261a5")

### _________________ CLICK EVENTS________________###

        self.ui.Browsefile.clicked.connect(self.browsefile)
        self.ui.BrowsFile_basisdaten1.clicked.connect(self.browsefile_1_Page4)
        self.ui.Submit1.clicked.connect(self.check_calculations_inputs)
        self.ui.Submit2.clicked.connect(self.check_Datenblatt_inputs)
        self.ui.Submit_basisdaten.pressed.connect(self.showProgressLabel)
        self.ui.Submit_basisdaten.clicked.connect(self.when_submit_page4)
        self.ui.calculat_all.clicked.connect(self.All_calculation)
        self.ui.calculat_specific.clicked.connect(self.Specific_calculation)
        self.ui.GraphsBtn.clicked.connect(self.showGraphsWindow)
        self.ui.wordTemplate.clicked.connect(self.Add_Template)
        self.ui.ExcelTemplate.clicked.connect(self.Add_Excel_Template)
        self.ui.mot3DateiImport.clicked.connect(self.importOEMdatei)
        self.ui.creatword.clicked.connect(self.CreateWord)
        self.ui.creatExcel.clicked.connect(self.exportExcel)
        self.ui.list_to_calculat.itemClicked.connect(self.items_selected)
        self.ui.Motoren_box.currentIndexChanged.connect(self.checkmotor)
        self.ui.sensorTypcomboBox.currentIndexChanged.connect(self.checkTypTempSensor)
        self.ui.Encoder_Auswertrichtung_comboBox.currentIndexChanged.connect(self.checkEncoder_Auswertrichtung)
        self.ui.sensorTypcomboBox_ANW.currentIndexChanged.connect(self.checkTypTempSensor_ANW)
        self.ui.exportBasisdaten_btn.clicked.connect(self.when_export_clicked_page4)
        self.ui.all_variations_btn.clicked.connect(self.when_mono_export_clicked)
        self.ui.calculat_all_2.clicked.connect(self.reverceCalculation)
        self.ui.importVorlageBtn.clicked.connect(self.Add_DatenBlattVorlage)
        self.ui.generatDataSheetBtn.clicked.connect(self.Export_Data_Sheet)

### ________ check all entred data in the line_edits ____________ ###   

        self.ui.Motoren_box.currentIndexChanged.connect(self.motorboxchangend)
        self.ui.Tw_lineedit.editingFinished.connect(self.checkWT)
        self.ui.Tw_lineedit_page4.editingFinished.connect(self.checkWT_page4)
        self.ui.Tk_lineedit.editingFinished.connect(self.checkKT)
        self.ui.Tk_lineedit_page4.editingFinished.connect(self.checkKT_page4)
        self.ui.laengeBox.currentIndexChanged.connect(self.laengeBoxChanged)
        self.ui.laengeBox.editTextChanged.connect(self.laengeBoxChanged)
        self.ui.laengeBox.currentIndexChanged.connect(self.laengeBoxChanged_pulsdauer)
        self.ui.laengeBox.editTextChanged.connect(self.laengeBoxChanged_pulsdauer)
        self.ui.Drehzahl_lineedit.editingFinished.connect(self.checkDrehzahl)
        self.ui.Teta_stern_lineedit.editingFinished.connect(self.checkThetaStern)
        self.ui.DeltaTeta_stern_lineedit.editingFinished.connect(self.checkDeltaThetaStern)
        self.ui.Pulsdauer_lineedit.editingFinished.connect(self.checkPulsDauer)
        self.ui.Strom_lineEdit.editingFinished.connect(self.checkStrom)
        
        self.ui.Tw_lineedit.cursorPositionChanged.connect(self.checkWT)
        self.ui.Tw_lineedit_page4.cursorPositionChanged.connect(self.checkWT_page4)
        self.ui.Tk_lineedit.cursorPositionChanged.connect(self.checkKT)
        self.ui.Tk_lineedit_page4.cursorPositionChanged.connect(self.checkKT_page4)
        self.ui.laengeBox.editTextChanged.connect(self.checkLaenge)
        self.ui.Drehzahl_lineedit.cursorPositionChanged.connect(self.checkDrehzahl)
        self.ui.Teta_stern_lineedit.cursorPositionChanged.connect(self.checkThetaStern)
        self.ui.DeltaTeta_stern_lineedit.cursorPositionChanged.connect(self.checkDeltaThetaStern)
        self.ui.Pulsdauer_lineedit.cursorPositionChanged.connect(self.checkPulsDauer)
        self.ui.Strom_lineEdit.cursorPositionChanged.connect(self.checkStrom)

        self.ui.Motortyp_lineedit.editingFinished.connect(self.checkMotorTyp)
        self.ui.Motor_ID_lineedit.editingFinished.connect(self.checkMotorID)
        self.ui.stator_ID_lineedit.editingFinished.connect(self.checkStatorID)
        self.ui.rotor_ID_lineedit.editingFinished.connect(self.checkRotorID)
        self.ui.motor_DBL_lineedit.editingFinished.connect(self.checkMotorDBL)
        self.ui.motor_AS_lineedit.editingFinished.connect(self.checkMotorAS)
        self.ui.Trgheitsmoment_lineedit.editingFinished.connect(self.TraegheitsMoment)
        
        self.ui.zwischenkreisspannung_lineedit.editingFinished.connect(self.checkZwischenkreisspannung)
        self.ui.Vorschaltdrossel_lineedit.editingFinished.connect(self.checkVorschaltdrossel)
        self.ui.Drehzahlbegrenzung_lineedit.editingFinished.connect(self.checkDrehzahlbegrenzung)
        
        self.ui.Motortyp_lineedit.cursorPositionChanged.connect(self.checkMotorTyp)
        self.ui.Motor_ID_lineedit.cursorPositionChanged.connect(self.checkMotorID)
        self.ui.stator_ID_lineedit.cursorPositionChanged.connect(self.checkStatorID)
        self.ui.rotor_ID_lineedit.cursorPositionChanged.connect(self.checkRotorID)
        self.ui.motor_DBL_lineedit.cursorPositionChanged.connect(self.checkMotorDBL)
        self.ui.motor_AS_lineedit.cursorPositionChanged.connect(self.checkMotorAS)
        self.ui.Trgheitsmoment_lineedit.cursorPositionChanged.connect(self.TraegheitsMoment)
        #self.ui.sensorTypcomboBox.cursorPositionChanged.connect(self.checkTypTempSensor)
        self.ui.zwischenkreisspannung_lineedit.cursorPositionChanged.connect(self.checkZwischenkreisspannung)
        self.ui.Vorschaltdrossel_lineedit.cursorPositionChanged.connect(self.checkVorschaltdrossel)
        self.ui.Drehzahlbegrenzung_lineedit.cursorPositionChanged.connect(self.checkDrehzahlbegrenzung)

        self.ui.DBL_version_lineedit.editingFinished.connect(self.checAnwendugsDBL)
        self.ui.DBL_ID_lineedit.editingFinished.connect(self.checkAnwendungsDBL_ID)
        self.ui.Anwendugs_AS_lineedit.editingFinished.connect(self.checkAnwendungs_AS)
        
        self.ui.Encoder_Strichzahl_lineedit.editingFinished.connect(self.checkEncoder_Strichzahl)
        #self.ui.Encoder_Schnitstelle_comboBox.editingFinished.connect(self.checkEncoder_Schnitstelle)
        
        #self.ui.Encoder_Hersteller_comboBox.editingFinished.connect(self.checkEncoder_Hersteller)
        self.ui.Encoder_Bezeichnung_lineedit.editingFinished.connect(self.checkEncoder_Bezeichnung)
        self.ui.Drehzahlbegrenzung2_lineedit.editingFinished.connect(self.checkDrehzahlbegrenzung_ANW)
        self.ui.Vorschaltdrossel2_lineedit.editingFinished.connect(self.checkVorschaltdrossel_ANW)

        self.ui.DBL_version_lineedit.cursorPositionChanged.connect(self.checAnwendugsDBL)
        self.ui.DBL_ID_lineedit.cursorPositionChanged.connect(self.checkAnwendungsDBL_ID)
        self.ui.Anwendugs_AS_lineedit.cursorPositionChanged.connect(self.checkAnwendungs_AS)
        #self.ui.sensorTypcomboBox_ANW.cursorPositionChanged.connect(self.checkTypTempSensor_ANW)
        self.ui.Encoder_Strichzahl_lineedit.cursorPositionChanged.connect(self.checkEncoder_Strichzahl)
        #self.ui.Encoder_Schnitstelle_comboBox.cursorPositionChanged.connect(self.checkEncoder_Schnitstelle)
        #self.ui.Encoder_Auswertrichtung_comboBox.cursorPositionChanged.connect(self.checkEncoder_Auswertrichtung)
        #self.ui.Encoder_Hersteller_comboBox.cursorPositionChanged.connect(self.checkEncoder_Hersteller)
        self.ui.Encoder_Bezeichnung_lineedit.cursorPositionChanged.connect(self.checkEncoder_Bezeichnung)
        self.ui.Drehzahlbegrenzung2_lineedit.cursorPositionChanged.connect(self.checkDrehzahlbegrenzung_ANW)
        self.ui.Vorschaltdrossel2_lineedit.cursorPositionChanged.connect(self.checkVorschaltdrossel_ANW)
        
        ## REMOVE WINDOW TITLE BAR
        #self.setWindowFlags(QtCore.Qt.FramelessWindowHint)

        ## ____________ STACKED PAGES NAVIGATION ________________ ##
        self.ui.pushButton_next.clicked.connect(self.NextButtonClicked)
        self.ui.pushButton_back.clicked.connect(self.BackButtonClicked)
        ## ________ Navigation using the side menu ________##
        # navigation to inputsPage:
        self.ui.Data_input_btn.clicked.connect(self.Data_input_btn_event)
        # navigate to calculationPage:
        self.ui.calculat_btn.clicked.connect(self.calculat_btn_event)
        # navigate to ExportPage:
        self.ui.Data_sheet_btn.clicked.connect(self.Data_sheet_btn_event)
        # navigation to BasisdatenPage:
        self.ui.basisdaten_btn.clicked.connect(self.Basisdaten_btn_event)

        ## optional frame slide (show and hide) ##
        self.ui.optionalBtn.clicked.connect(self.slideOptinalFrame)

        ## show the frame to change Strom ## :
        self.ui.Strom_checkBox.stateChanged.connect(self.slideStromnChangeFrame)

        ###  left menu toggle button (sow and hide)####
        self.ui.Menu_btn.clicked.connect(self.slideLeftMenu)

        ### Style cklicked menu button: ###
        for w in self.ui.menu_frame.findChildren(QPushButton):
            #Add click event listener:
            w.clicked.connect(self.applyButtonStyle)

    def NextButtonClicked(self):
        self.ui.stackedWidget.setCurrentIndex(self.ui.stackedWidget.currentIndex()+1)
        self.ui.stackedWidget_2.setCurrentIndex(self.ui.stackedWidget_2.currentIndex()+1)

        if self.ui.stackedWidget.currentIndex()==0:
            self.ui.Data_input_btn.setStyleSheet("border-bottom: 4px solid #0261a5")
            self.ui.calculat_btn.setStyleSheet("border-bottom: none;")
            self.ui.Data_sheet_btn.setStyleSheet("border-bottom: none;")
            self.ui.basisdaten_btn.setStyleSheet("border-bottom: none;")

        if self.ui.stackedWidget.currentIndex()==1:
            self.ui.Data_input_btn.setStyleSheet("border-bottom: none;")
            self.ui.calculat_btn.setStyleSheet("border-bottom: 4px solid #0261a5")
            self.ui.Data_sheet_btn.setStyleSheet("border-bottom: none;")
            self.ui.basisdaten_btn.setStyleSheet("border-bottom: none;")

        if self.ui.stackedWidget.currentIndex()==2:
            self.ui.Data_input_btn.setStyleSheet("border-bottom: none;")
            self.ui.calculat_btn.setStyleSheet("border-bottom: none;")
            self.ui.Data_sheet_btn.setStyleSheet("border-bottom: 4px solid #0261a5")
            self.ui.basisdaten_btn.setStyleSheet("border-bottom: none;")
        
        if self.ui.stackedWidget.currentIndex()==3:
            self.ui.Data_input_btn.setStyleSheet("border-bottom: none;")
            self.ui.calculat_btn.setStyleSheet("border-bottom: none;")
            self.ui.Data_sheet_btn.setStyleSheet("border-bottom: none;")
            self.ui.basisdaten_btn.setStyleSheet("border-bottom: 4px solid #0261a5")

        if 0<= self.ui.stackedWidget.currentIndex()<=2:
            self.ui.pushButton_next.setEnabled(True)
        else:
            self.ui.pushButton_next.setEnabled(False)

        if 1<= self.ui.stackedWidget.currentIndex()<=3:
            self.ui.pushButton_back.setEnabled(True)
        else:
            self.ui.pushButton_back.setEnabled(False)

    def BackButtonClicked(self):
        self.ui.stackedWidget.setCurrentIndex(self.ui.stackedWidget.currentIndex()-1)
        self.ui.stackedWidget_2.setCurrentIndex(self.ui.stackedWidget_2.currentIndex()-1)

        if self.ui.stackedWidget.currentIndex()==0:
            self.ui.Data_input_btn.setStyleSheet("border-bottom: 4px solid #0261a5")
            self.ui.calculat_btn.setStyleSheet("border-bottom: none;")
            self.ui.Data_sheet_btn.setStyleSheet("border-bottom: none;")
            self.ui.basisdaten_btn.setStyleSheet("border-bottom: none;")
            
        if self.ui.stackedWidget.currentIndex()==1:
            self.ui.Data_input_btn.setStyleSheet("border-bottom: none;")
            self.ui.calculat_btn.setStyleSheet("border-bottom: 4px solid #0261a5")
            self.ui.Data_sheet_btn.setStyleSheet("border-bottom: none;")
            self.ui.basisdaten_btn.setStyleSheet("border-bottom: none;")

        if self.ui.stackedWidget.currentIndex()==2:
            self.ui.Data_input_btn.setStyleSheet("border-bottom: none;")
            self.ui.calculat_btn.setStyleSheet("border-bottom: none;")
            self.ui.Data_sheet_btn.setStyleSheet("border-bottom: 4px solid #0261a5")
            self.ui.basisdaten_btn.setStyleSheet("border-bottom: none;")

        if self.ui.stackedWidget.currentIndex()==3:
            self.ui.Data_input_btn.setStyleSheet("border-bottom: none;")
            self.ui.calculat_btn.setStyleSheet("border-bottom: none;")
            self.ui.Data_sheet_btn.setStyleSheet("border-bottom: none;")
            self.ui.basisdaten_btn.setStyleSheet("border-bottom: 4px solid #0261a5")

        if 1<= self.ui.stackedWidget.currentIndex()<=2:
            self.ui.pushButton_back.setEnabled(True)
        else:
            self.ui.pushButton_back.setEnabled(False) 

        if 0<= self.ui.stackedWidget.currentIndex()<=3:
            self.ui.pushButton_next.setEnabled(True)
        else:
            self.ui.pushButton_next.setEnabled(False)

    def Data_input_btn_event(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.inputsPage)
        self.ui.stackedWidget_2.setCurrentWidget(self.ui.inputData_hinweise)
        if 0<= self.ui.stackedWidget.currentIndex()<=1:
            self.ui.pushButton_next.setEnabled(True)
        else:
            self.ui.pushButton_next.setEnabled(False)

        if 1<= self.ui.stackedWidget.currentIndex()<=2:
            self.ui.pushButton_back.setEnabled(True)
        else:
            self.ui.pushButton_back.setEnabled(False)
        
    def calculat_btn_event(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.calculationPage)
        self.ui.stackedWidget_2.setCurrentWidget(self.ui.calculation_hinweise)
        if 0<= self.ui.stackedWidget.currentIndex()<=2:
            self.ui.pushButton_next.setEnabled(True)
        else:
            self.ui.pushButton_next.setEnabled(False)

        if 1<= self.ui.stackedWidget.currentIndex()<=3:
            self.ui.pushButton_back.setEnabled(True)
        else:
            self.ui.pushButton_back.setEnabled(False)        

    def Data_sheet_btn_event(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.ExportPage)
        self.ui.stackedWidget_2.setCurrentWidget(self.ui.Export_hinweise)
        if 0<= self.ui.stackedWidget.currentIndex()<=2:
            self.ui.pushButton_next.setEnabled(True)
        else:
            self.ui.pushButton_next.setEnabled(False)

        if 1<= self.ui.stackedWidget.currentIndex()<=3:
            self.ui.pushButton_back.setEnabled(True)
        else:
            self.ui.pushButton_back.setEnabled(False)

    def Basisdaten_btn_event(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.BasisdatenPage)
        self.ui.stackedWidget_2.setCurrentWidget(self.ui.Basisdaten_hinweise)
        if 0<= self.ui.stackedWidget.currentIndex()<=2:
            self.ui.pushButton_next.setEnabled(True)
        else:
            self.ui.pushButton_next.setEnabled(False)

        if 1<= self.ui.stackedWidget.currentIndex()<=3:
            self.ui.pushButton_back.setEnabled(True)
        else:
            self.ui.pushButton_back.setEnabled(False)

    def applyButtonStyle(self):
        # Reset style for other buttons:
        for w in self.ui.menu_frame.findChildren(QPushButton):
            # if the Button name is not equal to clicked button name, then do the following:
            if w.objectName()!=self.sender().objectName():
                #creat default style by removing the left border, bottom border style also remove the left border style
                #apply the default style(no borders at all)
                w.setStyleSheet("border-bottom: none;")
        #apply the style to the clicked button:(set the bottom border to 2px)
        self.sender().setStyleSheet("border-bottom: 4px solid #0261a5")
        return

    def slideOptinalFrame(self):
        # get current left menu height
        height=self.ui.optionalFrame.height()

        # if minimized
        if height==0:
            newHeight=240
        else:
            newHeight=0
        self.animation=QPropertyAnimation(self.ui.optionalFrame,b"maximumHeight") # animate the maximumHeight of the frame
        self.animation.setDuration(200)
        self.animation.setStartValue(height) # strat from the the current height value of the frame
        self.animation.setEndValue(newHeight) # end to the new height value of the frame
        self.animation.setEasingCurve(QtCore.QEasingCurve.InOutQuart)
        self.animation.start()

    def slideStromnChangeFrame(self):
        height=self.ui.frame_20.height()
        if self.ui.Strom_checkBox.isChecked():
            self.checkStrom()
        # if minimized
            newHeight=100
            self.animation=QPropertyAnimation(self.ui.frame_20,b"maximumHeight") # animate the maximumHeight of the frame
            self.animation.setDuration(200)
            self.animation.setStartValue(height) # strat from the the current height value of the frame
            self.animation.setEndValue(newHeight) # end to the new height value of the frame
            self.animation.setEasingCurve(QtCore.QEasingCurve.InOutQuart)
            self.animation.start()
        else:

            newHeight=0
            self.animation=QPropertyAnimation(self.ui.frame_20,b"maximumHeight") # animate the maximumHeight of the frame
            self.animation.setDuration(200)
            self.animation.setStartValue(height) # strat from the the current height value of the frame
            self.animation.setEndValue(newHeight) # end to the new height value of the frame
            self.animation.setEasingCurve(QtCore.QEasingCurve.InOutQuart)
            self.animation.start()
            self.ui.pushButton_next.setEnabled(True)
            self.ui.calculat_btn.setEnabled(True)
            self.ui.Data_sheet_btn.setEnabled(True)
            
    def slideLeftMenu(self):
        # get current left menu width
        width=self.ui.left_menu_content_frame.width()
        # if maximized
        if width==200:
            newWidth=65
        # if minimized
        else:
            newWidth=200
        self.animation=QPropertyAnimation(self.ui.left_menu_content_frame,b"minimumWidth")# animate minimumWidth
        self.animation.setDuration(250)
        self.animation.setStartValue(width) #start value is the current menu width
        self.animation.setEndValue(newWidth) # end value is the new menu width
        self.animation.setEasingCurve(QtCore.QEasingCurve.InOutQuart)
        self.animation.start()

    def browsefile(self):
        self.filename=QFileDialog.getOpenFileName(self,'Open file',os.getenv('HOME'),'csv files(*.csv)')
        self.filepath=self.ui.filelabel.setText(self.filename[0])
        filepath=self.ui.filelabel.text()
        try:
            if filepath=='' or bool(filepath)==False or len(filepath)==0:
                self.ui.errorlabel1_11.show()
                self.ui.filelabel.setStyleSheet(self.errorstyle)
            else:
                self.ui.errorlabel1_11.hide()
                self.ui.filelabel.setStyleSheet(self.successstyle)
                self.ui.Motoren_box.setEnabled(True)
                self.ui.Submit1.setEnabled(True)
        except ValueError:
            self.ui.errorlabel1_11.show()
            self.ui.filelabel.setStyleSheet(self.errorstyle)
    
    def motorboxchangend(self):

        ''' i deleted the .apply(pd.to_numeric) from self.Daten'''

        self.Daten = pd.read_csv(self.filename[0],delimiter=';', index_col='Parameter',decimal=',').fillna('')
        self.Daten[["200HX", "200UHX", "240HX", "240UHX", "310HX", "310UHX", "360UHX", "410HX", "410UHX", "564HX"]] = self.Daten[["200HX", "200UHX", "240HX", "240UHX", "310HX", "310UHX", "360UHX", "410HX", "410UHX", "564HX"]].apply(pd.to_numeric).fillna('')

        self.all_VV=['1','2','3','4','5','6','8','10','12']
        print(self.all_VV)
        self.ui.VV_Box.clear()
        self.ui.VV_Box.setStyleSheet(self.normalstyle)
        #t_p_keys= ['25mm','40mm','50mm','75mm','100mm','125mm','150mm','175mm','200mm','225mm','250mm','275mm','300mm']
        vv_keys = ["Seriel", "2TM", "3TM", "4TM", "5TM", "6TM", "8TM", "10TM", "12TM"]
        Motor_Name=self.ui.Motoren_box.currentText()
        vv_values = self.Daten[Motor_Name][vv_keys]
        print('vv_values 1 =' ,vv_values)
        vv_values = vv_values[vv_values!=''].astype(int)
        print('vv_values 2 =' ,vv_values)
        vv_values = vv_values.to_string(header=False,index=False)
        print('vv_values 3 =' ,vv_values)
        self.VV_Updats=list(vv_values.split())
        #self.VV_Updats= [int(i) for i in self.VV_Updats]
        print('VV_Updates =' ,self.VV_Updats)
        print('VV_Updates type' ,type(self.VV_Updats))
       
        #print(self.VV_Updats)
        # self.pulsDauer=self.Daten[Motor_Name][t_p_keys]
        # self.pulsDauer=self.pulsDauer[self.pulsDauer!=''].astype(float)
        # self.pulsDauer=self.pulsDauer.to_string(header=False,index=False)
        # self.pulsDauer_updat=list(self.pulsDauer.split())
        # print(self.pulsDauer_updat)
        self.ui.laengeBox.setEnabled(True)
        try:
            if self.VV_Updats==['Series([],', ')'] :
                for i in self.all_VV:
                    item_index=self.all_VV.index(i)
                    self.ui.VV_Box.addItem(i)
                    self.ui.VV_Box.setItemData(item_index, QColor(255, 255, 0), role=Qt.BackgroundRole)

                # self.ui.VV_Box.setCurrentText('Keine')
                # self.ui.errorlabel1_12.show()
                # self.ui.errorlabel1_7.show()
                # self.ui.VV_Box.setStyleSheet(self.errorstyle_boxes)
                # self.ui.Motoren_box.setStyleSheet(self.errorstyle_boxes)
                # self.ui.pushButton_next.setEnabled(False)
                # self.ui.calculat_btn.setEnabled(False)
                # self.ui.Data_sheet_btn.setEnabled(False)
            else:
                # Convert list elements to integers for comparison
                self.all_VV = list(map(int, self.all_VV))
                self.VV_Updats = list(map(int, self.VV_Updats))

                # Find missing numbers
                missing_numbers = list(set(self.all_VV) - set(self.VV_Updats))

                # Insert missing numbers into list2 while maintaining the order
                for num in sorted(missing_numbers, reverse=False):
                    self.VV_Updats.insert(num - 1, int(num))
                self.VV_Updats.sort()

                missing_numbers= [str(j) for j in missing_numbers]
                print('missing_numbers:',missing_numbers)
                self.VV_Updats= [str(i) for i in self.VV_Updats]
                print("List 2:", self.VV_Updats)

                for i in self.VV_Updats:
                    if i in missing_numbers:
                        item_index=self.VV_Updats.index(i)
                        self.ui.VV_Box.addItem(i)
                        self.ui.VV_Box.setItemData(item_index, QColor(255, 255, 0), role=Qt.BackgroundRole)
                    else:
                        self.ui.VV_Box.addItem(i)
                #self.ui.Motortyp_lineedit.setText(Motor_Name+'-'int())

                self.condition10=self.ui.errorlabel1_7.hide()
                self.ui.errorlabel1_12.hide()
                self.ui.VV_Box.setStyleSheet(self.normalstyle)
                self.ui.Motoren_box.setStyleSheet(self.normalstyle)
        except ValueError:
            self.ui.errorlabel1_7.show()
            self.ui.VV_Box.setStyleSheet(self.errorstyle_boxes)
            self.ui.Motoren_box.setStyleSheet(self.errorstyle_boxes)
            self.ui.pushButton_next.setEnabled(False)
            self.ui.calculat_btn.setEnabled(False)
            self.ui.Data_sheet_btn.setEnabled(False)

        # Motor check __________________________:
    
    def checkmotor(self):
        Motor_Name=self.ui.Motoren_box.currentText()
        try: 
            if bool(Motor_Name)==False or Motor_Name=="" or len(Motor_Name)==0 or self.ui.Motoren_box.currentIndex==-1:
                self.ui.errorlabel1_1.show()
                self.ui.Motoren_box.setStyleSheet(self.errorstyle_boxes)
                self.ui.pushButton_next.setEnabled(False)
                self.ui.calculat_btn.setEnabled(False)
                self.ui.Data_sheet_btn.setEnabled(False)
                self.ui.list_to_calculat.clear()
                
            elif self.ui.Motoren_box.currentIndex!=-1:
                condition1=self.ui.errorlabel1_1.hide()
                self.ui.Motoren_box.setStyleSheet(self.normalstyle)
           
        except ValueError:
            self.ui.errorlabel1_1.show()
            self.ui.Motoren_box.setStyleSheet(self.errorstyle_boxes)
            self.ui.pushButton_next.setEnabled(False)
            self.ui.calculat_btn.setEnabled(False)
            self.ui.Data_sheet_btn.setEnabled(False)
            self.ui.list_to_calculat.clear()

        # Temp_Wicklung check __________________________:
    
    def checkWT(self):

        Temp_Wicklung= self.ui.Tw_lineedit.text()
        try:
            if bool(Temp_Wicklung)== False  or Temp_Wicklung=='' :
                self.ui.errorlabel1_2.show()
                self.ui.Tw_lineedit.setStyleSheet(self.errorstyle)
                self.ui.pushButton_next.setEnabled(False)
                self.ui.calculat_btn.setEnabled(False)
                self.ui.Data_sheet_btn.setEnabled(False)
                self.ui.list_to_calculat.clear()
                
            elif float(Temp_Wicklung.replace(',', '.')) in range(0,500):
                condition2=self.ui.errorlabel1_2.hide()
                self.ui.Tw_lineedit.setStyleSheet(self.successstyle)
                
        except ValueError:
            self.ui.errorlabel1_2.show()
            self.ui.Tw_lineedit.setStyleSheet(self.errorstyle)
            self.ui.pushButton_next.setEnabled(False)
            self.ui.calculat_btn.setEnabled(False)
            self.ui.Data_sheet_btn.setEnabled(False)
            self.ui.list_to_calculat.clear()

        # Temp_Kühlung check __________________________:

    def checkKT(self):
        Temp_Kühlung=self.ui.Tk_lineedit.text()
        try:
            if bool(Temp_Kühlung)==False or Temp_Kühlung is str or Temp_Kühlung=='':
                self.ui.errorlabel1_3.show()
                self.ui.Tk_lineedit.setStyleSheet(self.errorstyle)
                self.ui.pushButton_next.setEnabled(False)
                self.ui.calculat_btn.setEnabled(False)
                self.ui.Data_sheet_btn.setEnabled(False)
                self.ui.list_to_calculat.clear()
            elif float(Temp_Kühlung.replace(',', '.')) in range(0,500):
                condition3=self.ui.errorlabel1_3.hide()
                self.ui.Tk_lineedit.setStyleSheet(self.successstyle)
        except ValueError:
            self.ui.errorlabel1_3.show()
            self.ui.Tk_lineedit.setStyleSheet(self.errorstyle)
            self.ui.pushButton_next.setEnabled(False)
            self.ui.calculat_btn.setEnabled(False)
            self.ui.Data_sheet_btn.setEnabled(False)
            self.ui.list_to_calculat.clear()

        # Spannung check ________________________:
    
    def checkSpannung(self):
        try:
            if self.ui.radioButton_1.isChecked()==False and self.ui.radioButton_2.isChecked()==False and self.ui.radioButton_3.isChecked()==False:
                self.ui.errorlabel1_4.show()
                self.ui.pushButton_next.setEnabled(False)
                self.ui.calculat_btn.setEnabled(False)
                self.ui.Data_sheet_btn.setEnabled(False)
                self.ui.list_to_calculat.clear()
            else:    
                if self.ui.radioButton_1.isChecked()==True:
                    SpannungU=400
                    self.SpannungU1=SpannungU  
                    #print(self.SpannungU1)
                    condition4=self.ui.errorlabel1_4.hide()
                if self.ui.radioButton_2.isChecked()==True:
                    SpannungU=425
                    self.SpannungU1=SpannungU  
                    #print(self.SpannungU1)
                    condition4=self.ui.errorlabel1_4.hide()
                if self.ui.radioButton_3.isChecked()==True:
                    SpannungU=200
                    self.SpannungU1=SpannungU  
                    #print(self.SpannungU1)
                    condition4=self.ui.errorlabel1_4.hide()
                    self.ui.calculat_btn.setEnabled(True)
                    self.ui.Data_sheet_btn.setEnabled(True)
        except ValueError:
            self.ui.errorlabel1_4.show()
            self.ui.pushButton_next.setEnabled(False)
            self.ui.calculat_btn.setEnabled(False)
            self.ui.Data_sheet_btn.setEnabled(False)
            self.ui.list_to_calculat.clear()

        # Motor Länge check ________________________: 
            
    def laengeBoxChanged(self):
        self.laengeKeys=['25','50','75','100','125','150','175','200','225','250','275','300']
        selectedLaenge=self.ui.laengeBox.currentText()
        Motor_Name=self.ui.Motoren_box.currentText()

        try:    
            if selectedLaenge in self.laengeKeys==False :
                try:
                    self.ui.Drehzahl_lineedit.clear()
                    self.ui.Drehzahl_lineedit.setText('500')
                except KeyError:
                    self.ui.Drehzahl_lineedit.clear()
                    self.ui.Drehzahl_lineedit.setText('500')

            else:
                try:
                    TD_Drehzahl=self.Daten[Motor_Name][selectedLaenge]
                    if TD_Drehzahl=='':
                        self.ui.Drehzahl_lineedit.clear()
                        self.ui.Drehzahl_lineedit.setText('500')
                    else:
                        self.ui.Drehzahl_lineedit.setText(str(TD_Drehzahl))
                except KeyError:
                    self.ui.Drehzahl_lineedit.clear()
                    self.ui.Drehzahl_lineedit.setText('500')
        
        except ValueError:
            self.ui.Drehzahl_lineedit.clear()
            self.ui.Drehzahl_lineedit.setText('500')

    def laengeBoxChanged_pulsdauer(self):
        self.laengeKeys_pulsdauer=['25mm','50mm','75mm','100mm','125mm','150mm','175mm','200mm','225mm','250mm','275mm','300mm']
        selectedLaenge=self.ui.laengeBox.currentText()
        Motor_Name=self.ui.Motoren_box.currentText()

        LaengeDic={
            '25':'25mm',
            '50':'50mm',
            '75':'75mm',
            '100':'100mm',
            '125':'125mm',
            '175':'175mm',
            '150':'150mm',
            '200':'200mm',
            '225':'225mm',
            '250':'250mm',
            '275':'275mm',
            '300':'300mm'
        }

        selectedLaenge_pulsdauer=LaengeDic.get(selectedLaenge)
        try:
            if selectedLaenge_pulsdauer in self.laengeKeys_pulsdauer==False or selectedLaenge_pulsdauer==None :
                try:
                    self.ui.Pulsdauer_lineedit.clear()
                    self.ui.Pulsdauer_lineedit.setText('3')
                except KeyError:
                    self.ui.Pulsdauer_lineedit.clear()
                    self.ui.Pulsdauer_lineedit.setText('3')
            
            else:
                try:
                    TD_Pulsdauer=self.Daten[Motor_Name][selectedLaenge_pulsdauer]
                    if TD_Pulsdauer=='' :
                        self.ui.Pulsdauer_lineedit.clear()
                        self.ui.Pulsdauer_lineedit.setText('3')
                    else:
                        self.ui.Pulsdauer_lineedit.setText(str(TD_Pulsdauer))
                except KeyError:
                    self.ui.Pulsdauer_lineedit.clear()
                    self.ui.Pulsdauer_lineedit.setText('3')
        
        except ValueError:
            self.ui.Pulsdauer_lineedit.clear()
            self.ui.Pulsdauer_lineedit.setText('3')

    def checkLaenge(self):
        Motor_Laenge=self.ui.laengeBox.currentText()
        try:
            if bool(Motor_Laenge)==False or Motor_Laenge is str or Motor_Laenge=='':
                self.ui.errorlabel1_5.show()
                self.ui.laengeBox.setStyleSheet(self.errorstyle_boxes)
                self.ui.pushButton_next.setEnabled(False)
                self.ui.calculat_btn.setEnabled(False)
                self.ui.Data_sheet_btn.setEnabled(False)
                self.ui.list_to_calculat.clear()
            elif type(float(Motor_Laenge.replace(',', '.'))) is float:
                condition5=self.ui.errorlabel1_5.hide()
                self.ui.laengeBox.setStyleSheet(self.normalstyle)
        except ValueError:
            self.ui.errorlabel1_5.show()
            self.ui.laengeBox.setStyleSheet(self.errorstyle_boxes)
            self.ui.pushButton_next.setEnabled(False)
            self.ui.calculat_btn.setEnabled(False)
            self.ui.Data_sheet_btn.setEnabled(False)
            self.ui.list_to_calculat.clear()
        
        #Drehzahl check ________________________: 
    
    def checkDrehzahl(self):
        Drehzahl=self.ui.Drehzahl_lineedit.text()
        try:
            if bool(Drehzahl)==False or Drehzahl is str or Drehzahl=='':
                self.ui.errorlabel1_6.show()
                self.ui.Drehzahl_lineedit.setStyleSheet(self.errorstyle)
                self.ui.pushButton_next.setEnabled(False)
                self.ui.calculat_btn.setEnabled(False)
                self.ui.Data_sheet_btn.setEnabled(False)
                self.ui.list_to_calculat.clear()
            elif type(float(Drehzahl.replace(',', '.'))) is float:
                condition6=self.ui.errorlabel1_6.hide()
                self.ui.Drehzahl_lineedit.setStyleSheet(self.successstyle)
        except ValueError:
            self.ui.errorlabel1_6.show()
            self.ui.Drehzahl_lineedit.setStyleSheet(self.errorstyle)
            self.ui.pushButton_next.setEnabled(False)
            self.ui.calculat_btn.setEnabled(False)
            self.ui.Data_sheet_btn.setEnabled(False)
            self.ui.list_to_calculat.clear()
       
        # theta stern check ________________________ :
    
    def checkThetaStern(self):

        teta_stern=self.ui.Teta_stern_lineedit.text()
        try:
            if bool(teta_stern)==False or teta_stern is str or teta_stern=='':
                self.ui.errorlabel1_8.show()
                self.ui.Teta_stern_lineedit.setStyleSheet(self.errorstyle)
                self.ui.pushButton_next.setEnabled(False)
                self.ui.calculat_btn.setEnabled(False)
                self.ui.Data_sheet_btn.setEnabled(False)
                self.ui.list_to_calculat.clear()
            elif type(float(teta_stern)) is float:
                condition7=self.ui.errorlabel1_8.hide()
                self.ui.Teta_stern_lineedit.setStyleSheet(self.successstyle)
        except ValueError:
            self.ui.errorlabel1_8.show()
            self.ui.Teta_stern_lineedit.setStyleSheet(self.errorstyle)
            self.ui.pushButton_next.setEnabled(False)
            self.ui.calculat_btn.setEnabled(False)
            self.ui.Data_sheet_btn.setEnabled(False)  
            self.ui.list_to_calculat.clear()    
    
        #Delta_theta_stern check ___________________: 
    
    def checkDeltaThetaStern(self):
        Delta_Teta_stern=self.ui.DeltaTeta_stern_lineedit.text()
        try:
            if bool(Delta_Teta_stern)==False or Delta_Teta_stern is str or Delta_Teta_stern=='':
                self.ui.errorlabel1_9.show()
                self.ui.DeltaTeta_stern_lineedit.setStyleSheet(self.errorstyle)
                self.ui.pushButton_next.setEnabled(False)
                self.ui.calculat_btn.setEnabled(False)
                self.ui.Data_sheet_btn.setEnabled(False)
                self.ui.list_to_calculat.clear()
            elif type(float(Delta_Teta_stern)) is float:
                condition8=self.ui.errorlabel1_9.hide()
                self.ui.DeltaTeta_stern_lineedit.setStyleSheet(self.successstyle)
        except ValueError:
            self.ui.errorlabel1_9.show()
            self.ui.DeltaTeta_stern_lineedit.setStyleSheet(self.errorstyle)
            self.ui.pushButton_next.setEnabled(False)
            self.ui.calculat_btn.setEnabled(False)
            self.ui.Data_sheet_btn.setEnabled(False)
            self.ui.list_to_calculat.clear()
        
        # Puls_dauer check _________________________:
    
    def checkPulsDauer(self):

        Puls_dauer=self.ui.Pulsdauer_lineedit.text()
        try:
            if bool(Puls_dauer)==False or Puls_dauer is str or Puls_dauer=='':
                self.ui.errorlabel1_10.show()
                self.ui.Pulsdauer_lineedit.setStyleSheet(self.errorstyle)
                self.ui.pushButton_next.setEnabled(False)
                self.ui.calculat_btn.setEnabled(False)
                self.ui.Data_sheet_btn.setEnabled(False)
                self.ui.list_to_calculat.clear()
            elif type(float(Puls_dauer)) is float:
                condition9=self.ui.errorlabel1_10.hide()
                self.ui.Pulsdauer_lineedit.setStyleSheet(self.successstyle)
        except ValueError:
            self.ui.errorlabel1_10.show()
            self.ui.Pulsdauer_lineedit.setStyleSheet(self.errorstyle)
            self.ui.pushButton_next.setEnabled(False)
            self.ui.calculat_btn.setEnabled(False)
            self.ui.Data_sheet_btn.setEnabled(False)
            self.ui.list_to_calculat.clear()

######### _______________ Page 2 _________________ ######

    def checkStrom(self):
        self.stromWert= self.ui.Strom_lineEdit.text()
        try:
            if bool(self.stromWert)== False or self.stromWert is str or self.stromWert=='' :
                self.ui.errorlabel2_2.show()
                self.ui.Strom_lineEdit.setStyleSheet(self.errorstyle)
                self.ui.calculat_all_2.setEnabled(False)
                self.ui.pushButton_next.setEnabled(False)
                self.ui.Data_sheet_btn.setEnabled(False)
                
            elif type(float(self.stromWert.replace(',', '.'))) is float:
                self.conditionForcalculation=self.ui.errorlabel2_2.hide()
                self.ui.Strom_lineEdit.setStyleSheet(self.successstyle)
                self.ui.calculat_all_2.setEnabled(True)
                self.ui.pushButton_next.setEnabled(True)
                self.ui.calculat_btn.setEnabled(True)
                self.ui.Data_sheet_btn.setEnabled(True)
                   
        except ValueError:
            self.ui.errorlabel2_2.show()
            self.ui.Strom_lineEdit.setStyleSheet(self.errorstyle)
            self.ui.calculat_all_2.setEnabled(False)
            self.ui.pushButton_next.setEnabled(False)
            self.ui.Data_sheet_btn.setEnabled(False)

###### basedata  _______________________________________________ :

    def checkMotorTyp(self):

        self.MotorTyp= self.ui.Motortyp_lineedit.text()
        try:
            if bool(self.MotorTyp)== False or self.MotorTyp=='':
                self.ui.errorlabel1_13.show()
                self.ui.Motortyp_lineedit.setStyleSheet(self.errorstyle)

            else:
                self.ui.errorlabel1_13.hide()
                self.ui.Motortyp_lineedit.setStyleSheet(self.successstyle)
                
        except ValueError:
            self.ui.errorlabel1_13.show()
            self.ui.Motortyp_lineedit.setStyleSheet(self.errorstyle)
             

    def checkMotorID(self):
       
        self.MotorID= self.ui.Motor_ID_lineedit.text()
        try:
            if bool(self.MotorID)== False or self.MotorID=='' or len(self.MotorID)!=4:
                self.ui.errorlabel1_14.show()
                self.ui.Motor_ID_lineedit.setStyleSheet(self.errorstyle)
                
            else:
                self.ui.errorlabel1_14.hide()
                self.ui.Motor_ID_lineedit.setStyleSheet(self.successstyle)
                
        except ValueError:
            self.ui.errorlabel1_14.show()
            self.ui.Motor_ID_lineedit.setStyleSheet(self.errorstyle)
             

    def checkStatorID(self):
        
        self.StatorID= self.ui.stator_ID_lineedit.text()
        try:
            if bool(self.StatorID)== False or self.StatorID=='' or len(self.StatorID)!=4:
                self.ui.errorlabel1_15.show()
                self.ui.stator_ID_lineedit.setStyleSheet(self.errorstyle)
                
                
            else:
                self.ui.errorlabel1_15.hide()
                self.ui.stator_ID_lineedit.setStyleSheet(self.successstyle)
                
        except ValueError:
            self.ui.errorlabel1_15.show()
            self.ui.stator_ID_lineedit.setStyleSheet(self.errorstyle)
    

    def checkRotorID(self):

        self.RotorID= self.ui.rotor_ID_lineedit.text()
        try:
            if bool(self.RotorID)== False or self.RotorID=='' or len(self.RotorID)!=4:
                self.ui.errorlabel1_16.show()
                self.ui.rotor_ID_lineedit.setStyleSheet(self.errorstyle)
                
                
            else:
                self.ui.errorlabel1_16.hide()
                self.ui.rotor_ID_lineedit.setStyleSheet(self.successstyle)
                
        except ValueError:
            self.ui.errorlabel1_16.show()
            self.ui.rotor_ID_lineedit.setStyleSheet(self.errorstyle)
            

    def checkMotorDBL(self):
        
        self.MotorDBL= self.ui.motor_DBL_lineedit.text()
        try:
            if bool(self.MotorDBL)== False or self.MotorDBL=='':
                self.ui.errorlabel1_17.show()
                self.ui.motor_DBL_lineedit.setStyleSheet(self.errorstyle)
                
                
            else:
                self.ui.errorlabel1_17.hide()
                self.ui.motor_DBL_lineedit.setStyleSheet(self.successstyle)
                
        except ValueError:
            self.ui.errorlabel1_17.show()
            self.ui.motor_DBL_lineedit.setStyleSheet(self.errorstyle)
            

    def checkMotorAS(self):
        
        self.MotorAS= self.ui.motor_AS_lineedit.text()
        try:
            if bool(self.MotorAS)== False or self.MotorAS=='':
                self.ui.errorlabel1_18.show()
                self.ui.motor_AS_lineedit.setStyleSheet(self.errorstyle)
                
            else:
                self.ui.errorlabel1_18.hide()
                self.ui.motor_AS_lineedit.setStyleSheet(self.successstyle)
                
        except ValueError:
            self.ui.errorlabel1_18.show()
            self.ui.motor_AS_lineedit.setStyleSheet(self.errorstyle)
            

    def TraegheitsMoment(self):
        
        self.Traegheitsmoment=self.ui.Trgheitsmoment_lineedit.text()
        try:
            if bool(self.Traegheitsmoment)==False or self.Traegheitsmoment is str or self.Traegheitsmoment=='':
                self.ui.errorlabel1_19.show()
                self.ui.Trgheitsmoment_lineedit.setStyleSheet(self.errorstyle)
                
            elif type(float(self.Traegheitsmoment.replace(',', '.'))) is float:
                self.ui.errorlabel1_19.hide()
                self.ui.Trgheitsmoment_lineedit.setStyleSheet(self.successstyle)
        except ValueError:
            self.ui.errorlabel1_19.show()
            self.ui.Trgheitsmoment_lineedit.setStyleSheet(self.errorstyle)
            

    def checkTypTempSensor(self):
        
        self.TypTempSensor= self.ui.sensorTypcomboBox.currentText()
        try:
            if bool(self.TypTempSensor)== False or self.TypTempSensor=='' or self.ui.sensorTypcomboBox.currentIndex==-1 :
                self.ui.errorlabel1_20.show()
                self.ui.sensorTypcomboBox.setStyleSheet(self.errorstyle_boxes)
                
            elif self.ui.sensorTypcomboBox.currentIndex!=-1:
                self.ui.errorlabel1_20.hide()
                self.ui.sensorTypcomboBox.setStyleSheet(self.normalstyle)
                
        except ValueError:
            self.ui.errorlabel1_20.show()
            self.ui.sensorTypcomboBox.setStyleSheet(self.errorstyle_boxes)
            

    def checkZwischenkreisspannung(self):
        
        self.Zwischenkreisspannung= self.ui.zwischenkreisspannung_lineedit.text()
        try:
            if bool(self.Zwischenkreisspannung)== False or self.Zwischenkreisspannung is str or self.Zwischenkreisspannung=='' :
                self.ui.errorlabel1_21.show()
                self.ui.zwischenkreisspannung_lineedit.setStyleSheet(self.errorstyle)
                
                
            elif type(float(self.Zwischenkreisspannung)) is float:
                self.ui.errorlabel1_21.hide()
                self.ui.zwischenkreisspannung_lineedit.setStyleSheet(self.successstyle)
                
        except ValueError:
            self.ui.errorlabel1_21.show()
            self.ui.zwischenkreisspannung_lineedit.setStyleSheet(self.errorstyle)
            

    def checkVorschaltdrossel(self):
        
        self.Vorschaltdrossel= self.ui.Vorschaltdrossel_lineedit.text()
        try:
            if bool(self.Vorschaltdrossel)== False or self.Vorschaltdrossel is str or self.Vorschaltdrossel=='' :
                self.ui.errorlabel1_22.show()
                self.ui.Vorschaltdrossel_lineedit.setStyleSheet(self.errorstyle)
                
                
            elif type(float(self.Vorschaltdrossel.replace(',', '.'))) is float:
                self.ui.errorlabel1_22.hide()
                self.ui.Vorschaltdrossel_lineedit.setStyleSheet(self.successstyle)
                
        except ValueError:
            self.ui.errorlabel1_22.show()
            self.ui.Vorschaltdrossel_lineedit.setStyleSheet(self.errorstyle)
            

    def checkDrehzahlbegrenzung(self):
        
        self.Drehzahlbegrenzung= self.ui.Drehzahlbegrenzung_lineedit.text()
        try:
            if self.Drehzahlbegrenzung is str :
                self.ui.errorlabel1_23.show()
                self.ui.Drehzahlbegrenzung_lineedit.setStyleSheet(self.errorstyle)

            elif bool(self.Drehzahlbegrenzung)== False or self.Drehzahlbegrenzung=='' :
                self.ui.Drehzahlbegrenzung_lineedit.setText(str(round(self.DrehzahlbegrenzungWennLeer,0)))     
                self.ui.errorlabel1_23.hide()
                self.ui.Drehzahlbegrenzung_lineedit.setStyleSheet(self.successstyle)

            elif type(float(self.Drehzahlbegrenzung)) is float:
                self.ui.errorlabel1_23.hide()
                self.ui.Drehzahlbegrenzung_lineedit.setStyleSheet(self.successstyle)
                
        except ValueError:
            self.ui.errorlabel1_23.show()
            self.ui.Drehzahlbegrenzung_lineedit.setStyleSheet(self.errorstyle)    

###  Application  _________________________________________________ :

    def checAnwendugsDBL(self):
        
        self.AnwendungDBL_V= self.ui.DBL_version_lineedit.text()
        try:
            if bool(self.AnwendungDBL_V)== False or self.AnwendungDBL_V=='': # or len(AnwendungDBL_V)!=4:
                self.ui.errorlabel1_24.show()
                self.ui.DBL_version_lineedit.setStyleSheet(self.errorstyle)
                
                
            else:
                self.ui.errorlabel1_24.hide()
                self.ui.DBL_version_lineedit.setStyleSheet(self.successstyle)
                
        except ValueError:
            self.ui.errorlabel1_24.show()
            self.ui.DBL_version_lineedit.setStyleSheet(self.errorstyle)
             
    
    def checkAnwendungsDBL_ID(self):
        
        self.AnwendungDBL_ID= self.ui.DBL_ID_lineedit.text()
        try:
            if bool(self.AnwendungDBL_ID)== False or self.AnwendungDBL_ID==''or len(self.AnwendungDBL_ID)!=4:
                self.ui.errorlabel1_25.show()
                self.ui.DBL_ID_lineedit.setStyleSheet(self.errorstyle)
                
            else:
                self.ui.errorlabel1_25.hide()
                self.ui.DBL_ID_lineedit.setStyleSheet(self.successstyle)
                
        except ValueError:
            self.ui.errorlabel1_25.show()
            self.ui.DBL_ID_lineedit.setStyleSheet(self.errorstyle)
            

    def checkAnwendungs_AS(self):
        self.AnwendungsAS= self.ui.Anwendugs_AS_lineedit.text()
        try:
            if bool(self.AnwendungsAS)== False or self.AnwendungsAS=='' or len(self.AnwendungsAS)!=4:
                self.ui.errorlabel1_26.show()
                self.ui.Anwendugs_AS_lineedit.setStyleSheet(self.errorstyle)
                
            else:
                self.ui.errorlabel1_26.hide()
                self.ui.Anwendugs_AS_lineedit.setStyleSheet(self.successstyle)
        except ValueError:
            self.ui.errorlabel1_26.show()
            self.ui.Anwendugs_AS_lineedit.setStyleSheet(self.errorstyle)
            

    def checkTypTempSensor_ANW(self):
        
        self.TypTempSensor_Anw= self.ui.sensorTypcomboBox_ANW.currentText()
        try:
            if bool(self.TypTempSensor_Anw)== False or self.TypTempSensor_Anw=='' or self.ui.sensorTypcomboBox_ANW.currentIndex==-1 :
                self.ui.errorlabel1_27.show()
                self.ui.sensorTypcomboBox_ANW.setStyleSheet(self.errorstyle_boxes)
                
            elif self.ui.sensorTypcomboBox_ANW.currentIndex!=-1:
                self.ui.errorlabel1_27.hide()
                self.ui.sensorTypcomboBox_ANW.setStyleSheet(self.normalstyle)
                
        except ValueError:
            self.ui.errorlabel1_27.show()
            self.ui.sensorTypcomboBox_ANW.setStyleSheet(self.errorstyle_boxes)
            

    def checkEncoder_Strichzahl(self):
        
        self.Encoder_Strichzahl= self.ui.Encoder_Strichzahl_lineedit.text()
        try:
            if bool(self.Encoder_Strichzahl)== False or self.Encoder_Strichzahl=='' :
                self.ui.errorlabel1_28.show()
                self.ui.Encoder_Strichzahl_lineedit.setStyleSheet(self.errorstyle)
                
            else:
                self.ui.errorlabel1_28.hide()
                self.ui.Encoder_Strichzahl_lineedit.setStyleSheet(self.successstyle)
                
        except ValueError:
            self.ui.errorlabel1_28.show()
            self.ui.Encoder_Strichzahl_lineedit.setStyleSheet(self.errorstyle)
            

    def checkEncoder_Schnitstelle(self):
    
        self.Encoder_Schnitstelle= self.ui.Encoder_Schnitstelle_comboBox.currentText()
        try:
            if bool(self.Encoder_Schnitstelle)== False or self.Encoder_Schnitstelle=='' :
                self.ui.errorlabel1_29.show()
                self.ui.Encoder_Schnitstelle_comboBox.setStyleSheet(self.errorstyle_boxes)
                
            else:
                self.ui.errorlabel1_29.hide()
                self.ui.Encoder_Schnitstelle_comboBox.setStyleSheet(self.normalstyle)
                
        except ValueError:
            self.ui.errorlabel1_29.show()
            self.ui.Encoder_Schnitstelle_comboBox.setStyleSheet(self.errorstyle_boxes)
            

    def checkEncoder_Auswertrichtung(self):
        
        self.Encoder_Auswertrichtung= self.ui.Encoder_Auswertrichtung_comboBox.currentText()
        try:
            if self.ui.Encoder_Auswertrichtung_comboBox.currentIndex()==-1:
                self.ui.errorlabel1_30.show()
                self.ui.Encoder_Auswertrichtung_comboBox.setStyleSheet(self.errorstyle_boxes)
                
                
            elif self.ui.Encoder_Auswertrichtung_comboBox.currentIndex()!=-1:
                self.ui.errorlabel1_30.hide()
                self.ui.Encoder_Auswertrichtung_comboBox.setStyleSheet(self.normalstyle)
                
        except ValueError:
            self.ui.errorlabel1_30.show()
            self.ui.Encoder_Auswertrichtung_comboBox.setStyleSheet(self.errorstyle_boxes)
            

    def checkEncoder_Hersteller(self):
        
        self.Hersteller= self.ui.Encoder_Hersteller_comboBox.currentText()
        try:
            if bool(self.Hersteller)== False or self.Hersteller=='' :
                self.ui.errorlabel1_31.show()
                self.ui.Encoder_Hersteller_comboBox.setStyleSheet(self.errorstyle_boxes)
                
                
            else:
                self.ui.errorlabel1_31.hide()
                self.ui.Encoder_Hersteller_comboBox.setStyleSheet(self.normalstyle)
                
        except ValueError:
            self.ui.errorlabel1_31.show()
            self.ui.Encoder_Hersteller_comboBox.setStyleSheet(self.errorstyle_boxes)


    def checkEncoder_Bezeichnung(self):
        
        self.Encoder_Bezeichnung= self.ui.Encoder_Bezeichnung_lineedit.text()
        try:
            if bool(self.Encoder_Bezeichnung)== False or self.Encoder_Bezeichnung=='' :
                self.ui.errorlabel1_32.show()
                self.ui.Encoder_Bezeichnung_lineedit.setStyleSheet(self.errorstyle)
                
            else:
                self.ui.errorlabel1_32.hide()
                self.ui.Encoder_Bezeichnung_lineedit.setStyleSheet(self.successstyle)
                
        except ValueError:
            self.ui.errorlabel1_32.show()
            self.ui.Encoder_Bezeichnung_lineedit.setStyleSheet(self.errorstyle)
            

    def checkDrehzahlbegrenzung_ANW(self):
        
        self.DrehzahlbegrenzungANW= self.ui.Drehzahlbegrenzung2_lineedit.text()
        try:
            if self.DrehzahlbegrenzungANW is str :
                self.ui.errorlabel1_33.show()
                self.ui.Drehzahlbegrenzung2_lineedit.setStyleSheet(self.errorstyle)
                
            elif bool(self.DrehzahlbegrenzungANW)== False or self.DrehzahlbegrenzungANW=='':
                self.DrehzahlbegrenzungANW=''
                self.ui.errorlabel1_33.hide()
                self.ui.Drehzahlbegrenzung2_lineedit.setStyleSheet(self.successstyle)
            
            elif type(float(self.DrehzahlbegrenzungANW)) is float :
                self.ui.errorlabel1_33.hide()
                self.ui.Drehzahlbegrenzung2_lineedit.setStyleSheet(self.successstyle)
       
        except ValueError:
            self.ui.errorlabel1_33.show()
            self.ui.Drehzahlbegrenzung2_lineedit.setStyleSheet(self.errorstyle)
            

    def checkVorschaltdrossel_ANW(self):
        
        self.Vorschaltdrossel_ANW= self.ui.Vorschaltdrossel2_lineedit.text().replace(',', '.')
        try:
            if self.Vorschaltdrossel_ANW is str :
                self.ui.errorlabel1_34.show()
                self.ui.Vorschaltdrossel2_lineedit.setStyleSheet(self.errorstyle)

            elif bool(self.Vorschaltdrossel_ANW)== False or self.Vorschaltdrossel_ANW=='':
                self.ui.Vorschaltdrossel2_lineedit.setText('0')
                self.ui.errorlabel1_34.hide()
                self.ui.Vorschaltdrossel2_lineedit.setStyleSheet(self.successstyle)
                
            elif type(float(self.Vorschaltdrossel_ANW)) is float:
                self.ui.errorlabel1_34.hide()
                self.ui.Vorschaltdrossel2_lineedit.setStyleSheet(self.successstyle)
                
        except ValueError:
            self.ui.errorlabel1_34.show()
            self.ui.Vorschaltdrossel2_lineedit.setStyleSheet(self.errorstyle)
            

    def check_calculations_inputs(self):
        filepath=self.ui.filelabel.text()
        try:
            if filepath=='' or bool(filepath)==False or len(filepath)==0:
                self.ui.errorlabel1_11.show()
                self.ui.filelabel.setStyleSheet(self.errorstyle)
                self.ui.calculat_all.setEnabled(False)
                self.ui.calculat_specific.setEnabled(False)
                self.ui.GraphsBtn.setEnabled(False)
                self.ui.all_variations_btn.setEnabled(False)
            else:
                self.ui.errorlabel1_11.hide()
                self.ui.filelabel.setStyleSheet(self.successstyle)
                self.ui.Motoren_box.setEnabled(True)
                self.ui.Submit1.setEnabled(True)
                self.ui.calculat_all.setEnabled(True)
                self.ui.calculat_specific.setEnabled(True)
                self.ui.GraphsBtn.setEnabled(True)
                self.ui.all_variations_btn.setEnabled(True)

        except ValueError:
            self.ui.errorlabel1_11.show()
            self.ui.filelabel.setStyleSheet(self.errorstyle)
            self.ui.calculat_all.setEnabled(False)
            self.ui.calculat_specific.setEnabled(False)
            self.ui.GraphsBtn.setEnabled(False)
            self.ui.all_variations_btn.setEnabled(False)
        self.ui.list_to_calculat.clear()
        self.ui.pushButton_next.setEnabled(False)

        Motor_Name=self.ui.Motoren_box.currentText()
        for i in self.what_to_calculat :
            self.ui.list_to_calculat.addItem(i)
        self.ui.pushButton_next.setEnabled(True)

        data_to_check=['STK_Magnet','VH','VW','a','Bmax','b','KorrekturFaktor','Flussscheiteltwert','N1_Innenspule','F_NL','F_0','Nutflaeche_Anzahl']
        if self.ui.Motoren_box.currentIndex()!=-1:
            dtc=[self.Daten[Motor_Name].loc[dti]=='' for dti in data_to_check]
            if any(dtc)==True: 
                msg=QMessageBox()
                msg.setWindowTitle('Fehlermeldung')
                msg.setText('Da für einige Motoren keine vollständigen Daten verfügbar sind, können keine Berechnungen für den aktuellen Motor durchgeführt werden !\n\nBitte wählen Sie einen anderen Motor aus')
                msg.setStyleSheet('background-color: rgb(61, 64, 69);font: 75 12pt "Siemens Sans";color:rgb(255, 255, 255)')
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowIcon(QtGui.QIcon("Images for Program/logo.ico"))
                msg.exec_()
                self.ui.list_to_calculat.clear()
                self.ui.pushButton_next.setEnabled(False)
                self.ui.calculat_btn.setEnabled(False)
                self.ui.Data_sheet_btn.setEnabled(False)

        self.checkmotor() 
        self.checkWT()
        self.checkKT()
        self.checkSpannung()
        self.checkLaenge()
        self.checkDrehzahl()
        self.checkThetaStern()
        self.checkDeltaThetaStern()
        self.checkPulsDauer()
        self.is_there_any_visible_label=[label.isVisible() for label in self.errorlabels]
        if any(self.is_there_any_visible_label)==True:
            self.ui.calculat_btn.setEnabled(False)
            self.ui.Data_sheet_btn.setEnabled(False)
            self.ui.Submit2.setEnabled(False)
        else: 
            self.ui.calculat_btn.setEnabled(True)
            self.ui.Data_sheet_btn.setEnabled(True)
            self.ui.Submit2.setEnabled(True)
        
            self.Motor_Name=self.ui.Motoren_box.currentText()
            self.Temp_Wicklung=round(float(self.ui.Tw_lineedit.text().replace(',', '.')),1)
            self.Temp_Kühlung=round(float(self.ui.Tk_lineedit.text().replace(',', '.')),1)
            self.SpannungU1=round(float(self.SpannungU1),1)
            self.Motor_Laenge=round(float(self.ui.laengeBox.currentText().replace(',', '.')),1)
            #print('Motor länge=',self.Motor_Laenge)
            self.Drehzahl=round(float(self.ui.Drehzahl_lineedit.text().replace(',', '.')),1)
            #print('Drehzahl=',self.Drehzahl)
            self.Vv=float(self.ui.VV_Box.currentText())
            # self.Vv=float(''.join(c for c in self.Vv if c.isdecimal()))
            # print('self.Vv= ', self.Vv)
            self.teta_stern=float(self.ui.Teta_stern_lineedit.text())
            self.Delta_Teta_stern=float(self.ui.DeltaTeta_stern_lineedit.text()) # momentan ist 20[K]
            self.Puls_dauer=float(self.ui.Pulsdauer_lineedit.text())
            self.I=self.ui.Strom_lineEdit.text().replace(',', '.')
            
        #self.checkBoxStat=self.ui.Strom_checkBox.isChecked()

### _____________________________ CALCULATION ____________________________________ ###

    """ this is the main calculation function, contains all equations and functions to calculate the data
        to be printed on the DataSheet following the Theory provided from Mr.Josef Hodapp and discussed with Mr.Mohamed Akhiat,
        the Goal is to get rid of the excel Tables to creat DataSheets for the asked motors """

    def calculation_function(self,motorName,temWicklung,tempKühlung,spannung,motorLaenge,verschaltung,tetaStern,DeltaTetaStern,pulsDauer,strom):
        self.MotorName=motorName
        self.Spannung=spannung
        self.MotorLaenge=motorLaenge
        self.TM=verschaltung
        self.strom=strom  

        self.Delta_Teta_gesamt =  int(temWicklung) - int(tempKühlung)
        self.Delta_Teta_gesamt_stern=float(tetaStern-20)

    ## kaltwiderstand R1 bei 20C° _________________ :
        def R1_20C(data, motor_modell,X,VV):
            R1_20C=data[motor_modell]
            return ((R1_20C.loc["R_Strang"]*(1-R1_20C.loc["Want"])*(X/100))+R1_20C.loc["R_Strang"]*R1_20C.loc["Want"])/VV**2
        # if R1_20C(self.Daten,motorName,motorLaenge,verschaltung)<0:
        #     self.R1_20C= -R1_20C(self.Daten,motorName,motorLaenge,verschaltung)
        # else:
        #     self.R1_20C= R1_20C(self.Daten,motorName,motorLaenge,verschaltung)
        # print('R1_20C°= ', self.R1_20C)
        self.R1_20C= R1_20C(self.Daten,motorName,motorLaenge,verschaltung)
        #print('R1_20C°= ', self.R1_20C)

    ## warmwiderstand R1 bei 100C° ________________ :
        self.R1_100C=self.R1_20C*(1+(0.004*self.Delta_Teta_gesamt))
        #print('R1_100C°= ', self.R1_100C)

    ## Gesamtwiderstand R_th_ges gekühlt __________________ :
        def R_th_ges(data, motor_modell,X):
            R_th_ges=data[motor_modell]
            return R_th_ges.loc['Rth']*(100/X)
        self.R_th_ges=R_th_ges(self.Daten,motorName,motorLaenge)
        #print('Rth_ges= ', self.R_th_ges)

    ## S6_Betrieb verluste und Leistung und strom: _______________ :
        self.Pv_s6=80/self.R_th_ges
        print('Pv_s6=',self.Pv_s6)
        self.P_s6=(60/24)*self.Pv_s6
        print('P_s6=',self.P_s6)
        self.I_s6wc=np.sqrt(self.P_s6/(3*self.R1_20C*(1+0.004*100)))
        print('I_s6',self.I_s6wc)

    ## BemessungsStrom I_N _______ auch Strom-ungekühlt :
    ## _____ BemessungsStrom  _____ ## :
        def BemessungsStrom(R1_warm):
            if bool(self.strom)==True and self.ui.Strom_checkBox.isChecked():
                return float(round(strom,1))
            else:
                return np.sqrt(self.Delta_Teta_gesamt/(3*self.R_th_ges*R1_warm))
        self.IN = round(BemessungsStrom(self.R1_100C),2) #latexdatei seite:11, Formel:25
        #print('I_N= ',round(self.IN,1))

    ## SpannungsKonstante ____ ##:
        def Spannungskonstante(data, motor_modell,VV,X): # Berechnung des Spannungskonstante
            Ku = data[motor_modell]
            return 2*0.93*(Ku.loc['Flussscheiteltwert']/1000)*Ku.loc['N1_Innenspule']*(Ku.loc['STK_Magnet']/2)*((Ku.loc['Nutflaeche_Anzahl']/3/2/VV))*(Ku.loc['KorrekturFaktor']/np.sqrt(2))*(X/100)
        self.K_u = Spannungskonstante(self.Daten,motorName,verschaltung,motorLaenge)
        #print('Ku= ', self.K_u)

    ## strang indugtivität Ld_strg ______________ :
        def Induktivitaet_Ld(data, motor_modell,VV,X ): 
            Ld=data[motor_modell]
            #return (Ld.loc['L_ug']/VV**2)*(X/100)
            return ((Ld.loc['L_ug']/1000)*Ld.loc['N1_Innenspule']**2*(Ld.loc['Nutflaeche_Anzahl']/(2*3)))/VV**2*(X/100)
        self.Ld_strg=Induktivitaet_Ld(self.Daten,motorName,verschaltung,motorLaenge)
        #print('L_d_strg= ', self.Ld_strg)

    ## Motorinduktivität L ____________ :
        self.L=self.Ld_strg*2
        #print('L [mH] = ',self.L)

    ## widerstand für Peakstrom bei wiklunkgstemperatur von 80C° ____________ :
        self.R_str_80C=self.R1_20C*(1+(0.004*(self.Delta_Teta_gesamt_stern+(0.8*DeltaTetaStern))))
        #print('R_80C° = ',self.R_str_80C)

    ## Berechnung der Kupfermaße pro Spule m_cu/spule _____________ :
        def Cu_Maesse_Spule(data, motor_modell,X):
            Cu_m_Spule=data[motor_modell]
            # __ Brechnung der Kupfervolumen Wickelköpfe (V_cu) und Aktivkupfer (V_aktiv_cu) bei 100 mm +1mm Bauhöhe:
            V_cu=Cu_m_Spule.loc['V_Wickelkopf_Spule1']*Cu_m_Spule.loc['Nutflaeche_Anzahl']/10**6
            V_aktiv_cu=((Cu_m_Spule.loc['A_cu']*X)/10**6)*2*(Cu_m_Spule.loc['N1_Innenspule']+1)*Cu_m_Spule.loc['Nutflaeche_Anzahl']

            # __ Berechnung der cu gewicht wickelköpfe (m_wickelkopf) und cu gewicht nut (m_cu_Nut)
            m_wickelkopf=Cu_m_Spule.loc['rho_cu']*V_cu
            m_cu_Nut=Cu_m_Spule.loc['rho_cu']*V_aktiv_cu
            # __ cu gewicht ges. ___:
            m_cu=m_wickelkopf+m_cu_Nut
            return m_cu/Cu_m_Spule.loc['Nutflaeche_Anzahl']
        self.m_cu_pro_spule=round(Cu_Maesse_Spule(self.Daten,motorName,motorLaenge),2)
        print('m_cu_spule= ',self.m_cu_pro_spule)

    ## Peakstrom I_p ________________ :
        def Peakstrom(data,motor_modell,Deltatetastern,t_p):
            I_p=data[motor_modell]
            return np.sqrt((383*Deltatetastern*((I_p.loc['Nutflaeche_Anzahl']/3)*((self.m_cu_pro_spule)/1000)))/(self.R_str_80C*t_p))
        self.Ip=round(Peakstrom(self.Daten,motorName,DeltaTetaStern,pulsDauer),2)

        
    # ## GrenzStrom ___________________ :
    #     def grenzstrom(data,motor_modell,vv):
    #         Ig=data[motor_modell]
    #         return (Ig.loc['Igr_ug']/Ig.loc['N1_Innenspule'])*vv
    #     self.Ig=grenzstrom(self.Daten,motorName,verschaltung)

    ## Drehzahl formel _______________ :
        pp=self.Daten[motorName].loc['STK_Magnet']/2
        self.p=(2*self.K_u*self.R1_20C*self.IN)/(self.K_u**2+(pp*(self.Ld_strg/1000)*self.IN)**2)
        #print('p=',self.p)
        self.q=((self.R1_20C*self.IN)**2-230**2)/(self.K_u**2+(pp*(self.Ld_strg/1000)*self.IN)**2)
        #print('q=',self.q)
        self.lösung1=(-self.p/2)+np.sqrt((self.p/2)**2-self.q)
        self.lösung2=(-self.p/2)-np.sqrt((self.p/2)**2-self.q)
        if self.lösung1 > 0 and self.lösung2 <= 0 or self.lösung1>=self.lösung2 :
            self.nN=(self.lösung1/(2*np.pi))*60
        elif self.lösung2 > 0 and self.lösung1 <= 0 or self.lösung2>=self.lösung1:
            self.nN=(self.lösung2/(2*np.pi))*60

        self.n_max_18=round(self.nN*1.8,0)

        #print('lösung 1= ',self.lösung1)
        #print('lösung 2= ',self.lösung2)
        #print('NennDrehzehl= ',self.nN)         

## ___ Frequen z_____ ## :
        def Frequenz1(data,motor_modell,Drehzahl): # Hier muss die Frequenz berechnet werden
            f = data[motor_modell]
            return (((Drehzahl/60)*(float(f.loc['STK_Magnet'])/2)))    # das hier ist die formel f=(n*p)/60  (STK_Magnet is 2p) 
        self.f = Frequenz1(self.Daten,motorName,self.nN)

## __ Anzahl der Pole ______ ##:
        self.Pole = self.Daten[motorName].loc['STK_Magnet']
## __ Eisenverluste ______ ## :
        def EisenVerluste(data,motor_modell,frequenz):  # Spezifische Eisenverluste pro Kg
            Pvfe = data[motor_modell]                                    # diese formel befindet sich in latexDatei' Grundlage für die Berechnung der Motorkennlinien von Achsmotoren' seite:6 formel:14
            return (Pvfe.loc['VH']*(frequenz/50)+Pvfe.loc['VW']*(frequenz/50)**Pvfe.loc['a'])*(Pvfe.loc['Bmax']**Pvfe.loc['b'])
        self.P_fe_sptz=EisenVerluste(self.Daten,motorName,self.f)
        #print('Pfe_sptz= ',round(self.P_fe_sptz,1))

        def EisenVerluste1(data,motor_modell,Pvfe_sptz):  # gesamt Eisenverluste           
            Pfe = data[motor_modell]
            return (Pfe.loc['kb']*Pfe.loc['mb']*(Pvfe_sptz))             # diese formel befindet sich in latex Datei'...' seite 7 Formel 15
        self.P_fe = EisenVerluste1(self.Daten,motorName,self.P_fe_sptz)
        #print('Pfe= ',round(self.P_fe,1))

        # Stationäre Kennlinien: Das Modell wird ohne die Kapazitäten betrachtet und wird hauptsächlich die Eisen- und Kupferverluste berücksichtigen.
        Rwk=self.R_th_ges*10
        #print('Rwk= ',round(Rwk,3))
## _____ therm. Widerstand Wicklung-Blech _____ ## :
        Rwb=0.75*((self.R_th_ges*Rwk)/(Rwk-self.R_th_ges))
        #print('Rwb= ',round(Rwb,3))
        
## _____ therm. Widerstand Blech Kühlung _____ ## :
        Rbk=0.25*((self.R_th_ges*Rwk)/(Rwk-self.R_th_ges))
        #print('Rbk= ',round(Rbk,3))
## _____ therm. Gesamtwiderstand ungekühlt ____ ## :
        Rnk=3.6*self.R_th_ges
        #print('Rnk= ',round(Rnk,3))
## _____ therm. Kapazität Wicklung  ______ ## :
        Cw=self.Daten[motorName].loc['Cw']*(motorLaenge/100)

## _____ therm. Kapazität Blech  ____ ## :
        Cb=self.Daten[motorName].loc['Cb']*(motorLaenge/100)

## _____ Kupververluste ______ ## :
        def Kupferverlust(R1_warm,I,Pfe):
            if bool(self.strom)==True and self.ui.Strom_checkBox.isChecked():
                return 3*R1_warm*I**2
            else:
                return (self.Delta_Teta_gesamt*(Rwk+Rwb+Rbk)-Rbk*Pfe)/(Rwk*(Rwb+Rbk))
                #return 3*R1_warm*self.IN**2
        self.P_cu = Kupferverlust(self.R1_100C,strom,self.P_fe)
        #print('Pcu = ',round(self.P_cu,1))
        # Die Formel für Kupferverlust befindet sich in der latexDatei seite: 8 formel: 17

## ____ Gesamtverluste _____ ## :
        self.Pv_ges= self.P_fe + self.P_cu

## _____ BemessungsStrom für diagram : 
## _____ BemessungsStrom  _____ ## :
        def BemessungsStromDiagram(Pcu,R1_warm):
            if bool(self.strom)==True and self.ui.Strom_checkBox.isChecked():
                return float(round(strom,1))
            else:
                return np.sqrt((Pcu)/(3*R1_warm))
        self.I_N = BemessungsStromDiagram(self.P_cu, self.R1_100C) #latexdatei seite:11, Formel:25
        #print('I_N= ',round(self.I_N,1))

## ____ FlussScheitelWert ____ ## :
        def Flussscheitelwert(data,motor_modell,X):
            Fluss_scheitelwert=data[motor_modell]
            return (float(Fluss_scheitelwert.loc['Flussscheiteltwert'])*float(X))/100
        self.Fluss_SW = Flussscheitelwert(self.Daten,motorName,motorLaenge)

        # die Formel von Ku befindet sich in LatexDatei an der Seite:13 Formel:24

## ____ Drehmomentskonstante ____ ## :
        def Drehmomentkonstante(data, motor_modell,ku): # Berechnung des Drehmomentkonstante
            KT = data[motor_modell]
            return (3*ku)/KT.loc['KorrekturFaktor']
        self.K_T = Drehmomentkonstante(self.Daten,motorName,self.K_u)
        # Formel ist: KT=(3*Ku/1.4142)/Fku; die Formel befindet sich in word_Datei: 2021_06_30 Glgen Motordatenblatt weiter mit 2.3

        # muss die Einheit der Drehzahl in 1/s verwendet. LatexDatei Seite:14 Formel:25

        # Berechnung der Koeffizienten: gedruckte version von LatexDatei seite: 12
        # Bemerkung: im code wurde Mabs als Mquad gerchnet, Mquad als Mabs gerechnet und Mlin wurde korrekt gerechnet
        # aber sie sind richtig in der Formel des Drehmoments angewendet

## ____ PolradSpannung Up _____ ## :
        def PolradSpannung(Ku,Drehzahl):
            return (Ku*2*np.pi*Drehzahl/60)
        self.U_p = PolradSpannung(self.K_u,self.nN)  
        # muss die Einheit der Drehzahl in 1/s verwendet.
        #die Formel von Up_strang befindet sich in LatexDatei an der Seite:14 Formel:25

## ____ Drehzahl-Maximal ____ ## :
        def Drehzahl_max(Ku):
            return ((self.Spannung/np.sqrt(3))/(Ku*2*np.pi))*60
        self.n_N_max = Drehzahl_max(self.K_u)
        #print('n_max=',self.n_N_max)
        self.DrehzahlbegrenzungWennLeer=self.n_N_max

## ____ Drehmoment _____ ## :
        def Drehmoment(data,motor_modell,Strom):
            M_abs = data[motor_modell].loc['Mabs']*(int(motorLaenge)/100)
            M_lin =(data[motor_modell].loc['Mlin'])*(int(motorLaenge)/100)*(np.sqrt(2)*(data[motor_modell].loc['N1_Innenspule'])/verschaltung) 
            M_quad = data[motor_modell].loc['Mquad']*(int(motorLaenge)/100)*(np.sqrt(2)*(data[motor_modell].loc['N1_Innenspule'])/verschaltung)**2

            return M_quad*(Strom)**2 + M_lin*Strom + M_abs

        # Bemessungs_Drehmoment
        self.M_N= Drehmoment(self.Daten,motorName,self.IN)

        # ____________________________________________  ab 13-09-2022 __________________________________________________________#

## ____ Strang Widerstand _____ ## :
        def Strang_Widerstand(data,motor_modell,delta_theta):
            R_strang=data[motor_modell]
            return (R_strang.loc['R_pro_Spule']*(R_strang.loc['Nutflaeche_Anzahl']/3)*(1/verschaltung**2))*(1+(0.004*delta_theta))
        self.R_strang=Strang_Widerstand(self.Daten,motorName,self.Delta_Teta_gesamt)
        
## ____ Leiterspannung /1000rpm nach Siemens-Def _____ ## :
        self.Ku_stern= self.K_u*181.40  #__ Berechnungsweg: 2022_09_20 Grundlagen Motordatenblatt - Update 2022_09_18 Seite 5

## ____ Stillstand_Drehmoment ____ ## :
        self.M_n0=self.M_N*self.Daten[motorName].loc['F_0']   # aus TD

## ____ Stillstand_Strom _____ ## :
        self.I_n0= round(self.Daten[motorName].loc['F_0']*self.IN,2)  # aus TD

## ____ Bemessungs_Leistung _____ ## :
        self.P_N=2*np.pi*self.M_N*(self.nN/60)

        # # R_Strang für peak_Moment:
        # def Strang_stern_Widerstand(data,motor_modell):
        #     R_strang_stern=data[motor_modell]
        #     return (R_strang_stern.loc['R_pro_Spule']*(R_strang_stern.loc['Nutflaeche_Anzahl']/3)*(1/verschaltung)**2)*(1+(0.004*80))      

## ____ Peak_moment ____ ## :
        self.M_p=Drehmoment(self.Daten,motorName,self.Ip)

## ____ strom_100% Air chilled ______ :
        self.I_s1ac=np.sqrt(80/(3*self.R1_100C*(self.R_th_ges*3.61)))

## ____ Drehmoment_100% Air chilled ______ :
        self.M_s1ac=Drehmoment(self.Daten,motorName,self.I_s1ac)

## ____ Drehmoment_40% water chilled _____ :
        self.M_s6wc=Drehmoment(self.Daten,motorName,self.I_s6wc)
        print('M_s6=',self.M_s6wc)

## ____ Leistung_S6_40% Water chilled ______:
        self.P_s6wc=(2*np.pi*self.M_s6wc*(self.nN)/60)/1000
        print('P_s6_40%_wc=',self.P_s6wc)

## ____ Drehzahl_S6_40% Water chilled ______:
        self.n_s6= (self.P_s6wc*1000*60)/(2*np.pi*self.M_s6wc)

        self.f_s6=(self.n_s6/60)*(self.Pole/2)

                                                               
## ____ GrenzStrom _____ ## :
        def Grenzstrom(frequenz,Up):
            Z_quadrat=self.R1_100C**2+(2*np.pi*frequenz*(self.Ld_strg*0.001))**2
            return np.sqrt(((Up*self.R1_100C)/Z_quadrat)**2-((Up**2-(spannung/np.sqrt(3))**2)/Z_quadrat))-((Up*self.R1_100C)/Z_quadrat)
        self.Ig_new=Grenzstrom(self.f,self.U_p)
        print('Ig_new',self.Ig_new)

        # def Stromgrenze(data, motor_modell,frequenz,U_p):
        #     I1g = data[motor_modell]
        #     #if U_p*1.005>=spannung/np.sqrt(3):
        #     if U_p**2<(spannung/np.sqrt(3))**2: #and frequenz > 0:
        #         return (1/(np.pi*frequenz*(self.Ld_strg*0.001)))*np.sqrt((spannung/np.sqrt(3))**2-U_p**2)
        #     else:
        #         return (0)
        # self.I1_g=Stromgrenze(self.Daten,motorName,self.f,self.U_p)

## ____ Kurzschlussstrom ____ ## :
        def Kurzschlusssctrom(data,motor_modell):
            I_k=data[motor_modell]
            return self.Ku_stern/(45.34*I_k.loc['STK_Magnet']*(self.L*0.001))

## ____ Einsatzdrehzahl n_FS ____ ## :
        self.n_FS=60*(np.sqrt(((380/np.sqrt(3))**2)/(((self.K_u)**2)+(self.Pole/2*self.IN*0.25)**2))/2*np.pi)
        #print('n_fs=',self.n_FS)
## ____ Bemessungs_Drehzahl n_N ____ ## :
        #self.n_N=(self.P_N*60)/(2*np.pi*self.M_N)
        #print('Nenndrehzehl= ',round(n_N,0))   

### _________________ Graphs axis definition ____________________ ### :

        self.x1=np.arange(0,self.n_N_max,0.1)
        #self.x1=np.linspace(0,self.nN,100)    # x_axis is Drehzahl [1/min]
        self.y1=PolradSpannung(self.K_u,self.x1)

        Frequenz1(self.Daten,motorName,self.x1)

        self.x2=np.arange(0.1,self.f,0.5)              # x_axis is frequenz [Hz]
        self.y2=EisenVerluste1(self.Daten,motorName,EisenVerluste(self.Daten,motorName,self.x2))
        self.yLeistung=EisenVerluste1(self.Daten,motorName,EisenVerluste(self.Daten,motorName,Frequenz1(self.Daten,motorName,self.x1)))

        # x3 the x_axis is the same as x2 : is frequenz
        self.y3= Kupferverlust(self.R1_100C,strom,self.y2)
        self.y3Leistung= Kupferverlust(self.R1_100C,strom,self.yLeistung)

        # x4 the x_axis is the same as x2 and x3: is frequenz
        self.y4=BemessungsStromDiagram(self.y3, self.R1_100C)
        self.y4Leistung=BemessungsStromDiagram(self.y3Leistung, self.R1_100C)

        self.x5=np.arange(0,self.IN,0.5)
        self.y5=Drehmoment(self.Daten,motorName,self.x5)
        # x6 the x_axis is the same as x2 and x3 and x4: is frequenz

        
        self.I_list=[]         # Liste der Werte des Grenzstroms (Bemesssung)
        self.y6_list=[]         # Liste der Werte des Drehmoments (Bemessungsmoment)
        self.I_list_p=[]         # Liste der Werte des Grenzstroms (peak)
        self.y6_list_p=[]         # Liste der Werte des Drehmoments (Peakmoment)
        self.I_list_S6=[]         # Liste der Werte des Grenzstroms (S6)
        self.y6_list_S6=[]         # Liste der Werte des Drehmomens (S6)

        #self.I1_N=BemessungsStromDiagram(Kupferverlust(self.R1_100C,strom,EisenVerluste1(self.Daten,motorName,EisenVerluste(self.Daten,motorName,Frequenz1(self.Daten,motorName,self.nN)))), self.R1_100C)

# für nennBetrieb
        for n in self.x1:       #Berechnung von Grenzstrom für jedes Drehzahl punkt, dann in einer liste hinzufügen (Bemessung)
            #self.I_g=round(Stromgrenze(self.Daten,motorName,Frequenz1(self.Daten,motorName,n),PolradSpannung(self.K_u,n)),2)
            self.I_g=round(Grenzstrom(Frequenz1(self.Daten,motorName,n),PolradSpannung(self.K_u,n)),2)
            #self.I1_N=BemessungsStromDiagram(Kupferverlust(self.R1_100C,strom,EisenVerluste1(self.Daten,motorName,EisenVerluste(self.Daten,motorName,Frequenz1(self.Daten,motorName,n)))), self.R1_100C)
            # if self.I_g==0 :
            #     self.I_list.append(0)
            if self.I_g>self.IN and n<=self.nN:
                 self.I_list.append(self.IN)
            # elif self.I_g<=self.IN and n<=self.nN:
            #      self.I_list.append(self.IN)
            else:
                self.I_list.append(self.I_g)
        for i in self.I_list:
            self.y6=Drehmoment(self.Daten,motorName,i)
            # if self.y6<=0:
            #     self.y6_list.append(0)
            # else:
            self.y6_list.append(self.y6)
        #print('I_list=',self.I_list)
        #print('________________________')
        #print('M_list=',self.y6_list)

        self.x_axis=np.linspace(0,self.n_N_max,len(self.y6_list))
        #self.x_axis2=np.linspace(0,self.n_N_max,len(self.y6_list))
        #self.x_axis=np.arange(0,self.nN+10,(self.nN+10)/len(self.y6_list))
        #self.xNew=np.linspace(0,self.n_max_b+50,len(self.I_list))
        #print(self.I_list)


# für pulsBetrieb
        for n in self.x1:       #Berechnung von Grenzstrom für jedes Drehzahl punkt, dann in einer liste hinzufügen (Peak)
            #self.I_g=round(Stromgrenze(self.Daten,motorName,Frequenz1(self.Daten,motorName,n),PolradSpannung(self.K_u,n)),2)
            self.I_g=round(Grenzstrom(Frequenz1(self.Daten,motorName,n),PolradSpannung(self.K_u,n)),2)
            # if self.I_g==0 :
            #     self.I_list_p.append(0)
            if self.I_g>self.Ip:
                 self.I_list_p.append(self.Ip)
            else:
                self.I_list_p.append(self.I_g)      
        for j in self.I_list_p:
            self.y6_p=Drehmoment(self.Daten,motorName,j)
            # if self.y6_p<=0:
            #     self.y6_list_p.append(0)
            # else:
            self.y6_list_p.append(self.y6_p)

# für S6-Betrieb:
            
        for n in self.x1:       #Berechnung von Grenzstrom für jedes Drehzahl punkt, dann in einer liste hinzufügen (Peak)
            #self.I_g=round(Stromgrenze(self.Daten,motorName,Frequenz1(self.Daten,motorName,n),PolradSpannung(self.K_u,n)),2)
            self.I_g=round(Grenzstrom(Frequenz1(self.Daten,motorName,n),PolradSpannung(self.K_u,n)),2)
            # if self.I_g==0 :
            #     self.I_list_p.append(0)
            if self.I_g>self.I_s6wc and n<=self.n_s6:
                self.I_list_S6.append(self.I_s6wc)
            # elif self.I_g<=self.I_s6wc and n<=self.n_s6:
            #      self.I_list.append(self.I_s6wc)
            else:
                self.I_list_S6.append(self.I_g)      
        for w in self.I_list_S6:
            self.y6_s6=Drehmoment(self.Daten,motorName,w)
            # if self.y6_p<=0:
            #     self.y6_list_p.append(0)
            # else:
            self.y6_list_S6.append(self.y6_s6)


        ## ____ Bemessungs_Drehzahl n_N ____ ## :
        predict1=interp1d(self.y6_list,self.x_axis, kind='linear')
        self.n_Nenn=predict1(self.M_N)
        #print('nN_ ausdiagram=',self.n_Nenn)
        self.n_MAX=predict1(self.y6_list[-1])
        #print('Ip_list=',self.I_list_p)
        #print('________________________')
        #print('Mp_list=',self.y6_list_p)
        
        # self.M_abs = int(self.Daten[motorName].loc['Mabs']*(int(motorLaenge)/100))
        # self.M_fast_null=Drehmoment(self.Daten,motorName,0)
        # print(self.M_fast_null)
        
        predict2=interp1d(self.y6_list_p,self.x_axis, kind='linear')
        self.n_p=predict2(self.M_p)

        # puls leistung
        self.P_p= (self.n_p*2*np.pi*self.M_p)/60
        # puls frequenz
        self.f_p = (self.n_p/60)*(self.Pole/2)

        # S6 Drehmoment punkt für das diagramm
        self.M_S6_list=[self.M_s6wc,self.M_s6wc,0]
        
        self.n_nenn_list=[0,self.nN-10,self.n_N_max]
        self.n_Puls_list=[0,self.n_p,self.n_N_max]
        self.n_s6_list=[0,self.n_s6,self.n_N_max]

        self.P_N_leistung = [0,self.P_N,self.P_N]
        self.P_N_leistung_Puls=[0,self.P_p, self.P_p]
        self.P_N_leistung_S6=[0,self.P_s6wc*1000, self.P_s6wc*1000]
        #self.x_axis_MS6=np.linspace(0,self.n_N_max,len(self.M_S6_list))

        # self.S6_Strom=np.linspace(0,self.I_s6wc,len(self.x_axis))
        # self.I_N_diagram=np.linspace(0,self.IN,len(self.x_axis))

        # self.M_N_DB_Diagram=[e*2*np.pi*(self.x_axis/60) for e in self.y6_list]
        # self.M_p_DB_Diagram=[e*2*np.pi*(self.x_axis/60) for e in self.y6_list_p]

        self.P_N_leistung_2 = [e*2*np.pi*(n/60) for e,n in zip(self.y6_list,self.x_axis)]
        self.P_N_leistung_Puls_2=[e*2*np.pi*(n/60) for e,n in zip(self.y6_list_p,self.x_axis)]
        self.P_N_leistung_S6_2=[e*2*np.pi*(n/60) for e,n in zip(self.y6_list_S6,self.x_axis)]

        #leistung_interp = np.interp(self.x_axis, [0,self.n_s6,self.n_N_max], [0,self.P_s6wc,self.P_s6wc])
# _________________________________________________________________ add graphs for the datasheet _______________________________________________________________#
        # Drehzahl_Drehmoment digram
        self.fig1, ax1 = plt.subplots()
        #Graphs_Window.Graphs_function(self,self.x1,self.x2,self.x2,self.x2,self.x5,self.x_axis,self.x_axis,self.y1,self.y2,self.y3,self.y4,self.y5,self.y6_list,self.y6_list_p)
        #ax1.plot(self.x_axis,self.y6_list,self.x_axis,self.y6_list_p,self.n_s6_list,self.M_S6_list)
        #ax1.plot(self.x_axis,self.y6_list,self.x_axis,self.y6_list_p,self.x_axis,self.y6_list_S6)

        # Plot each curve with a specified linestyle
        ax1.plot(self.x_axis, self.y6_list, linestyle='-', label='M_N(n)')
        ax1.plot(self.x_axis, self.y6_list_p, linestyle='--', label='M_p(n)')
        ax1.plot(self.x_axis, self.y6_list_S6, linestyle='-.', label='M_s6(n)')
    
        ax1.set_xlabel('Drehzahl in [1/min]', fontsize=7)
        ax1.set_ylabel('Drehmoment in [Nm]', fontsize=7)
        ax1.legend(['M_N(n)','M_p(n)','M_s6(n)'], fontsize=7)
        ax1.tick_params(axis='both', which='both', labelsize=7)
        #ax1.yaxis.set_major_locator(plt.MultipleLocator(100))
        ax1.grid()
        ax1.margins(x=0,y=0)
        current_ylim = ax1.get_ylim()
        ax1.set_ylim(current_ylim[0], current_ylim[1] + 20)
        # Save the Matplotlib figure to a BytesIO object
        self.fig1.set_size_inches(8.6,1.7)
        self.fig1.set_dpi(300)
        #plt.tight_layout()
        self.img_buffer = io.BytesIO()
        self.fig1.savefig(self.img_buffer,format='png', dpi=300 ,bbox_inches='tight')
        self.img_buffer.seek(0)
        plt.close(self.fig1)

        # Drehzahl_Leistung diagram:
        self.fig2, ax2 = plt.subplots()
        #ax2.plot(self.x_axis,self.P_N_leistung_2,self.x_axis,self.P_N_leistung_Puls_2,self.x_axis,self.P_N_leistung_S6_2)
        #ax2.plot(self.n_nenn_list,self.P_N_leistung,self.n_Puls_list,self.P_N_leistung_Puls,self.n_s6_list,self.P_N_leistung_S6)
        
        # Plot each curve with a specified linestyle
        ax2.plot(self.n_nenn_list, self.P_N_leistung, linestyle='-', label='P_S1(n)')
        ax2.plot(self.n_Puls_list, self.P_N_leistung_Puls, linestyle='--', label='P_p(n)')
        ax2.plot(self.n_s6_list, self.P_N_leistung_S6, linestyle='-.', label='P_S6(n)')

        #ax2.step(self.x_axis,leistung_interp, where='post')
        ax2.set_xlabel('Drehzahl in [1/min]', fontsize=7)
        ax2.set_ylabel('Leistung in [KW]', fontsize=7)
        ax2.set_xlim(0,self.n_N_max)
        ax2.legend(['P_S1(n)','P_p(n)','P_S6(n)'], fontsize=7)
        ax2.tick_params(axis='both', which='both', labelsize=7)
        #ax1.yaxis.set_major_locator(plt.MultipleLocator(100))
        ax2.grid()
        ax2.margins(x=0,y=0)
        # Save the Matplotlib figure to a BytesIO object
        current_ylim2 = ax2.get_ylim()
        ax2.set_ylim(current_ylim2[0], current_ylim2[1] + 500)
        ax2.yaxis.set_major_formatter(FuncFormatter(lambda value, _: f"{value / 1000:.0f}"))
        self.fig2.set_size_inches(8.6,1.7)
        self.fig2.set_dpi(300)
        #plt.tight_layout()
        self.img_buffer2 = io.BytesIO()
        self.fig2.savefig(self.img_buffer2,format='png', dpi=300 ,bbox_inches='tight')
        self.img_buffer2.seek(0)
        plt.close(self.fig2)

        self.MotorSpannung_S1=round(np.sqrt(3)*np.sqrt((((self.nN/1000)*(self.K_u*np.sqrt(2)*74.05))+(self.R1_100C*self.IN))**2 + (self.IN*2*np.pi*(self.nN/60)*(self.Pole)/2*(self.L/2000))**2))
        self.MotorSpannung_S6=round(np.sqrt(3)*np.sqrt((((self.n_s6/1000)*(self.K_u*np.sqrt(2)*74.05))+(self.R1_100C*self.I_s6wc))**2 + (self.I_s6wc*2*np.pi*(self.n_s6/60)*(self.Pole)/2*(self.L/2000))**2))
        self.MotorSpannung_Puls=round(np.sqrt(3)*np.sqrt((((self.n_p/1000)*(self.K_u*np.sqrt(2)*74.05))+(self.R1_100C*self.Ip))**2 + (self.Ip*2*np.pi*(self.n_p/60)*(self.Pole)/2*(self.L/2000))**2))
        print('U= ',self.MotorSpannung_S1)
        print('U= ',self.MotorSpannung_S6)
        print('U= ',self.MotorSpannung_Puls)

    # ---------------------------------------------- Ergebnisse auf dem Screen -----------------------------------------------------#

        self.calculated_results1=[self.Pole,self.Delta_Teta_gesamt,round(Rwb,4),round(Rbk,4),round(Cw,4),round(Cb,4)\
            ,round(self.P_fe,2),round(self.P_cu,1),round(self.Pv_ges/1000,1),round(self.P_N/1000,1),round(self.P_s6wc)\
            ,round(self.P_p/1000,1),round(self.R1_20C,3),round(self.R1_100C,3),round(self.R_th_ges,3),round(self.R1_100C,3)\
            ,round(self.L,1),round(self.f),round(self.f_p),round(self.f_s6)\
            ,round(int(self.nN)),round(int(self.n_p)),round(int(self.n_N_max)),self.n_max_18, round(self.n_s6)]#,round(self.n_FS,1)

        self.calculated_results2=[spannung,round(self.U_p,1),round(self.K_u,1),round(self.Ku_stern,2)\
            ,round(self.IN,1),round(self.Ip,1), round(self.I_n0,2),round(Kurzschlusssctrom(self.Daten,motorName),1)\
            ,round(self.Ig_new,1),round(self.I_s1ac,1),round(self.IN,1),round(self.I_s6wc,1)\
            ,round(self.M_N),round(self.M_p),round(self.M_n0),round(self.M_s1ac)\
            ,round(self.M_N),round(self.M_s6wc),round(self.K_T,1),round(self.n_FS,1)]

        #print('bemessungsstrom',self.calculated_results2[3])

        self.list1=['Anzahl der Pole','DeltaTetagesamt','therm.Widerstand (Wicklung-Blech)','therm.Widerstand (Blech Kühlung)',\
        'therm.Kapazität (Wicklung)','therm.Kapazität (Blech)','Eisenverluste','kupferverluste','Gesamtverluste','Bemessungsleistung',\
        'S6_Leistung','Pulsleistung','kaltwiderstand [20C°]','Warmwiderstand [100C°]','Thermischer Gesamtwiderstand','Strangwiderstand','Motorinduktivität',\
        'Frequenz','Puls_Frequenz','S6_Frequenz','Bemessungsdrehzahl','Puls-Drehzahl','Maximaldrehzahl mechanisch','Maximaldrehzahl [1.8*n_N]','S6_Drehzahl']#,'Einsatzdrehzahl'

        self.list2=['Bemessungsspannung','Polradspannung','Spannungskonstante','Leiterspannung/1000rpm',\
        'Bemessungsstrom','Puls-Strom','stillstandsstrom','Kurzschlussstrom','Grenzstrom','Strom S1 (100% Air chilled)',\
        'Strom S1 (100% Water chilled)','Strom S6 (40% water chilled)','Drehmoment_Bemessungs','Puls-Drehmoment','Stillstandsdrehmoment','Drehmoment S1 (100% Air chilled)',\
        'Drehmoment S1 (100% Water chilled)','Drehmoment S6 (40% Water chilled)','Drehmomntkonstante','Einsatzdrehzahl_FS']

        self.Kurzbezeichnungen1=['p','ΔΘ','Rth1','Rth2','Cth1','Cth2','Pv_fe','Pv_cu','Pv_ges','P_N','P_S6','P_p','R1_20C°','R1_100C°','R_th','R_strang','L','f','f_p','f_S6','n_N','n_p','n_max','n_max_1.8','n_S6']#,'n_FS'
        self.Kurzbezeichnungen2=['U_N','Up_str','Ku','Ku*','I_N','I_p','I_0','I_k','I1_g','I_s1ac','I_s1wc','I_s6wc','M_N','M_p','M_0','M_s1ac','M_s1wc','M_s6wc','KT','n_FS']

        self.gesamtBezeichnungen=self.Kurzbezeichnungen1+self.Kurzbezeichnungen2

        self.Einheiten1=['','[K]','[K/W]','[K/W]','[Ws/K]','[Ws/K]','[W]','[W]','[KW]','[KW]','[KW]','[KW]','[Ω]','[Ω]','[Ω]','[Ω]','[mH]','[Hz]','[Hz]','[Hz]','[1/min]','[1/min]','[1/min]','[1/min]','[1/min]']#,'[1/min]'
        self.Einheiten2=['[V]','[V]','[Vs]','[V/1000rpm]','[A]','[A]','[A]','[A]','[A]','[A]','[A]','[A]','[Nm]','[Nm]','[Nm]','[Nm]','[Nm]','[Nm]','[Nm/A]','[1/min]']

        self.calculated_results_Specific = self.calculated_results1 + self.calculated_results2
        
        self.results_Dict={}
        for d,r in zip(self.gesamtBezeichnungen,self.calculated_results_Specific):
            self.results_Dict[d]=r
        # print(len(self.results_Dict))
        # print(self.results_Dict)

### __________ Assignment of the calculated values in the variables of the Datasheet Template ______ ###: 
                   
        self.Werte = {}
        self.Werte['I_N'] = self.results_Dict['I_N'] 
        self.Werte['N_N'] = self.results_Dict['n_N']
        self.Werte['PP'] = round(self.results_Dict['p']/2)
        self.Werte['K_T'] = self.results_Dict['KT']
        self.Werte['N_MAX'] = self.results_Dict['n_max']
        self.Werte['I_MAX'] = self.results_Dict['I_p']
        self.Werte['I_LIM'] = self.results_Dict['I_p']
        self.Werte['P_N'] = self.results_Dict['P_N']
        self.Werte['M_N'] = self.results_Dict['M_N']
        self.Werte['K_U'] = round(74.05*self.results_Dict['Ku']*(np.sqrt(3)*np.sqrt(2)),0)
        self.Werte['I_0'] = self.results_Dict['I_0'] # Stillstandstrom
        self.Werte['M_0'] = self.results_Dict['M_0'] # Stillstandmoment
        self.Werte['I_K'] = self.results_Dict['I_k'] # Kurzschlussstrom
        self.Werte['I_POLL'] = self.results_Dict['I_p']/4 # Strom Pollageidentif.
        self.Werte['N_FS'] = round(self.n_FS,1) #self.results_Dict['n_FS']    #absprechen, bei HT = Daten!C77; bei HS und UHT = Daten!C75. #C75=WURZEL((220^2)/((Daten!C85/2,45)^2+(Daten!F15/2*(Daten!G62/1000)*Daten!G44*0,25)^2))*60/2/PI()
        self.Werte['R_20'] = self.results_Dict['R1_20C°']
        self.Werte['L_STRANG'] = self.results_Dict['L']/2
        self.Werte['U_N'] = self.results_Dict['U_N']
        self.Werte['F_N'] = round(self.results_Dict['f'],0)     #RUNDEN(Daten!C79/60*Daten!F15/2;0)
        self.PWM=round(((self.n_N_max/60)*(self.Pole/2))*6/1000,2)
        if self.PWM<=4:
            self.Werte['PWM']=4   #max.Ausgangsfrequenz = N_MAX*PP/60; benötigte PWM = max.Ausgangsfrequenz * 6/1000(weil kHz); =WENN(benötigte PWM<=4;4;) + WENN(UND(benötigte PWM>4;benötigte PWM<=8);8;) + WENN(UND(benötigte PWM>8;benötigte PWM<=12);12;) + WENN(UND(benötigte PWM>12;benötigte PWM<=16);16;0)
        elif 4<self.PWM<=8:
            self.Werte['PWM']=8
        elif 8<self.PWM<=12:
            self.Werte['PWM']=12
        elif 12<self.PWM<=16:
            self.Werte['PWM']=16
        else:
            self.Werte['PWM']=round(((self.n_N_max/60)*(self.Pole/2))*6/1000,2)
            
        self.PWM_H=round(((self.n_N_max/60)*(self.Pole/2))*5,2)
        if self.PWM_H <=5000:
            self.Werte['PWM_H']=5000
        elif 5000< self.PWM_H <=6666:
            self.Werte['PWM_H']=6666
        elif  6666 < self.PWM_H <=8000:
            self.Werte['PWM_H']=8000
        elif 8000<self.PWM_H<=10000:
            self.Werte['PWM_H']=10000
        else:
            self.Werte['PWM_H']=round(((self.n_N_max/60)*(self.Pole/2))*5,2)
        
        self.Werte['OU_PROTEC'] =round(11.695*(820/self.results_Dict['KT']),0) #WENN((11,695*(820/K_T))<N_MAX;"benötigt ab "&RUNDEN(11,695*(820/K_T);0)&" U/min";"---")
        self.Werte['M_MAX'] = self.results_Dict['M_p']
        self.Werte['NMAX_VPM'] = round(11.695*(820/self.results_Dict['KT']),0)  #RUNDEN(11,695*(820/Daten!C82);0)
        #self.Werte['n_lim']= round(11.695*(820/self.results_Dict['n_N']),0)
        self.Werte['s1ac_p'] = round((self.results_Dict['n_N']*2*np.pi/60)*self.results_Dict['M_s1ac']/1000,1)
        self.Werte['s1wc_p'] = round(self.results_Dict['P_N'],1)
        self.Werte['s6wc_p'] = self.results_Dict['P_S6']
        self.Werte['pl_n'] = self.results_Dict['Pv_ges']   #GANZZAHL(0,1*Daten!C57)*10
        self.Werte['s1ac_i'] = self.results_Dict['I_s1ac'] #Daten!$P$39
        self.Werte['s1wc_i'] = self.results_Dict['I_s1wc'] #Daten!G44
        self.Werte['s6wc_i'] = self.results_Dict['I_s6wc']
        self.Werte['s1ac_m'] = self.results_Dict['M_s1ac'] #GANZZAHL(Daten!$P$40)
        self.Werte['s1wc_m'] = self.results_Dict['M_s1wc'] #GANZZAHL(Daten!C42)
        self.Werte['s6wc_m'] = self.results_Dict['M_s6wc']
        #self.Werte['PWM_apl'] = round(((int(self.DrehzahlbegrenzungANW)*(self.Pole/2))/60)*6/1000,3) #max.Ausgangsfrequenz = n_lim_apl*PP/60; benötigte PWM = max.Ausgangsfrequenz * 6/1000(weil kHz); =WENN(benötigte PWM<=4;4;) + WENN(UND(benötigte PWM>4;benötigte PWM<=8);8;) + WENN(UND(benötigte PWM>8;benötigte PWM<=12);12;) + WENN(UND(benötigte PWM>12;benötigte PWM<=16);16;0)

        # Heidenhain-Fragebogen
        self.Werte['M_0_H'] = self.results_Dict['M_N'] # H, entspricht M_N, Stillstands-Dauerdrehmoment
        self.Werte['I_0_H'] = self.results_Dict['I_0'] # H, entspricht I_0, Stillstandsstrom
        self.Werte['P_N_H'] = self.results_Dict['P_N']*1000 # H   Nennleistung (bei Nenndrehzahl)
        self.Werte['U_0_H'] = round(self.results_Dict['Ku']*np.sqrt(2)*np.sqrt(3)*0.074*self.results_Dict['n_N'],1) # H   Leerlaufspannung (bei Nenndrehzahl und Nennfluss) (Bitte die Effektivspannung zwischen zwei Phasen/Klemmen eintragen)
        self.Werte['K_T_H'] = self.results_Dict['KT'] # H
        self.Werte['R_20_H'] = 2*self.results_Dict['R1_20C°'] # H    Wicklungswiderstand bei 20 °C(Bitte den Messwert zwischen zwei Phasen/Klemmen eintragen)
        self.Werte['L_MOT_H'] = self.results_Dict['L'] # H,F,R    Induktivität der Motorwicklung (Bitte den Messwert zwischen zwei Phasen eintragen)
        self.Werte['Rth1_H'] = self.results_Dict['Rth1'] # H
        self.Werte['Rth2_H'] = self.results_Dict['Rth2'] # H
        self.Werte['Cth1_H'] = self.results_Dict['Cth1'] # H
        self.Werte['Cth2_H'] = self.results_Dict['Cth2'] # H
        condition1=(850*self.results_Dict['n_N'])/(self.Werte['U_0_H']*np.sqrt(2))
        if condition1< self.n_N_max:
            self.Werte['OU_PROTEC_H'] = round(condition1,0) # H
        else:
            self.Werte['OU_PROTEC_H']='--'
        
        # Fanuc-Fragebogen
        self.Werte['R_20_F'] = self.results_Dict['R1_20C°'] # F,R    Ankerwiderstand (L-L) 
        self.Werte['I_p_F'] = round(self.results_Dict['I_p']*1.2,1) # F

        return (self.Werte)
    # do the all calculation    
    
    def check_Datenblatt_inputs(self):
        self.checkTemplate()
        self.checkMotorTyp()
        self.checkMotorID()
        self.checkStatorID()
        self.checkRotorID()
        self.checkMotorDBL()
        self.checkMotorAS()
        self.TraegheitsMoment()
        self.checkTypTempSensor()
        self.checkZwischenkreisspannung()
        self.checkVorschaltdrossel()
        self.checkDrehzahlbegrenzung()
        self.checAnwendugsDBL()
        self.checkAnwendungsDBL_ID()
        self.checkAnwendungs_AS()
        self.checkTypTempSensor_ANW()
        self.checkEncoder_Strichzahl()
        self.checkEncoder_Schnitstelle()
        self.checkEncoder_Auswertrichtung()
        self.checkEncoder_Hersteller()
        self.checkEncoder_Bezeichnung()
        self.checkDrehzahlbegrenzung_ANW()
        self.checkVorschaltdrossel_ANW()
        Motor=self.Motor_Name
        motor_only_number=re.sub('[^0-9]','',Motor)  # this will delete the characters from the motor name. 360UHX==>360
        self.ui.Motortyp_lineedit.setText('RMHX'+f'{motor_only_number}'+'-'+f'{int(self.Motor_Laenge)}'+'-'+f'{int(self.Vv)}'+'TM')
        
        self.MotorTyp= self.ui.Motortyp_lineedit.text()
        self.fieldNames1 = ["Motortyp", 
                  "Motor-ID 166-", 
                  "Stator-ID 166-",
                  "Rotor-ID 166-", 
                  "Motor-DBL-Version", 
                  "Motor-AS", 
                  "Trägheitsmoment [kgm²]", 
                  "Typ Temperatursensor", 
                  "Zwischenkreisspannung", 
                  "Vorschaltdrossel [mH]",
                  "Drehzahlbegrenzung"]

        self.fieldValues1 = [self.MotorTyp,
        self.MotorID,
        self.StatorID,
        self.RotorID,
        self.MotorDBL,
        self.MotorAS,
        self.Traegheitsmoment,
        self.TypTempSensor,
        self.Zwischenkreisspannung,
        self.Vorschaltdrossel,
        self.Drehzahlbegrenzung]

        self.basedata = dict(zip(self.fieldNames1, self.fieldValues1))
        #print(self.basedata)

        self.fieldNames2 = ["Anwendungs-DBL-Version", 
                  "Anwendungs-DBL-ID 205-", 
                  "Anwendungs-AS", 
                  "Typ Temperatursensor-Anwendung", 
                  "Encoder-Strichzahl",
                  "Encoder-Schnittstelle",
                  "Encoder-Auswertrichtung",
                  "Encoder-Hersteller",
                  "Encoder-Bezeichnung",
                  "Drehzahlbegrenzung",
                  "Vorschaltdrossel-Anwendung [mH]"]   

        self.fieldValues2=[self.AnwendungDBL_V,
        self.AnwendungDBL_ID,
        self.AnwendungsAS,
        self.TypTempSensor_Anw,
        self.Encoder_Strichzahl,
        self.Encoder_Schnitstelle,
        self.Encoder_Auswertrichtung,
        self.Hersteller,
        self.Encoder_Bezeichnung,
        self.DrehzahlbegrenzungANW,
        self.Vorschaltdrossel_ANW,]

        self.applicationdata = dict(zip(self.fieldNames2, self.fieldValues2))

        filepath=self.ui.filelabel.text()
        try:
            if filepath=='' or bool(filepath)==False or len(filepath)==0:
                self.ui.errorlabel1_11.show()
                self.ui.filelabel.setStyleSheet(self.errorstyle)
                msg=QMessageBox()
                msg.setWindowTitle('Fehlermeldung')
                msg.setText('Sie haben kein TechnoTab am anfang ausgewählt \n und keine Berechnungen durchgeführt!\n\nBitte wählen Sie die TechnoTab bei "Data input" seite aus')
                msg.setStyleSheet('background-color: rgb(61, 64, 69);font: 75 12pt "Siemens Sans";color:rgb(255, 255, 255)')
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowIcon(QtGui.QIcon("Images for Program/logo.ico"))
                msg.exec_()
            else:
                self.ui.errorlabel1_11.hide()
                self.ui.filelabel.setStyleSheet(self.successstyle)
        except ValueError:
            self.ui.errorlabel1_11.show()
            self.ui.filelabel.setStyleSheet(self.errorstyle)

        self.is_there_any_visible_label=[label.isVisible() for label in self.errorlabels]
        try:
            if any(self.is_there_any_visible_label)==True:
                self.ui.creatword.setEnabled(False)
                self.ui.creatword.setStyleSheet(desabledButtonStyleSheet)
                
            else:
                self.ui.creatword.setEnabled(True) 
                self.ui.creatword.setStyleSheet(enabledButtonStyleSheet)
                self.calculation_function(self.Motor_Name,self.Temp_Wicklung,self.Temp_Kühlung,self.SpannungU1,self.Motor_Laenge,self.Vv,self.teta_stern,self.Delta_Teta_stern,self.Puls_dauer,self.I)
        except AttributeError:
                self.ui.creatword.setEnabled(False)
                msg1=QMessageBox()
                msg1.setWindowTitle('Fehlermeldung')
                msg1.setText('Sie haben noch keine Berechnungen durchgeführt!\n\nBitte gehen Sie wieder auf die erste seite und wählen Sie die TechnoTab aus\nund starten sie die berechnungen mit dem click auf den "Submit"\n')
                msg1.setStyleSheet('background-color: rgb(61, 64, 69);font: 75 12pt "Siemens Sans";color:rgb(255, 255, 255)')
                msg1.setIcon(QMessageBox.Warning)
                msg1.setWindowIcon(QtGui.QIcon("Images for Program/logo.ico"))
                msg1.exec_()


    """ mini calculation function is a light copy of the main calculation function for the BasisDatenTabelle calculation,
        we are not using the main calculation function because it's taking longer to calculat one cycle,
        because it's should pass through the calculation steps for the diagramms to get more data,
        and those data are not importent and not needed for the BasisDatenTabelle """

    def mini_calculation_function(self,motorName,temWicklung,tempKühlung,spannung,motorLaenge,verschaltung,tetaStern,DeltaTetaStern,pulsDauer,strom):
        self.MotorName=motorName
        self.Spannung=spannung
        self.MotorLaenge=motorLaenge
        self.TM=verschaltung
        self.strom=strom  
        self.Delta_Teta_gesamt =  int(temWicklung) - int(tempKühlung)
        self.Delta_Teta_gesamt_stern=float(tetaStern-20)

    ## kaltwiderstand R1 bei 20C° _________________ :
        def R1_20C(data, motor_modell,X,VV):
            R1_20C=data[motor_modell]
            return ((R1_20C.loc["R_Strang"]*(1-R1_20C.loc["Want"])*(X/100))+R1_20C.loc["R_Strang"]*R1_20C.loc["Want"])/VV**2
        # if R1_20C(self.Daten,motorName,motorLaenge,verschaltung)<0:
        #     self.R1_20C= -R1_20C(self.Daten,motorName,motorLaenge,verschaltung)
        # else:
        #     self.R1_20C= R1_20C(self.Daten,motorName,motorLaenge,verschaltung)
        # print('R1_20C°= ', self.R1_20C)
        self.R1_20C= R1_20C(self.Daten,motorName,motorLaenge,verschaltung)
        #print('R1_20C°= ', self.R1_20C)

    ## warmwiderstand R1 bei 100C° ________________ :
        self.R1_100C=self.R1_20C*(1+(0.004*self.Delta_Teta_gesamt))
        #print('R1_100C°= ', self.R1_100C)

    ## Gesamtwiderstand R_th_ges gekühlt __________________ :
        def R_th_ges(data, motor_modell,X):
            R_th_ges=data[motor_modell]
            return R_th_ges.loc['Rth']*(100/X)
        self.R_th_ges=R_th_ges(self.Daten,motorName,motorLaenge)
        #print('Rth_ges= ', self.R_th_ges)

    ## BemessungsStrom I_N _______ auch Strom-ungekühlt :
## _____ BemessungsStrom _____ ## :
        def BemessungsStrom(R1_warm):                       
            if bool(self.strom)==True and self.ui.Strom_checkBox.isChecked():
                return float(round(strom,1))
            else:
                return np.sqrt(self.Delta_Teta_gesamt/(3*self.R_th_ges*R1_warm))
        self.IN = round(BemessungsStrom(self.R1_100C),2) #latexdatei seite:11, Formel:25
        #print('I_N= ',round(self.IN,1))

    ## SpannungsKonstante ____ ##:
        def Spannungskonstante(data, motor_modell,VV,X): # Berechnung des Spannungskonstante
            Ku = data[motor_modell]
            return 2*0.93*(Ku.loc['Flussscheiteltwert']/1000)*Ku.loc['N1_Innenspule']*(Ku.loc['STK_Magnet']/2)*((Ku.loc['Nutflaeche_Anzahl']/3/2/VV))*(Ku.loc['KorrekturFaktor']/np.sqrt(2))*(X/100)
        self.K_u = Spannungskonstante(self.Daten,motorName,verschaltung,motorLaenge)
        #print('Ku= ', self.K_u)

    ## strang indugtivität Ld_strg ______________ :
        def Induktivitaet_Ld(data, motor_modell,VV,X ): 
            Ld=data[motor_modell]
            #return (Ld.loc['L_ug']/VV**2)*(X/100)
            return ((Ld.loc['L_ug']/1000)*Ld.loc['N1_Innenspule']**2*(Ld.loc['Nutflaeche_Anzahl']/(2*3)))/VV**2*(X/100)
        self.Ld_strg=Induktivitaet_Ld(self.Daten,motorName,verschaltung,motorLaenge)
        #print('L_d_strg= ', self.Ld_strg)

    ## Motorinduktivität L ____________ :
        self.L=self.Ld_strg*2
        #print('L [mH] = ',self.L)

    ## widerstand für Peakstrom bei wiklunkgstemperatur von 80C° ____________ :
        self.R_str_80C=self.R1_20C*(1+(0.004*(self.Delta_Teta_gesamt_stern+0.8*DeltaTetaStern)))
        #print('R_80C° = ',self.R_str_80C)

        self.List_ohne_spule2=['200UHX','240UHX','310UHX','360UHX','410UHX','564HX','564UHX']
    ## Berechnung der Kupfermaße pro Spule m_cu/spule _____________ :
        def Cu_Maesse_Spule(data, motor_modell,X):
            Cu_m_Spule=data[motor_modell]
            # __ Brechnung der Kupfervolumen Wickelköpfe (V_cu) und Aktivkupfer (V_aktiv_cu) bei 100 mm +1mm Bauhöhe:
            if motorName in self.List_ohne_spule2:
                V_cu=Cu_m_Spule.loc['V_Wickelkopf_Spule1']*Cu_m_Spule.loc['Nutflaeche_Anzahl']/10**6
                V_aktiv_cu=((Cu_m_Spule.loc['A_cu']*X)/10**6)*2*(Cu_m_Spule.loc['N1_Innenspule']+1)*Cu_m_Spule.loc['Nutflaeche_Anzahl']
            else:
                V_cu=((Cu_m_Spule.loc['V_Wickelkopf_Spule1']*Cu_m_Spule.loc['Nutflaeche_Anzahl']/2)+(Cu_m_Spule.loc['V_Wickelkopf_Spule2']*Cu_m_Spule.loc['Nutflaeche_Anzahl']/2))/10**6
                V_aktiv_cu=((Cu_m_Spule.loc['A_cu']*X)/10**6)*2*Cu_m_Spule.loc['N1_Innenspule']*Cu_m_Spule.loc['Nutflaeche_Anzahl']
            
            # __ Berechnung der cu gewicht wickelköpfe (m_wickelkopf) und cu gewicht nut (m_cu_Nut)
            m_wickelkopf=Cu_m_Spule.loc['rho_cu']*V_cu
            m_cu_Nut=Cu_m_Spule.loc['rho_cu']*V_aktiv_cu
            # __ cu gewicht ges. ___:
            m_cu=m_wickelkopf+m_cu_Nut
            return m_cu/Cu_m_Spule.loc['Nutflaeche_Anzahl']
        self.m_cu_pro_spule=round(Cu_Maesse_Spule(self.Daten,motorName,motorLaenge),2)
        print('m_cu_spule= ',self.m_cu_pro_spule)
        
    ## Peakstrom I_p ________________ :
        def Peakstrom(data,motor_modell,Deltatetastern,t_p):
            I_p=data[motor_modell]
            return np.sqrt((383*Deltatetastern*((I_p.loc['Nutflaeche_Anzahl']/3)*((self.m_cu_pro_spule)/1000)))/(self.R_str_80C*t_p))
        self.Ip=round(Peakstrom(self.Daten,motorName,DeltaTetaStern,pulsDauer),2)

    ## Drehzahl formel _______________ :
        pp=self.Daten[motorName].loc['STK_Magnet']/2
        self.p=(2*self.K_u*self.R1_20C*self.IN)/(self.K_u**2+(pp*(self.Ld_strg/1000)*self.IN)**2)
        #print('p=',self.p)
        self.q=((self.R1_20C*self.IN)**2-230**2)/(self.K_u**2+(pp*(self.Ld_strg/1000)*self.IN)**2)
        #print('q=',self.q)
        self.lösung1=(-self.p/2)+np.sqrt((self.p/2)**2-self.q)
        self.lösung2=(-self.p/2)-np.sqrt((self.p/2)**2-self.q)
        if self.lösung1 > 0 and self.lösung2 <= 0 or self.lösung1>=self.lösung2 :
            self.nN=(self.lösung1/(2*np.pi))*60
        elif self.lösung2 > 0 and self.lösung1 <= 0 or self.lösung2>=self.lösung1:
            self.nN=(self.lösung2/(2*np.pi))*60    
    
## ____ Drehzahl-Maximal ____ ## :
        def Drehzahl_max(Ku):
            return ((self.Spannung/np.sqrt(3))/(Ku*2*np.pi))*60
        self.n_N_max = Drehzahl_max(self.K_u)
        #print('n_max=',self.n_N_max)
        self.DrehzahlbegrenzungWennLeer=self.n_N_max

## ____ Drehmoment _____ ## :
        def Drehmoment(data,motor_modell,Strom):
            M_abs = data[motor_modell].loc['Mabs']*(int(motorLaenge)/100)
            M_lin =(data[motor_modell].loc['Mlin'])*(int(motorLaenge)/100)*(np.sqrt(2)*(data[motor_modell].loc['N1_Innenspule'])/verschaltung) 
            M_quad = data[motor_modell].loc['Mquad']*(int(motorLaenge)/100)*(np.sqrt(2)*(data[motor_modell].loc['N1_Innenspule'])/verschaltung)**2

            return M_quad*(Strom)**2 + M_lin*Strom + M_abs

        # Bemessungs_Drehmoment
        self.M_N= Drehmoment(self.Daten,motorName,self.IN)

## ____ Stillstand_Drehmoment ____ ## :
        self.M_n0=self.M_N*self.Daten[motorName].loc['F_0']   # aus TD

## ____ Stillstand_Strom _____ ## :
        self.I_n0= round(self.Daten[motorName].loc['F_0']*self.IN,2)  # aus TD 

## ____ Peak_moment ____ ## :
        self.M_p=Drehmoment(self.Daten,motorName,self.Ip)

#_______________________________________________________________________________
    def reverceCalculation(self):
        self.check_calculations_inputs()
        self.checkStrom()
        self.I_new=round(float(self.stromWert),1)
        self.All_calculation()

    def All_calculation(self):
        
        self.results=Calculation_window()
        if bool(self.I)==True and self.I!='' and self.ui.Strom_checkBox.isChecked():
            self.calculation_function(self.Motor_Name,self.Temp_Wicklung,self.Temp_Kühlung,self.SpannungU1,self.Motor_Laenge,self.Vv,self.teta_stern,self.Delta_Teta_stern,self.Puls_dauer,self.I_new)
        else:
            self.calculation_function(self.Motor_Name,self.Temp_Wicklung,self.Temp_Kühlung,self.SpannungU1,self.Motor_Laenge,self.Vv,self.teta_stern,self.Delta_Teta_stern,self.Puls_dauer,self.I)
        self.results.results_table_1.setRowCount(len(self.calculated_results1))
        self.results.results_table_2.setRowCount(len(self.calculated_results2))
        row_index1=0
        header1=self.results.results_table_1.horizontalHeader()
        header1.setSectionResizeMode(0,QHeaderView.ResizeMode.ResizeToContents)
        header1.setSectionResizeMode(1,QHeaderView.ResizeMode.ResizeToContents)
        self.results.results_table_1.setColumnWidth(2,100)
        self.results.results_table_1.setColumnWidth(3,100)
        for data1,kurzbezeichnung1,einheit1,wert1 in zip(self.list1,self.Kurzbezeichnungen1,self.Einheiten1,self.calculated_results1):
            
            d1=QTableWidgetItem(str(data1))
            d1.setFont(QFont('Noto Sans',9))
            #d1.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
            self.results.results_table_1.setItem(row_index1, 0, d1)

            kb1=QTableWidgetItem(str(kurzbezeichnung1))
            kb1.setFont(QFont('Noto Sans',9, QFont.Bold))
            kb1.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
            self.results.results_table_1.setItem(row_index1,1,kb1)

            ein1=QTableWidgetItem(str(einheit1))
            ein1.setFont(QFont('Noto Sans',9, QFont.Bold))
            ein1.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
            self.results.results_table_1.setItem(row_index1,2,ein1)
            r1=QTableWidgetItem(str(wert1))

            r1.setFont(QFont('Noto Sans',9, QFont.Bold))
            r1.setForeground(QBrush(QColor('#087dbd'),QtCore.Qt.SolidPattern))
            r1.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
            self.results.results_table_1.setItem(row_index1, 3, r1)
            row_index1=row_index1+1

        row_index2=0
        header2=self.results.results_table_2.horizontalHeader()
        header2.setSectionResizeMode(0,QHeaderView.ResizeMode.ResizeToContents)
        header2.setSectionResizeMode(1,QHeaderView.ResizeMode.ResizeToContents)
        self.results.results_table_2.setColumnWidth(2,100)
        self.results.results_table_2.setColumnWidth(3,100)
        for data2,kurzbezeichnung2,einheit2,wert2 in zip(self.list2,self.Kurzbezeichnungen2,self.Einheiten2,self.calculated_results2):
            
            d2=QTableWidgetItem(str(data2))
            d2.setFont(QFont('Noto Sans',9))
            #d2.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
            self.results.results_table_2.setItem(row_index2, 0, d2)

            kb2=QTableWidgetItem(str(kurzbezeichnung2))
            kb2.setFont(QFont('Noto Sans',9, QFont.Bold))
            kb2.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
            self.results.results_table_2.setItem(row_index2,1,kb2)

            ein2=QTableWidgetItem(str(einheit2))
            ein2.setFont(QFont('Noto Sans',9, QFont.Bold))
            ein2.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
            self.results.results_table_2.setItem(row_index2,2,ein2)

            r2=QTableWidgetItem(str(wert2))
            r2.setFont(QFont('Noto Sans',9, QFont.Bold))
            r2.setForeground(QBrush(QColor('#087dbd'),QtCore.Qt.SolidPattern))
            r2.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
            self.results.results_table_2.setItem(row_index2, 3, r2)
            row_index2=row_index2+1
        #self.results.setWindowIcon(QIcon("Images for Program/logo.ico"))

        self.Eingangs_Kurzbezeichnung=['','T_w_max','T_kü_max','U_N','l','Vv','Θ*','ΔΘ*','t_p']
        self.results.results_table_3.setRowCount(len(self.Eingangs_Kurzbezeichnung))
        self.Eingangs_Einheit=['','°C','°C','V','mm','','°C','K','s']
        self.Eingangs_Werte=[self.Motor_Name,self.Temp_Wicklung,self.Temp_Kühlung,self.SpannungU1,self.Motor_Laenge,self.Vv,self.teta_stern,self.Delta_Teta_stern,self.Puls_dauer]

        row_index3=0
        header3=self.results.results_table_2.horizontalHeader()
        header3.setSectionResizeMode(1,QHeaderView.ResizeMode.ResizeToContents)
        self.results.results_table_3.setColumnWidth(2,100)
        self.results.results_table_3.setColumnWidth(3,100)
        for kurzbezeichnung3,einheit3,wert3 in zip(self.Eingangs_Kurzbezeichnung,self.Eingangs_Einheit,self.Eingangs_Werte):
            kb3=QTableWidgetItem(str(kurzbezeichnung3))
            kb3.setFont(QFont('Noto Sans',9, QFont.Bold))
            kb3.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
            self.results.results_table_3.setItem(row_index3,1,kb3)

            ein3=QTableWidgetItem(str(einheit3))
            ein3.setFont(QFont('Noto Sans',9, QFont.Bold))
            ein3.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
            self.results.results_table_3.setItem(row_index3,2,ein3)

            r3=QTableWidgetItem(str(wert3))
            r3.setFont(QFont('Noto Sans',9, QFont.Bold))
            r3.setForeground(QBrush(QColor('#087dbd'),QtCore.Qt.SolidPattern))
            r3.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
            self.results.results_table_3.setItem(row_index3, 3, r3)
            row_index3=row_index3+1
        self.results.show()

    def items_selected(self):
        self.num_ITEMS=[self.item.row() for self.item in self.ui.list_to_calculat.selectedIndexes()]

    # do a specific calculation
    def Specific_calculation(self):
        self.results_Specific=Calculation_window_Specific()
        if bool(self.I)==True and self.I!='' and self.ui.Strom_checkBox.isChecked():
            self.calculation_function(self.Motor_Name,self.Temp_Wicklung,self.Temp_Kühlung,self.SpannungU1,self.Motor_Laenge,self.Vv,self.teta_stern,self.Delta_Teta_stern,self.Puls_dauer,self.I_new)
        else:
            self.calculation_function(self.Motor_Name,self.Temp_Wicklung,self.Temp_Kühlung,self.SpannungU1,self.Motor_Laenge,self.Vv,self.teta_stern,self.Delta_Teta_stern,self.Puls_dauer,self.I)
        
        if self.ui.list_to_calculat.selectedIndexes()==[] or bool(self.ui.list_to_calculat.selectedIndexes())==False:
            self.ui.errorlabel2_1.show()
        else:
            self.ui.errorlabel2_1.hide()
            self.what_to_calculat_Specific = self.list1 + self.list2
            self.Kurzbezeichnungen=self.Kurzbezeichnungen1+self.Kurzbezeichnungen2
            self.Einheiten=self.Einheiten1+self.Einheiten2
            
            self.dataSpecific=[]
            self.Kurzbezeichnung=[]
            self.Einheit=[]
            self.Specific_results=[]
               
            for idx in self.num_ITEMS:
                self.dataSpecific.append(self.what_to_calculat_Specific[idx])
                self.Kurzbezeichnung.append(self.Kurzbezeichnungen[idx])
                self.Einheit.append(self.Einheiten[idx])
                self.Specific_results.append(self.calculated_results_Specific[idx])

            self.results_Specific.results_table_Specific.setRowCount(len(self.Kurzbezeichnung))

            row_index=0
            for data, Kbz, unit, value in zip(self.dataSpecific,self.Kurzbezeichnung,self.Einheit,self.Specific_results):
                
                d=QTableWidgetItem(str(data))
                d.setFont(QFont('Noto Sans',9, QFont.Bold))

                self.results_Specific.results_table_Specific.setItem(row_index, 0, d)

                kb=QTableWidgetItem(str(Kbz))
                kb.setFont(QFont('Noto Sans',9, QFont.Bold))
                kb.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
                self.results_Specific.results_table_Specific.setItem(row_index,1,kb)

                ein=QTableWidgetItem(str(unit))
                ein.setFont(QFont('Noto Sans',9, QFont.Bold))
                ein.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
                self.results_Specific.results_table_Specific.setItem(row_index,2,ein)

                r=QTableWidgetItem(str(value))
                r.setFont(QFont('Noto Sans',9, QFont.Bold))
                r.setForeground(QBrush(QColor('#087dbd'),QtCore.Qt.SolidPattern))
                r.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
                self.results_Specific.results_table_Specific.setItem(row_index, 3, r)
                row_index=row_index+1
            self.results_Specific.show()



    def Add_DatenBlattVorlage(self):
        self.response1,filter=QFileDialog.getOpenFileName(self,"Bitte die Vorlage auswählen",os.getenv('HOME'),'docx files(*.docx)')
        self.ui.vorlagePfadLineEdit.setText(self.response1)
        if self.response1 == None or self.response1=='' or bool(self.response1)==False:
            self.ui.errorlabel_21.show()
            self.ui.vorlagePfadLineEdit.setStyleSheet(self.errorstyle)
            self.ui.generatDataSheetBtn.setEnabled(False)
        else:
            self.ui.errorlabel_21.hide()
            self.ui.vorlagePfadLineEdit.setStyleSheet(self.successstyle)
            self.ui.generatDataSheetBtn.setEnabled(True)

    def Export_Data_Sheet(self):
        Motor=self.Motor_Name
        UHT_or_HT_NUMBERS=re.sub('[0-9]','',Motor) # this will delete the numbers from the motor name. 360UHX==>UHX
        UHT_or_HT_CHARACTERS=re.sub('[^0-9]','',Motor)  # this will delete the characters from the motor name. 360UHX==>360
        
        self.calculation_function(self.Motor_Name,self.Temp_Wicklung,self.Temp_Kühlung,self.SpannungU1,self.Motor_Laenge,self.Vv,self.teta_stern,self.Delta_Teta_stern,self.Puls_dauer,self.I)
        DatenBlattName='RM-'+UHT_or_HT_CHARACTERS+'-'+UHT_or_HT_NUMBERS +'-'+str(int(self.Motor_Laenge))+'-'+str(int(self.Vv))+'TM'
        docx_tpl = DocxTemplate(self.response1)

        self.readyFig1=InlineImage(docx_tpl, self.img_buffer)
        self.readyFig2=InlineImage(docx_tpl, self.img_buffer2)

        context = {
            'UHT':UHT_or_HT_NUMBERS,
            'type' : DatenBlattName,
            'D7': int(self.Daten[self.Motor_Name].loc['D7']),
            'D0': int(self.Daten[self.Motor_Name].loc['D0']),
            'H': int(self.Daten[self.Motor_Name].loc['H']),
            's1wc_i':self.Werte['s1wc_i'],
            's6wc_i': self.Werte['s6wc_i'],
            'P_v': self.results_Dict['Pv_ges'],
            'R_20': self.Werte['R_20'],
            'L_STRANG': self.Werte['L_STRANG'],
            'M_0':self.Werte['M_0'],
            'I_N' : self.Werte['I_N'],
            'N_N' : self.Werte['N_N'],
            'P_N' : self.Werte['P_N'],
            'M_N' : self.Werte['M_N'],
            'V_S1_ph_ph' :self.MotorSpannung_S1,
            'F_S1': self.results_Dict['f'],

            'N_S6':self.results_Dict['n_S6'],
            's6wc_p' : self.Werte['s6wc_p'],
            's6wc_m' : self.Werte['s6wc_m'],
            'V_S6_ph_ph' :self.MotorSpannung_S6 ,
            's6wc_i' : self.Werte['s6wc_i'],
            'F_S6': self.results_Dict['f_p'] ,

            'N_p': self.results_Dict['n_p'],
            'P_p':self.results_Dict['P_p'],
            'M_p': self.results_Dict['M_p'],
            'V_p_ph_ph' :self.MotorSpannung_Puls,
            'I_p':self.results_Dict['I_p'],
            'F_p': self.results_Dict['f_S6'] ,

            'Bild1':self.readyFig1,
            'Bild2':self.readyFig2,
            
            'create_date' : date.today(),

            }

        docx_tpl.render(context)
        self.response1_1,filter=QFileDialog.getSaveFileName(self,"Bitte die Vorlage speichern",DatenBlattName,'docx files(*.docx)')#,os.getenv('HOME')
        if self.response1_1==None or self.response1_1=='' or bool(self.response1_1)==False:
            self.ui.errorlabel_22.show()
        else:
            docx_tpl.save(rf'{self.response1_1}')
            self.ui.errorlabel_22.hide()


    def Add_Template(self):
        self.response,filter=QFileDialog.getOpenFileName(self,"Bitte die Vorlage auswählen",os.getenv('HOME'),'docx files(*.docx)')
        self.ui.filelabel_2.setText(self.response)
        if self.response == None or self.response=='' or bool(self.response)==False:
            self.ui.errorlabel1_35.show()
            self.ui.filelabel_2.setStyleSheet(self.errorstyle)
        else:
            self.ui.errorlabel1_35.hide()
            self.ui.filelabel_2.setStyleSheet(self.successstyle)

    def Add_Excel_Template(self):
        self.response_excel,filter=QFileDialog.getOpenFileName(self,"Bitte die Vorlage auswählen",os.getenv('HOME'),'excel files(*.xlsx)')
        self.ui.filelabel_3.setText(self.response_excel)
        if self.response_excel == None or self.response_excel=='' or bool(self.response_excel)==False:
            self.ui.errorlabel1_41.show()
            self.ui.filelabel_3.setStyleSheet(self.errorstyle)
            self.ui.creatExcel.setEnabled(False)
            self.ui.creatExcel.setStyleSheet(desabledButtonStyleSheet)
        else:
            self.is_there_any_visible_label=[label.isVisible() for label in self.errorlabels]
            if any(self.is_there_any_visible_label)==True:
                self.ui.creatExcel.setEnabled(False)
                self.ui.creatExcel.setStyleSheet(desabledButtonStyleSheet)
            else:
                self.ui.creatExcel.setEnabled(True)
                self.ui.creatExcel.setStyleSheet(enabledButtonStyleSheet)
                self.ui.errorlabel1_41.hide()
                self.ui.filelabel_3.setStyleSheet(self.successstyle)
                print(type(self.response_excel))

    def checkTemplate(self):
        try:
            if self.ui.filelabel_2.text() == None or self.ui.filelabel_2.text()=='' or bool(self.ui.filelabel_2.text())==False:
                self.ui.errorlabel1_35.show()
                self.ui.filelabel_2.setStyleSheet(self.errorstyle)
            else:
                self.ui.errorlabel1_35.hide()
                self.ui.filelabel_2.setStyleSheet(self.successstyle) 
        except ValueError:
            self.ui.errorlabel1_35.show()
            self.ui.filelabel_2.setStyleSheet(self.errorstyle)             

    def checkExcelTemplate(self):
        try:
            if self.ui.filelabel_3.text() == None or self.ui.filelabel_3.text()=='' or bool(self.ui.filelabel_3.text())==False:
                self.ui.errorlabel1_41.show()
                self.ui.filelabel_3.setStyleSheet(self.errorstyle)
                self.ui.creatExcel.setEnabled(False)
                self.ui.creatExcel.setStyleSheet(desabledButtonStyleSheet)
            else:
                self.ui.errorlabel1_41.hide()
                #self.ui.filelabel_2.setStyleSheet(self.successstyle)
                self.ui.creatExcel.setEnabled(True)
                self.ui.filelabel_3.setStyleSheet(self.successstyle)
        except ValueError:
            self.ui.errorlabel1_41.show()
            self.ui.filelabel_3.setStyleSheet(self.errorstyle)
            self.ui.creatExcel.setEnabled(False)
            self.ui.creatExcel.setStyleSheet(desabledButtonStyleSheet)

    # Erzeugung Worddateien
    def CreateWord(self): 
        Motor_Laenge=self.ui.laengeBox.currentText()
        self.checkTemplate()
        docx_tpl = DocxTemplate(self.response)
        print('p1800',self.Werte['PWM'])
        # Zurdnen der self.Werte zu Varablen aus Word-Vorlage
        if self.basedata['Vorschaltdrossel [mH]']=='0':
            l_serie='--'
        else:
            l_serie=self.basedata['Vorschaltdrossel [mH]']

        if self.applicationdata['Vorschaltdrossel-Anwendung [mH]']=='0':
            l_serie_apl='--'
        else:
            l_serie_apl=self.applicationdata['Vorschaltdrossel-Anwendung [mH]']

        #condition0=round(11.695*(820/self.results_Dict['KT']),0)
        condition0=self.Werte['OU_PROTEC']
        if condition0< int(self.DrehzahlbegrenzungANW):
            OU_PROTEC= '--' # H
        else:
            OU_PROTEC=round(condition0,0)

        self.PWM_benoetigt= round(((int(self.DrehzahlbegrenzungANW)*(self.Pole/2))/60)*6/1000,3)
        if self.PWM_benoetigt<=4:
            self.PWM_apl = 4   #max.Ausgangsfrequenz = N_MAX*PP/60; benötigte PWM = max.Ausgangsfrequenz * 6/1000(weil kHz); =WENN(benötigte PWM<=4;4;) + WENN(UND(benötigte PWM>4;benötigte PWM<=8);8;) + WENN(UND(benötigte PWM>8;benötigte PWM<=12);12;) + WENN(UND(benötigte PWM>12;benötigte PWM<=16);16;0)
        elif 4<self.PWM_benoetigt<=8:
            self.PWM_apl=8
        elif 8<self.PWM_benoetigt<=12:
            self.PWM_apl=12
        elif 12<self.PWM_benoetigt<=16:
            self.PWM_apl=16
        else:
            self.PWM_apl=round(((int(self.DrehzahlbegrenzungANW)*(self.Pole/2))/60)*6/1000,3)
        
        tempsens = (self.applicationdata['Typ Temperatursensor-Anwendung']).lower() # Umwandlung des self.Wertes in Kleinschreibung '.lower()' damit Vergleich nicht an Case-sensitive scheitert.

        encInterface = (self.applicationdata['Encoder-Schnittstelle']).lower()
        if '1vss' in encInterface:
            ENC_PRO = 0
            ENC_ADJ = 0
            ENC_ABS = 0
            ENC_LIN = 0
            ENC_REF = 0
        elif 'endat 2.1' or 'endat 2.2' in encInterface:
            ENC_PRO = 1
            ENC_ADJ = 0
            ENC_ABS = 1
            ENC_LIN = 0
            ENC_REF = 1
        elif 'endat 2.2 fs' in encInterface:
            ENC_PRO = 2
            ENC_ADJ = 0
            ENC_ABS = 1
            ENC_LIN = 0
            ENC_REF = 1

        if self.applicationdata['Typ Temperatursensor-Anwendung']=='PT1000':
            TSENSOR_APPLICATION = 1
        else:
            TSENSOR_APPLICATION = 0
        context = {
            'type' : self.basedata['Motortyp'],
            'dbl_ver' : self.basedata['Motor-DBL-Version'],
            'mot_id' : self.basedata['Motor-ID 166-'],
            'stat_id' : self.basedata['Stator-ID 166-'],
            'rot_id' : self.basedata['Rotor-ID 166-'],
            'mot_as' : self.basedata['Motor-AS'],
            'I_N' : self.Werte['I_N'],
            'N_N' : self.Werte['N_N'],
            'PP' : self.Werte['PP'],
            'K_T' : self.Werte['K_T'],
            'N_MAX' : self.Werte['N_MAX'],
            'I_MAX' : self.Werte['I_MAX'],
            'I_LIM' : self.Werte['I_LIM'],
            'J' : self.basedata['Trägheitsmoment [kgm²]'],
            'P_N' : self.Werte['P_N'],
            'M_N' : self.Werte['M_N'],
            'K_U' : self.Werte['K_U'],
            'I_0' : self.Werte['I_0'],
            'M_0' : self.Werte['M_0'],
            'I_K' : self.Werte['I_K'],
            'I_POLL' : self.Werte['I_POLL'],
            'N_FS' : self.Werte['N_FS'],
            'R_20' : self.Werte['R_20'],
            'L_STRANG' : self.Werte['L_STRANG'],
            'U_N' : self.Werte['U_N'],
            'F_N' : self.Werte['F_N'],
            'PWM' : self.Werte['PWM'],
            'l_series' : l_serie,
            't_sens' : self.basedata['Typ Temperatursensor'],
            'OU_PROTEC' : OU_PROTEC,
            'u_zk' : self.basedata['Zwischenkreisspannung'],
            'M_MAX' : self.Werte['M_MAX'],
            'n_lim' : self.basedata['Drehzahlbegrenzung'],# wird das verwendet?
            'NMAX_VPM' : self.Werte['NMAX_VPM'],

            's1ac_p' : self.Werte['s1ac_p'],
            's1wc_p' : self.Werte['s1wc_p'],
            's6wc_p' : self.Werte['s6wc_p'],
            'pl_n' : self.Werte['pl_n'],
            's1ac_i' : self.Werte['s1ac_i'],
            's1wc_i' : self.Werte['s1wc_i'],
            's6wc_i' : self.Werte['s6wc_i'],
            's1ac_m' : self.Werte['s1ac_m'],
            's1wc_m' : self.Werte['s1wc_m'],
            's6wc_m' : self.Werte['s6wc_m'],

            'dbl_id' : self.applicationdata['Anwendungs-DBL-ID 205-'],
            'as_apl' : self.applicationdata['Anwendungs-AS'],
            't_sens_apl' : self.applicationdata['Typ Temperatursensor-Anwendung'],
            'n_lim_apl' : self.applicationdata['Drehzahlbegrenzung'],
            'count_nr_apl' : self.applicationdata['Encoder-Strichzahl'],
            'enc_type_apl_H' : self.applicationdata['Encoder-Schnittstelle'],
            'count_dir_apl' : self.applicationdata['Encoder-Auswertrichtung'],
            'enc_manuf_apl' : self.applicationdata['Encoder-Hersteller'],
            'enc_name_apl' : self.applicationdata['Encoder-Bezeichnung'],
            'PWM_apl' : self.PWM_apl,
            'l_series_apl' : l_serie_apl,

            # Heidenhain: ___________
            'M_0_H' : self.Werte['M_0'],
            'I_0_H' : self.Werte['I_0_H'],
            #'T_MAX' : self.Werte['M_MAX'],
            'P_N_H' : self.Werte['P_N_H'],
            'U_0_H' : self.Werte['U_0_H'],
            'K_T_H' : self.Werte['K_T_H'],
            'R_20_H' : self.Werte['R_20_H'],
            'L_MOT_H' : self.Werte['L_MOT_H'],
            'Rth1_H' : self.Werte['Rth1_H'],
            'Rth2_H' : self.Werte['Rth2_H'],
            'T_TH1' : self.Werte['Cth1_H']*self.Werte['Rth1_H'],
            'T_TH2' : self.Werte['Cth2_H']*self.Werte['Rth2_H'],
            'Cth1_H' : self.Werte['Cth1_H'],
            'Cth2_H' : self.Werte['Cth2_H'],
            't_sens_apl_HH': TSENSOR_APPLICATION, 
            'F_PWM' : self.Werte['PWM_H'],
            'OU_PROTEC_H' : self.Werte['OU_PROTEC_H'],
            'ENC_PRO' : ENC_PRO,
            'ENC_ADJ' : ENC_ADJ,
            'ENC_ABS' : ENC_ABS,
            'ENC_LIN' : ENC_LIN,
            'ENC_REF' : ENC_REF,

            # Fanuc-Fragebogen
            'R_20_F' : self.Werte['R_20_F'],
            'I_p_F' : self.Werte['I_p_F'],

            'create_date' : date.today(),
        }

        docx_tpl.render(context)
        self.response2,filter=QFileDialog.getSaveFileName(self,"Bitte die Vorlage speichern",'205-'+self.applicationdata['Anwendungs-DBL-ID 205-']+'-'+self.basedata['Motortyp'],'docx files(*.docx)')#,os.getenv('HOME')
        if self.response2==None or self.response2=='' or bool(self.response2)==False:
            self.ui.errorlabel1_36.show()
        else:
            docx_tpl.save(rf'{self.response2}')
            self.ui.errorlabel1_36.hide()


    def importOEMdatei(self):
        self.mot3_file, filter = QFileDialog.getOpenFileName(self, "motor_oem.mot3 file auswählen", os.getenv('HOME'), "Mot3 Files (*.mot3)")
        self.ui.filelabel_4.setText(self.mot3_file)

    def exportExcel(self):
        self.checkExcelTemplate()
        # !!!!!!!!!!!!!!!!!! wenn die Werte hier als Text vorliegen, müssen sie in "float" umgewandelt werden!!!!!!!!!!!!!!!!!! #
        wb = load_workbook(filename = self.response_excel)
        self.type='CT-RM'+self.basedata['Motortyp']+ '_TDB' + self.basedata['Motor-ID 166-']
        #ws = wb.active
        #print(self.response_excel)
        ws = wb['SM-Kopie-In']
        ws2=wb['Tab3']
        ws['C3'] = self.type+'-prep'
        ws["C5"] = self.Werte['I_N']
        ws['C6'] = self.Werte['I_0']
        ws['C7'] = self.Werte['I_N']
        ws['C9'] = self.Werte['I_MAX']
        ws['C10'] = self.Werte['U_N']
        ws['C11'] = self.Werte['U_0_H']
        ws['C12'] = self.Werte['N_N']
        ws['C14'] = self.Werte['F_N']
        ws['C15'] = self.Werte['R_20']
        ws['C17'] = self.Werte['L_STRANG']
        ws['C18'] = self.Werte['L_STRANG']
        ws['C19'] = self.Werte['L_STRANG']
        ws['C24'] = self.Werte['N_MAX']
        ws['C28'] = self.Werte['PP']
        ws['C31'] = self.applicationdata['Encoder-Auswertrichtung']
        ws['C33'] = self.Werte['P_N_H']
        ws['C34'] = self.basedata['Trägheitsmoment [kgm²]']
        ws['C36'] = self.basedata['Vorschaltdrossel [mH]']
        ws['C38'] = 0.1 / self.Werte['PP'] * 60
        ws['C39'] = round(self.Werte['Rth1_H'] * self.Werte['Cth1_H'],3)
        ws['C40'] = self.Werte['Rth1_H']
        ws['C41'] = round(self.Werte['Rth2_H'] * self.Werte['Cth2_H'],3)
        ws['C42'] = self.Werte['Rth2_H']
        ws['C43'] = self.Werte['M_0_H']
        ws['C45'] = self.Werte['M_MAX']
        ws['C47'] = self.Werte['PWM_H']
        ws['C52'] = self.applicationdata['Typ Temperatursensor-Anwendung']
        encInterface = (self.applicationdata['Encoder-Schnittstelle']).lower()
        if '1vss' in encInterface:
            ws['C62'] = 0
            ws['C63'] = 0
            ws['C64'] = 0
            ws['C65'] = 0
            ws['C66'] = 0
        elif 'endat 2.1' or 'endat 2.2' in encInterface:
            ws['C62'] = 1
            ws['C63'] = 0
            ws['C64'] = 1
            ws['C65'] = 0
            ws['C66'] = 1
        elif 'endat 2.2 fs' in encInterface:
            ws['C62'] = 2
            ws['C63'] = 0
            ws['C64'] = 1
            ws['C65'] = 0
            ws['C66'] = 1
        
        ws['C67'] = self.applicationdata['Encoder-Strichzahl']
        ws2['C5'] = 'SM'
                                      
        self.response_excel_export,filter=QFileDialog.getSaveFileName(self,"Bitte die Vorlage speichern",'205-'+self.applicationdata['Anwendungs-DBL-ID 205-']+'_'+self.basedata['Motortyp']+ '_TDB' + self.basedata['Motor-ID 166-'],'excel files(*.xlsx)')#,os.getenv('HOME')
        if self.response_excel_export==None or self.response_excel_export=='' or bool(self.response_excel_export)==False:
            self.ui.errorlabel1_42.show()
        else:
            try:

                # Save the workbook
                wb.save(self.response_excel_export)
                
                # Open the Excel file
                excel = win32.gencache.EnsureDispatch('Excel.Application')
                excel.Visible = False
                workbook = excel.Workbooks.Open(self.response_excel_export)

                # Set calculation mode to automatic
                workbook.Application.Calculation = win32.constants.xlCalculationAutomatic
                workbook.Save()

                # Close the Excel file
                workbook.Close()

                # Release resources
                excel.Quit()
                excel = None


                if self.mot3_file:
                    try:
                        # Read the content of the .mot3 file
                        with open(self.mot3_file, 'r') as f:
                            existing_lines = f.readlines()

                        # Extract the words from the second line of the .mot3 file
                        second_line_words = existing_lines[1].split()
                        print(second_line_words)

                        # Calculate the spaces between each pair of adjacent words
                        spaces_between_items = []
                        word_started = False
                        space_count = 0
                        for char in existing_lines[1]:
                            if char != ' ' and not word_started:
                                word_started = True
                            elif char == ' ' and word_started:
                                space_count += 1
                            elif char != ' ' and word_started and space_count > 0:
                                spaces_between_items.append(space_count)
                                space_count = 0
                            elif char != ' ' and word_started and space_count == 0:
                                continue
                        
                        # place to reserve :
                        total_space_list=[]
                        for i, word in zip(spaces_between_items,second_line_words):
                            total_space = i+len(word)
                            total_space_list.append(total_space)
                        print(f"Total spaces list is {total_space_list}")

                        # Read data from row 3 of the Excel sheet
                        df = pd.read_excel(self.response_excel_export, sheet_name='motor.mot3 line', header=None)
                        row_3_data = df.iloc[2].tolist()

                        # Format the row 3 data based on the calculated spaces
                        formatted_row_3_data = []
                        for i, word in enumerate(row_3_data):
                            spaceToAdd=total_space_list[i]-len(str(word))
                            formatted_word = str(word) + (' ' * spaceToAdd)
                            formatted_row_3_data.append(formatted_word)
                            # formatted_word = str(word).ljust(total_space_list[i]-len(str(word)))
                            # formatted_row_3_data.append(formatted_word)

                        # Insert the formatted row 3 data before line 3 of the .mot3 file
                        formatted_row_3_line = ''.join(formatted_row_3_data)
                        existing_lines.insert(2, formatted_row_3_line + '\n')

                        # Write the modified content back to the .mot3 file
                        with open(self.mot3_file, 'w') as f:
                            f.writelines(existing_lines)

                        # Inform the user about the successful update
                        QMessageBox.information(self, "Success", "Data from the third row of the Excel sheet has been added to the .mot3 file.")

                    except Exception as e:
                        QMessageBox.warning(self, "Error", f"An error occurred: {str(e)}")

                print(f"Data from rows 2 and 3 of 'motor.mot3 line' sheet of the Excel file has been exported to {self.mot3_file}.")
                self.ui.errorlabel1_42.hide() 

                # Create a QMessageBox
                msg_box = CustomMessageBox()
                msg_box.setWindowTitle("Files Saved")
                msg_box.setText("Files saved successfully.")
                msg_box.setInformativeText("Do you want to open them?")
                msg_box.setStyleSheet('background-color: rgb(61, 64, 69);font: 75 12pt "Siemens Sans";color:rgb(255, 255, 255)')
                msg_box.setIcon(QMessageBox.Information)

                # Add buttons for opening the files
                open_excel_button = msg_box.addButton("Open Excel", QMessageBox.AcceptRole)
                open_text_button = msg_box.addButton("Open Text", QMessageBox.AcceptRole)
                cancel_button = msg_box.addButton("Cancel", QMessageBox.RejectRole)

                # Connect the button clicks to lambda functions to open the files without closing the QMessageBox
                open_excel_button.clicked.connect(lambda: self.open_file_and_keep_box_open(self.response_excel_export, msg_box))
                open_text_button.clicked.connect(lambda: self.open_file_and_keep_box_open(self.mot3_file, msg_box))

                # Execute the QMessageBox
                msg_box.exec_()
            except Exception as e:
                print("An error occurred:", str(e))
        #wb.save(('205-'+self.applicationdata['Anwendungs-DBL-ID 205-']+'_'+self.basedata['Motortyp']+ '_TDB' + self.basedata['Motor-ID 166-']+'.xlsx'))

    # Function to open a file and keep the QMessageBox open
    def open_file_and_keep_box_open(self, file_path, msg_box):
        QDesktopServices.openUrl(QUrl.fromLocalFile(file_path))
        msg_box.activateWindow() 

    def revercedCalculation(self):
        self.I=self.ui.Strom_lineEdit.text()
        def KupferVerlust(I,R1_warm):
            return 3*R1_warm*I**2
        #self.P_cu_new = KupferVerlust(self.I, self.R1_100C)

    def monoMotor_multiVariation(self):
                       
        #self.monoMotor_multiVariation()
        self.blueFill = PatternFill('solid',fgColor='8EA9DB')
        self.darkGreyFill=PatternFill('solid',fgColor='D0CECE')
        self.grayFill = PatternFill('solid',fgColor='F2F2F2')
        self.greenFill = PatternFill('solid',fgColor='E2EFDA')
        self.goldFill=PatternFill('solid',fgColor='FFF2CC') 
        #self.ui.progressBarLabel.show()

        self.WT=float(self.ui.Tw_lineedit_page4.text())
        self.KT=float(self.ui.Tk_lineedit_page4.text())
        laenge=[f'{int(self.Motor_Laenge)}',]
        #spannung=[self.SpannungU1]
        spannung=int(self.SpannungU1)
        self.spannungsliste=[spannung]
        self.spannungStr=[f'{spannung}V']
        self.MotorenList=[self.Motor_Name]
        self.TM_listStr= [f'{self.ui.VV_Box.itemText(i)}TM' for i in range(self.ui.VV_Box.count())]
        self.TM_list= [self.ui.VV_Box.itemText(i) for i in range(self.ui.VV_Box.count())]

        #print(self.TM_list)

        DataToCalculat=['M_N','M_p','M_0','I_N','I_p','I_0','n_N','n_p','n_MAX']
        
        # __________ creating the dataFrame fo rthe results ___________ #
        columns=[np.array(self.spannungStr),np.array(self.MotorenList),np.array(DataToCalculat)]
        columnsList=pd.MultiIndex.from_product(columns,names=['Voltage:','Motor:',''])

        rows=[np.array(laenge),np.array(self.TM_listStr)]
        indexList=pd.MultiIndex.from_product(rows,names=['h [mm]','TM'])

        self.df_Mono=pd.DataFrame(index=indexList, columns=columnsList)

        cycles=len(self.TM_list)
        progress=QProgressDialog('Calculation in progress...','Cancel',0,cycles)
        progress.setWindowTitle('Please wait')
        progress.setWindowIcon(QtGui.QIcon("Images for Program/logo.ico"))
        progress.setFixedSize(300, 100)
        #progress.setCancelButton(None)
        progress.show()
        QApplication.processEvents()

    ## _____________ loop for the calculation fanction ___________ ## 
        for U1 in self.spannungsliste:
            for vv,i in zip(self.TM_list,range(len(self.TM_list))):
                self.TM=float(vv)
                progress.setValue(i)
                #self.calculation_function(self.Motor_Name,self.WT,self.KT,spannung,self.Motor_Laenge,self.TM,80.0,20.0,self.Puls_dauer,strom_basisDB)
                self.calculation_function(self.Motor_Name,self.WT,self.KT,self.SpannungU1,self.Motor_Laenge,self.TM,self.teta_stern,self.Delta_Teta_stern,self.Puls_dauer,self.I)
                DataToCalculat_values=[self.results_Dict['M_N'],self.results_Dict['M_p'],self.results_Dict['M_0'],self.IN,self.Ip,self.I_n0,self.results_Dict['n_N'],self.results_Dict['n_p'],self.results_Dict['n_max'],]
                self.df_Mono.loc[(str(int(self.MotorLaenge)),f'{int(self.TM)}TM'),(f'{U1}V',str(self.MotorName),np.array(DataToCalculat))]=np.array(DataToCalculat_values)
                if progress.wasCanceled():
                    break
                QApplication.processEvents()
            progress.setValue(cycles)
        
        print(self.df_Mono)

    def when_mono_export_clicked(self):
        self.monoMotor_multiVariation()

        def fill_colour(sheet): 
        ### Blue Fill
            
            col = 'C'
            row = 1
            fill_cell = sheet[f'{col}{row}']
            if fill_cell.value is not None:
                fill_cell.fill = self.blueFill
                    
            ### Gray & Brown Fill
            ### Fill header cells
            head_dict = {4: 4+len(self.TM_listStr)}
            for hrow, v in head_dict.items():
                ### Create list of cells in the Header section to fill
                ### Columns C to S for row 3 or 107 (excludes cols J, K & L)
                head_list = [f'{chr(i)}{hrow - 1}' for i in range(ord('C'), ord('K') + 1)]
            #print(hrow)
                #print(v)
                ### Background fill each cell in the head_list
                for hcell in head_list:
                    gray_cell = sheet[hcell]
                    if gray_cell.value is not None:
                        gray_cell.fill = self.grayFill

                ### Add the cells for the 'Motor' row in light  brown
                motorHX = f'C{hrow-2}'
                darkGrey_cell = sheet[motorHX]
                if darkGrey_cell.value is not None:
                    darkGrey_cell.fill = self.darkGreyFill

                ### Fill cells in columns A, B and K, L if used
                for row in sheet.iter_rows(min_col=2, max_col=2, min_row=hrow, max_row=v):
                    for cell in row:
                        if cell.value is not None:
                            cell.fill = self.grayFill
                        col_list = [-1,]
                        for col in col_list:
                            if cell.offset(column=col).value is not None:
                                cell.offset(column=col).fill = self.grayFill

                ### Add borders to the celss of each frame if only if the cell is not empty
                for row in sheet.iter_rows(min_col=0, max_col=(self.df_V.shape[1]+10), min_row=0, max_row=(self.df_V.shape[0]+10)):
                    for cell in row:
                        if cell.value is not None:
                            bd = Side(style='thin', color="000000")
                            cell.border=Border(left=bd, top=bd, right=bd, bottom=bd)

        self.response4,filter=QFileDialog.getSaveFileName(self,"Bitte das Datenblatt speichern",rf'Motortabelle-{self.Motor_Name}','Excel files(*.xlsx)')#,os.getenv('HOME')
        if self.response4==None or self.response4=='' or bool(self.response4)==False:
            self.ui.errorlabel_20.show()
        else:
            self.ui.errorlabel_20.hide()
            writer= pd.ExcelWriter(rf'{self.response4}', engine='openpyxl')
            for motor in self.MotorenList:
                sheetName=motor
                self.df_V=pd.DataFrame(self.df_Mono[[f'{int(self.SpannungU1)}V']].xs(motor,level='Motor:',axis=1,drop_level=False))
                self.df_V.to_excel(writer,sheet_name=sheetName,startrow=0,startcol=0)
            fill_colour(writer.sheets[sheetName])
            writer.close()

    # ___________ show graphs window
    def showGraphsWindow(self):
        if bool(self.I)==True and self.I!='' and self.ui.Strom_checkBox.isChecked():
            self.calculation_function(self.Motor_Name,self.Temp_Wicklung,self.Temp_Kühlung,self.SpannungU1,self.Motor_Laenge,self.Vv,self.teta_stern,self.Delta_Teta_stern,self.Puls_dauer,self.I_new)
        else:
            self.calculation_function(self.Motor_Name,self.Temp_Wicklung,self.Temp_Kühlung,self.SpannungU1,self.Motor_Laenge,self.Vv,self.teta_stern,self.Delta_Teta_stern,self.Puls_dauer,self.I)
       
        Graphs_Window.Graphs_function(self,self.x1,self.x2,self.x2,self.x2,self.x5,self.x_axis,self.x_axis,self.y1,self.y2,self.y3,self.y4,self.y5,self.y6_list,self.y6_list_p)
        #Graphs_Window.Graphs_function(self,self.x1,self.x2,self.x2,self.x2,self.x5,self.x1,self.x1,self.y1,self.y2,self.y3,self.y4,self.y5,self.y6_list,self.y6_list_p)

    def browsefile_1_Page4(self):
        self.file_name1=QFileDialog.getOpenFileName(self,'Open file',os.getenv('HOME'),'csv files(*.csv)')
        file_path1=self.ui.filelabel_page4_1.setText(self.file_name1[0])
        file_path1=self.ui.filelabel_page4_1.text()
        try:
            if file_path1=='' or bool(file_path1)==False or len(file_path1)==0:
                self.ui.errorlabel1_37.show()
                self.ui.filelabel_page4_1.setStyleSheet(self.errorstyle)
                self.ui.exportBasisdaten_btn.setEnabled(False)
            else:
                self.ui.errorlabel1_37.hide()
                self.ui.filelabel_page4_1.setStyleSheet(self.successstyle)
        except ValueError:
            self.ui.errorlabel1_37.show()
            self.ui.filelabel_page4_1.setStyleSheet(self.errorstyle)
            self.ui.exportBasisdaten_btn.setEnabled(False)

    def checkWT_page4(self):
        Temp_Wicklung= self.ui.Tw_lineedit_page4.text()
        try:
            if bool(Temp_Wicklung)== False or Temp_Wicklung is str or Temp_Wicklung=='' :
                self.ui.errorlabel1_38.show()
                self.ui.Tw_lineedit_page4.setStyleSheet(self.errorstyle)
                self.ui.pushButton_next.setEnabled(False)
                self.ui.exportBasisdaten_btn.setEnabled(False)
                
            elif float(Temp_Wicklung) in range(0,500):
                condition2=self.ui.errorlabel1_38.hide()
                self.ui.Tw_lineedit_page4.setStyleSheet(self.successstyle)
                
        except ValueError:
            self.ui.errorlabel1_38.show()
            self.ui.Tw_lineedit_page4.setStyleSheet(self.errorstyle)
            self.ui.pushButton_next.setEnabled(False)
            self.ui.exportBasisdaten_btn.setEnabled(False)

    def checkKT_page4(self):
        Temp_Kühlung=self.ui.Tk_lineedit_page4.text()
        try:
            if bool(Temp_Kühlung)==False or Temp_Kühlung is str or Temp_Kühlung=='':
                self.ui.errorlabel1_39.show()
                self.ui.Tk_lineedit_page4.setStyleSheet(self.errorstyle)
                self.ui.pushButton_next.setEnabled(False)
                self.ui.exportBasisdaten_btn.setEnabled(False)

            elif float(Temp_Kühlung) in range(0,500):
                condition3=self.ui.errorlabel1_39.hide()
                self.ui.Tk_lineedit_page4.setStyleSheet(self.successstyle)
                    
        except ValueError:
            self.ui.errorlabel1_39.show()
            self.ui.Tk_lineedit_page4.setStyleSheet(self.errorstyle)
            self.ui.pushButton_next.setEnabled(False)
            self.ui.exportBasisdaten_btn.setEnabled(False)

    def showProgressLabel(self):
        if self.ui.filelabel_page4_1.text() == None or self.ui.filelabel_page4_1.text()=='' or bool(self.ui.filelabel_page4_1.text())==False:
            self.ui.errorlabel1_37.show()
            self.ui.filelabel_page4_1.setStyleSheet(self.errorstyle)
            self.ui.progressBarLabel.hide()
        else:
            self.ui.errorlabel1_37.hide()
            self.ui.filelabel_page4_1.setStyleSheet(self.successstyle) 
            self.ui.progressBarLabel.show()
   
    def ExecutelongTaskFunction(self,progress_bar):
        #self.ui.progressBarLabel.show()

        self.checkWT_page4()
        self.checkKT_page4()
        strom_basisDB=self.ui.Strom_lineEdit.text()
        if self.ui.filelabel_page4_1.text() == None or self.ui.filelabel_page4_1.text()=='' or bool(self.ui.filelabel_page4_1.text())==False:
            self.ui.errorlabel1_37.show()
            self.ui.filelabel_page4_1.setStyleSheet(self.errorstyle)
            
        else:
            self.ui.errorlabel1_37.hide()
            self.ui.filelabel_page4_1.setStyleSheet(self.successstyle)         
            # reading the csv_file :
            self.Daten = pd.read_csv(self.file_name1[0],delimiter=';', index_col='Parameter',decimal=',').fillna('')
            self.Daten[["200HX", "200UHX", "240HX", "240UHX", "310HX", "310UHX", "360UHX", "410HX", "410UHX", "564HX"]] = self.Daten[["200HX", "200UHX", "240HX", "240UHX", "310HX", "310UHX", "360UHX", "410HX", "410UHX", "564HX"]].apply(pd.to_numeric).fillna('')
            print(type(self.Daten))
            self.MotorenListe=["200HX", "200UHX", "240HX", "310HX", "360UHX","564HX"]
            self.WT=float(self.ui.Tw_lineedit_page4.text())
            self.KT=float(self.ui.Tk_lineedit_page4.text())
            self.SpannungListe=[400,200]
            vv_keys = ["Seriel", "2TM", "3TM", "4TM", "5TM", "6TM", "8TM", "10TM", "12TM"]
            laengeListeDrehzahl=['50','75','100','125','150','175','200','225','250','275','300']
            laengeKeys_pulsdauer=['50mm','75mm','100mm','125mm','150mm','175mm','200mm','225mm','250mm','275mm','300mm']

            LaengeDic={
                '50':'50mm',
                '75':'75mm',
                '100':'100mm',
                '125':'125mm',
                '175':'175mm',
                '150':'150mm',
                '200':'200mm',
                '225':'225mm',
                '250':'250mm',
                '275':'275mm',
                '300':'300mm'
            }

            laenge=['50','75','100','125','150','175','200','225','250','275','300']
            TM=["S", "2TM", "3TM", "4TM", "5TM", "6TM", "8TM", "10TM", "12TM"]
            TM_werte=[1,2,3,4,5,6,8,10,12]

            spannung=['400V','200V']
            self.MotorenList=["200HX", "200UHX", "240HX", "310HX","360UHX","564HX"]
            DataToCalculat=['M_N','M_0','I_N','I_p','I_0','n_N','n_MAX']

            # __________ creating the dataFrame fo r the results ___________ #
            columns=[np.array(spannung),np.array(self.MotorenList),np.array(DataToCalculat)]
            columnsList=pd.MultiIndex.from_product(columns,names=['Voltage:','Motor:',''])

            rows=[np.array(laenge),np.array(TM)]
            indexList=pd.MultiIndex.from_product(rows,names=['h','TM'])

            self.df=pd.DataFrame(index=indexList, columns=columnsList)

    ## _________ Add Formatting to the Excel file with openpyxl and pandas 

        ## _____________ loop for the calculation fanction ___________ ## 
            i=0
            for U1 in self.SpannungListe:
                self.spannung=U1
                for motor in self.MotorenListe:
                    self.MotorName=motor
                    vv_values = self.Daten[self.MotorName][vv_keys]
                    vv_values = vv_values[vv_values!=''].astype(int)
                    vv_values = vv_values.to_string(header=False,index=False)
                    VV_Updats=list(vv_values.split())
            
                    for x in laengeListeDrehzahl: 
                        self.Laenge=float(x)
                        selectedLaenge_pulsdauer=LaengeDic.get(x)
                        try:
                            if selectedLaenge_pulsdauer in laengeKeys_pulsdauer==False or selectedLaenge_pulsdauer==None :
                                try:
                                    self.pulsDauer=3
                                except KeyError:
                                    self.pulsDauer=3
                            else:
                                try:
                                    TD_Pulsdauer=self.Daten[self.MotorName][selectedLaenge_pulsdauer]
                                    if TD_Pulsdauer=='' :
                                        self.pulsDauer=3
                                    else:
                                        self.pulsDauer=TD_Pulsdauer
                                except KeyError:
                                    self.pulsDauer=3
                        except ValueError:
                            self.pulsDauer=3

                        for vv in TM_werte:
                            self.TM=float(vv)
                            self.mini_calculation_function(self.MotorName,self.WT,self.KT,self.spannung,self.Laenge,self.TM,80.0,20.0,self.pulsDauer,strom_basisDB)
                            i+=1
                            DataToCalculat_values=[round(self.M_N),round(self.M_n0),round(self.IN,1),round(self.Ip,1),round(self.I_n0,1),round(self.nN),round(self.n_N_max)]
                            #print('cycles= ',i)
                            
                            if self.TM==1:
                                self.df.loc[(str(int(self.MotorLaenge)),'S'),(f'{int(self.spannung)}V',str(self.MotorName),np.array(DataToCalculat))]=np.array(DataToCalculat_values)

                            else:
                                self.df.loc[(str(int(self.MotorLaenge)),f'{int(self.TM)}TM'),(f'{int(self.spannung)}V',str(self.MotorName),np.array(DataToCalculat))]=np.array(DataToCalculat_values)
                            
                            progress_bar.emit(i)
                            print('cycle= ',i)
                            #self.ui.progressBar.setValue(i)
            #return i
            # if i==572:
            #     self.ui.exportBasisdaten_btn.setEnabled(True)
            #     self.ui.progressBarLabel.setText('Progress successfully completed')
            #     print("Thread Complete")                     

    def updateProgressBar(self,value):
        self.ui.progressBar.setValue(value)
        if value==1188:
            self.ui.exportBasisdaten_btn.setEnabled(True)
            self.ui.progressBarLabel.setText('Progress successfully completed')
 
    # def thread_complete(self):
    #     # if value==572:
    #     #     self.ui.exportBasisdaten_btn.setEnabled(True)
    #     #     self.ui.progressBarLabel.setText('Progress successfully completed')
    #         print("Thread Complete")

    def when_submit_page4(self):

        worker=Worker(self.ExecutelongTaskFunction)
        worker.signals.progress.connect(self.updateProgressBar)
        #worker.signals.finished.connect(self.thread_complete)
        self.threadpool.start(worker)
        #worker.signals.finished.connect(self.thread_complete) 

        # _______________ deviding the dataframes depending on the voltage and the motortypes _________ #

    def when_export_clicked_page4(self):
      
        def fill_colour(sheet):
        ### Blue Fill
            self.blueFill = PatternFill('solid',fgColor='8EA9DB')
            self.darkGreyFill=PatternFill('solid',fgColor='D0CECE')
            self.grayFill = PatternFill('solid',fgColor='F2F2F2')
            self.greenFill = PatternFill('solid',fgColor='E2EFDA')
            self.goldFill=PatternFill('solid',fgColor='FFF2CC') 
            col_400V = ['C','M']
            for col in col_400V:
                row_400V = 1
                fill_cell_400V = sheet[f'{col}{row_400V}']
                if fill_cell_400V.value is not None:
                    fill_cell_400V.fill = self.blueFill

            col_200V = ['C','M']
            for col in col_200V:
                row_200V = 105
                fill_cell_200V = sheet[f'{col}{row_200V}']
                if fill_cell_200V.value is not None:
                    fill_cell_200V.fill = self.goldFill

            ### Gray & Brown Fill
            ### Fill header cells
            head_dict = {4: 103, 108: 207}
            for hrow, v in head_dict.items():
                ### Create list of cells in the Header section to fill
                ### Columns C to S for row 3 or 107 (excludes cols J, K & L)
                head_list = [f'{chr(i)}{hrow - 1}' for i in range(ord('C'), ord('S') + 1)
                            if chr(i) != 'J' and chr(i) != 'K' and chr(i) != 'L']
                #print(hrow)
                #print(v)
                ### Background fill each cell in the head_list
                for hcell in head_list:
                    gray_cell = sheet[hcell]
                    if gray_cell.value is not None:
                        gray_cell.fill = self.grayFill
                                                                      
                ### Add the cells for the 'Motor' row in light  brown
                motorHX = f'C{hrow-2}'
                darkGrey_cell = sheet[motorHX]
                if darkGrey_cell.value is not None:
                    darkGrey_cell.fill = self.darkGreyFill

                motorUHX =f'M{hrow-2}'
                brn_cell = sheet[motorUHX]

                if brn_cell.value is not None:
                    brn_cell.fill = self.greenFill

                ### Fill cells in columns A, B and K, L if used
                for row in sheet.iter_rows(min_col=2, max_col=2, min_row=hrow, max_row=v):
                    for cell in row:
                        if cell.value is not None:
                            cell.fill = self.grayFill
                        col_list = [-1, 9, 10]
                        for col in col_list:
                            if cell.offset(column=col).value is not None:
                                cell.offset(column=col).fill = self.grayFill
    
                ### Add borders to the cells of each frame if only if the cell is not empty
                for row in sheet.iter_rows(min_col=0, max_col=(self.df_400V.shape[1]+5)*2, min_row=0, max_row=(self.df_400V.shape[0]+5)*2):
                    for cell in row:
                        if cell.value is not None:
                            bd = Side(style='thin', color="000000")
                            cell.border=Border(left=bd, top=bd, right=bd, bottom=bd)  
                            
        self.response3,filter=QFileDialog.getSaveFileName(self,"Bitte das Basisdatenblatt speichern",'Motorentabelle','Excel files(*.xlsx)')#,os.getenv('HOME')
        if self.response3==None or self.response3=='' or bool(self.response3)==False:
            self.ui.errorlabel1_40.show()
        else:
            self.ui.errorlabel1_40.hide()
            writer= pd.ExcelWriter(rf'{self.response3}', engine='openpyxl')
            for motor in self.MotorenList:
                if motor=='200HX' or motor=='200UHX':
                    sheetName='RM 200'
                    self.df_400V=pd.DataFrame(self.df[['400V']].xs(motor,level='Motor:',axis=1,drop_level=False))
                    df_200V=pd.DataFrame(self.df[['200V']].xs(motor,level='Motor:',axis=1,drop_level=False))
                    if motor=='200UHX':
                        self.df_400V.to_excel(writer,sheet_name=sheetName,startrow=0,startcol=self.df_400V.shape[1]+3)
                        df_200V.to_excel(writer,sheet_name=sheetName,startrow=self.df_400V.shape[0]+5,startcol=self.df_400V.shape[1]+3)
                    else:
                        self.df_400V.to_excel(writer,sheet_name=sheetName,startrow=0,startcol=0)
                        df_200V.to_excel(writer,sheet_name=sheetName,startrow=self.df_400V.shape[0]+5,startcol=0)
                    fill_colour(writer.sheets[sheetName])

                elif motor=='240HX' or motor=='240UHX':
                    sheetName='RM 240'
                    self.df_400V=pd.DataFrame(self.df[['400V']].xs(motor,level='Motor:',axis=1,drop_level=False))
                    df_200V=pd.DataFrame(self.df[['200V']].xs(motor,level='Motor:',axis=1,drop_level=False))
                    if motor=='240UHX':
                        self.df_400V.to_excel(writer,sheet_name=sheetName,startrow=0,startcol=self.df_400V.shape[1]+3)
                        df_200V.to_excel(writer,sheet_name=sheetName,startrow=self.df_400V.shape[0]+5,startcol=self.df_400V.shape[1]+3)
                    else:
                        self.df_400V.to_excel(writer,sheet_name=sheetName,startrow=0,startcol=0)
                        df_200V.to_excel(writer,sheet_name=sheetName,startrow=self.df_400V.shape[0]+5,startcol=0)
                    fill_colour(writer.sheets[sheetName])

                elif motor=='310HX' or motor=='310UHX':
                    sheetName='RM 310'
                    self.df_400V=pd.DataFrame(self.df[['400V']].xs(motor,level='Motor:',axis=1,drop_level=False))
                    df_200V=pd.DataFrame(self.df[['200V']].xs(motor,level='Motor:',axis=1,drop_level=False))
                    if motor=='310UHX':
                        self.df_400V.to_excel(writer,sheet_name=sheetName,startrow=0,startcol=self.df_400V.shape[1]+3)
                        df_200V.to_excel(writer,sheet_name=sheetName,startrow=self.df_400V.shape[0]+5,startcol=self.df_400V.shape[1]+3)
                    else:
                        self.df_400V.to_excel(writer,sheet_name=sheetName,startrow=0,startcol=0)
                        df_200V.to_excel(writer,sheet_name=sheetName,startrow=self.df_400V.shape[0]+5,startcol=0)
                    fill_colour(writer.sheets[sheetName])

                elif motor=='360HX' or motor=='360UHX':
                    sheetName='RM 360'
                    self.df_400V=pd.DataFrame(self.df[['400V']].xs(motor,level='Motor:',axis=1,drop_level=False))
                    df_200V=pd.DataFrame(self.df[['200V']].xs(motor,level='Motor:',axis=1,drop_level=False))
                    if motor=='360UHX':
                        self.df_400V.to_excel(writer,sheet_name=sheetName,startrow=0,startcol=self.df_400V.shape[1]+3)
                        df_200V.to_excel(writer,sheet_name=sheetName,startrow=self.df_400V.shape[0]+5,startcol=self.df_400V.shape[1]+3)
                    else:
                        self.df_400V.to_excel(writer,sheet_name=sheetName,startrow=0,startcol=0)
                        df_200V.to_excel(writer,sheet_name=sheetName,startrow=self.df_400V.shape[0]+5,startcol=0)
                    fill_colour(writer.sheets[sheetName])

                elif motor=='564HX' or motor=='564UHX':
                    sheetName='RM 564'
                    self.df_400V=pd.DataFrame(self.df[['400V']].xs(motor,level='Motor:',axis=1,drop_level=False))
                    df_200V=pd.DataFrame(self.df[['200V']].xs(motor,level='Motor:',axis=1,drop_level=False))
                    if motor=='564UHX':
                        self.df_400V.to_excel(writer,sheet_name=sheetName,startrow=0,startcol=self.df_400V.shape[1]+3)
                        df_200V.to_excel(writer,sheet_name=sheetName,startrow=self.df_400V.shape[0]+5,startcol=self.df_400V.shape[1]+3)
                    else:
                        self.df_400V.to_excel(writer,sheet_name=sheetName,startrow=0,startcol=0)
                        df_200V.to_excel(writer,sheet_name=sheetName,startrow=self.df_400V.shape[0]+5,startcol=0)
                    fill_colour(writer.sheets[sheetName])
            writer.close()
                
class Graphs_Window(QMainWindow):
    def __init__(self): 
        super().__init__()
        uic.loadUi("ui_windows/Graphs_Window.ui",self)
        self.setWindowTitle("Graphs Window")
        self.setWindowIcon(QtGui.QIcon("Images for Program/logo.ico"))
        self.setGeometry(100, 100, 800, 600)   

    def Graphs_function(self,x1,x2,x3,x4,x5,x6,x7,y1,y2,y3,y4,y5,y6,y7):
        self.xList=[x1,x2,x3,x4,x5,x6]
        self.yList=[y1,y2,y3,y4,y5,y6]
        self.TabsTitle=['Induzierte Spannung','Eisenverlust','Kupferverlust','Bemessungsstrom','Drehmoment/strom']
        self.Lables=["Up(n)","Pvfe(f)","Pvcu(f)","I1(f)","M(I1)",'M_N(n)','M_p(n)']
        self.xLables=['Drehzahl [1/s]',"Frequenz [1/s]","Frequenz [1/s]","Frequenz [1/s]","Strom [A]"]
        self.yLables=["Up-Strang [V]","Pv_fe [W]","Pv_cu [W]","I1 [A]","Drehmoment [Nm]"]

        self.Graphs=Graphs_Window()
        self.Graphs.DrehzahlLineEdit.hide()
        self.Graphs.Drehzahl_label.hide()
        self.Graphs.DrehmomentLineEdit.hide()
        self.Graphs.Drehmoment_label.hide()
        
        for i in range(len(self.TabsTitle)):
            tab=QWidget()
            tab.layout=QVBoxLayout(tab)
            fig, ax = plt.subplots()
            self.Graphs.tabwidget.addTab(tab, f"{self.TabsTitle[i]}")
            ax.set_xlabel(f"{self.xLables[i]}")
            ax.set_ylabel(f"{self.yLables[i]}")
            ax.grid()
            ax.plot(self.xList[i], self.yList[i], label=f"{self.Lables[i]}")
            ax.legend()
            ax.margins(x=0,y=0.01)
            canvas = FigureCanvas(fig)
            toolbar = NavigationToolbar(canvas, tab)
            tab.layout.addWidget(toolbar)
            tab.layout.addWidget(canvas)

        # _______ Diagram für Drehmoment/Drehzahl ________ #
        tabNew=QWidget()
        tabNew.layout=QVBoxLayout(tabNew)
        figNew, axNew = plt.subplots()
        self.Graphs.tabwidget.addTab(tabNew, "Drehmoment/Drehzahl") 
        axNew.plot(x6,y6,x7,y7)
        axNew.set_xlabel('Drehzahl in [1/min]')
        axNew.set_ylabel('Drehmoment in [Nm]')
        axNew.legend(['M_N(n)','M_p(n)'])
        axNew.grid()
        axNew.margins(x=0,y=0.01)
        canvasNew = FigureCanvas(figNew)
        toolbarNew = NavigationToolbar(canvasNew, tabNew)
        tabNew.layout.addWidget(toolbarNew)
        tabNew.layout.addWidget(canvasNew)
        coords=[]

        def mouse_click(event):
            x,y=event.xdata,event.ydata
            self.Graphs.DrehzahlLineEdit.setText(str(round(x,1)))
            self.Graphs.DrehmomentLineEdit.setText(str(round(y,1)))
            #print ('x = %d, y = %d'%(x, y))
            coords.append((x,y))
            axNew.scatter(x, y, marker='.')
            figNew.canvas.draw()
        cid=figNew.canvas.mpl_connect('button_press_event',mouse_click)

        def TabIndexChanged():
            tab=self.Graphs.tabwidget.currentIndex()
            if tab==5:
                self.Graphs.DrehzahlLineEdit.show()
                self.Graphs.Drehzahl_label.show()
                self.Graphs.DrehmomentLineEdit.show()
                self.Graphs.Drehmoment_label.show()
            else:
                self.Graphs.DrehzahlLineEdit.hide()
                self.Graphs.Drehzahl_label.hide()
                self.Graphs.DrehmomentLineEdit.hide()
                self.Graphs.Drehmoment_label.hide()
        self.Graphs.tabwidget.currentChanged.connect(TabIndexChanged)
        self.Graphs.show()    

if __name__=="__main__":
    app=QApplication(sys.argv)
    window=MainWindow()
    window.setWindowIcon(QtGui.QIcon("Images for Program/logo.ico"))
    window.setWindowTitle(('Data Sheet Manager'))
    window.show()
    sys.exit(app.exec_())
